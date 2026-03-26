# Single-cell Market Event Calendar notebook

# Optional configuration
CUSTOM_ECO_UNIVERSE = [
    'US Country', 'EU Country', 'JN Country', 'CN Country', 'KR Country',
    'AU Country', 'NZ Country', 'BZ Country', 'MX Country', 'CA Country',
    'CL Country', 'CO Country',
]

CUSTOM_COMDTY_TICKERS = [
    'CL1 Comdty', 'CO1 Comdty', 'NG1 Comdty', 'GC1 Comdty', 'SI1 Comdty',
    'HG1 Comdty', 'ES1 Comdty', 'RTY1 Comdty', 'NQ1 Comdty', 'Z1 Comdty',
]

# Economic calendar types pulled by the live BQL refresh.
# This mirrors:
# bql.data.calendar(type=['central_banks', 'ECONOMIC_RELEASES'], dates=DATE_RANGE)
CUSTOM_ECO_CALENDAR_TYPES = ['central_banks', 'ECONOMIC_RELEASES']

# Flagged-event alarm defaults.
# The notebook/browser clock is used against the event date and time shown in the app.
ALARM_ENABLED_BY_DEFAULT = True
ALARM_SOUND_ENABLED_BY_DEFAULT = True
ALARM_CHECK_INTERVAL_SECONDS = 5
ALARM_TRIGGER_WINDOW_SECONDS = 15

TOP_N_MARKET_CAP = 250
ANNOTATIONS_FILE = "event_annotations.csv"
LEGACY_STATE_FILE = "event_calendar_state.json"  # Optional migration source
EXPORT_DIR = "event_calendar_exports"
MONTHS_PER_PANE_DEFAULT = 3

# Leave as None to use the embedded holiday matrix from the supplied closure schedule.
CUSTOM_MARKET_HOLIDAYS_TSV = None

# Override market labels here if your header meanings differ.
# Current defaults assume: CH = China, SI = Singapore, SK = South Korea.
CUSTOM_MARKET_CODE_MAP = {
    "US": "United States",
    "UK": "United Kingdom",
    "GE": "Germany",
    "FR": "France",
    "SP": "Spain",
    "IT": "Italy",
    "NO": "Norway",
    "SW": "Sweden",
    "CA": "Canada",
    "MX": "Mexico",
    "JP": "Japan",
    "CH": "China",
    "SI": "Singapore",
    "SK": "South Korea",
    "AU": "Australia",
    "NZ": "New Zealand",
}

# Leave as None to use the embedded always-important key-event table from the prompt.
CUSTOM_ALWAYS_IMPORTANT_TSV = None

# Optional alias overrides used when matching the always-important key-event rules.
# Embedded defaults already include aliases such as:
# SZ = Switzerland, SW = Sweden, EC = Eurozone, JN = Japan, CH = China.
CUSTOM_IMPORTANT_COUNTRY_ALIASES = None

import json
import math
import re
import calendar
from io import StringIO
from pathlib import Path

from datetime import date, datetime
import pandas as pd
try:
    import ipywidgets as widgets
except Exception as exc:
    raise ImportError("This notebook requires ipywidgets. Install it in the notebook environment before running the app.") from exc
from IPython.display import display, HTML, clear_output

try:
    import bql
except Exception:
    bql = None

try:
    from docx import Document
    from docx.enum.section import WD_ORIENT
    from docx.enum.table import WD_TABLE_ALIGNMENT, WD_CELL_VERTICAL_ALIGNMENT
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    from docx.oxml import OxmlElement
    from docx.oxml.ns import qn
    from docx.shared import Inches, Pt, RGBColor
except Exception:
    Document = None
    WD_ORIENT = None
    WD_TABLE_ALIGNMENT = None
    WD_CELL_VERTICAL_ALIGNMENT = None
    WD_ALIGN_PARAGRAPH = None
    OxmlElement = None
    qn = None
    Inches = None
    Pt = None
    RGBColor = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
except Exception:
    Workbook = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None
    get_column_letter = None
    PageMargins = None

APP_CSS = """
<style>
:root {
    --bg: #0b1020;
    --panel: #111827;
    --panel-2: #172033;
    --panel-3: #0f1729;
    --text: #e5eefb;
    --muted: #94a3b8;
    --line: #243248;
    --line-soft: #1b263a;
    --accent: #5da8ff;
    --accent-soft: rgba(93, 168, 255, 0.14);
    --econ: #2ec4b6;
    --earn: #a78bfa;
    --cmdty: #fb7185;
    --holiday: #59e0b5;
    --watch: #facc15;
    --flag: #fb7185;
    --success: #34d399;
}

.event-app {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Helvetica, Arial, sans-serif;
    color: var(--text);
}
.app-title {
    font-size: 30px;
    font-weight: 700;
    margin: 0 0 4px 0;
    color: #d1d7e0;
}
.app-subtitle {
    color: var(--muted);
    font-size: 13px;
    margin-bottom: 4px;
}
.app-footnote {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.45;
}
.section-title {
    font-size: 15px;
    font-weight: 700;
    letter-spacing: 0.01em;
    margin-bottom: 10px;
}
.summary-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(130px, 1fr));
    gap: 10px;
}
.summary-item {
    background: linear-gradient(180deg, rgba(20, 31, 50, 0.85) 0%, rgba(13, 22, 38, 0.98) 100%);
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 12px;
}
.summary-label {
    color: var(--muted);
    font-size: 12px;
    margin-bottom: 4px;
}
.summary-value {
    font-size: 24px;
    font-weight: 700;
}
.badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 999px;
    font-size: 11px;
    font-weight: 700;
    margin-right: 8px;
    border: 1px solid transparent;
}
.badge-econ {
    color: var(--econ);
    background: rgba(46, 196, 182, 0.12);
    border-color: rgba(46, 196, 182, 0.25);
}
.badge-earn {
    color: var(--earn);
    background: rgba(167, 139, 250, 0.13);
    border-color: rgba(167, 139, 250, 0.26);
}
.badge-cmdty {
    color: var(--cmdty);
    background: rgba(251, 113, 133, 0.12);
    border-color: rgba(251, 113, 133, 0.24);
}
.badge-holiday {
    color: var(--holiday);
    background: rgba(245, 158, 11, 0.12);
    border-color: rgba(245, 158, 11, 0.24);
}
.badge-watch {
    color: #ffe98a;
    background: rgba(250, 204, 21, 0.13);
    border-color: rgba(250, 204, 21, 0.28);
}
.badge-flag {
    color: #ffd1d7;
    background: rgba(251, 113, 133, 0.14);
    border-color: rgba(251, 113, 133, 0.3);
}
.detail-title {
    font-size: 22px;
    font-weight: 700;
    margin: 10px 0 6px 0;
}
.detail-subtitle {
    color: var(--muted);
    font-size: 13px;
    margin-bottom: 12px;
}
.field-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(180px, 1fr));
    gap: 10px;
}
.field {
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 10px 12px;
    background: var(--panel-3);
}
.field-label {
    color: var(--muted);
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.03em;
    margin-bottom: 4px;
}
.field-value {
    font-size: 15px;
    font-weight: 600;
}
.note-preview {
    margin-top: 12px;
    border-left: 3px solid var(--accent);
    padding-left: 10px;
    color: #dbe7fa;
    white-space: pre-wrap;
}
.page-caption {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.35;
}
.month-header-bar {
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 12px;
    margin-bottom: 10px;
}
.page-indicator {
    color: var(--muted);
    font-size: 12px;
}
.weekday-label {
    text-align: center;
    font-size: 11px;
    color: var(--muted);
    font-weight: 700;
    letter-spacing: 0.03em;
    text-transform: uppercase;
    padding: 6px 0 2px 0;
}

.calendar-day-box {
    background: var(--panel-3);
    border: 1px solid var(--line-soft);
    border-radius: 12px;
    padding: 6px 8px 8px 8px;
    min-height: 96px;
    box-sizing: border-box;
}
.calendar-day-box.has-events {
    background: rgba(93, 168, 255, 0.08);
    border-color: rgba(93, 168, 255, 0.22);
}
.calendar-day-box.holiday-day {
    box-shadow: inset 0 0 0 1px rgba(245, 158, 11, 0.35);
}
.calendar-day-box.watch-day {
    box-shadow: inset 0 0 0 1px rgba(250, 204, 21, 0.35);
}
.calendar-day-box.flagged-day {
    box-shadow: inset 0 0 0 1px rgba(251, 113, 133, 0.3);
}
.calendar-day-box.selected-day {
    outline: 2px solid var(--accent);
    outline-offset: 0;
}
.calendar-day-box.out-month {
    opacity: 0.42;
}
.calendar-day-box.no-click {
    opacity: 0.7;
}
.day-meta {
    margin-top: 4px;
    font-size: 10px;
    color: var(--muted);
    line-height: 1.25;
}
.day-meta strong {
    color: var(--text);
    font-weight: 700;
}
.holiday-chip {
    display: inline-block;
    border-radius: 999px;
    padding: 1px 6px;
    margin: 1px 4px 0 0;
    font-size: 10px;
    color: #ffd590;
    background: rgba(245, 158, 11, 0.14);
    border: 1px solid rgba(245, 158, 11, 0.28);
}
.day-panel-header {
    margin-bottom: 12px;
}
.day-panel-title {
    font-size: 20px;
    font-weight: 700;
    margin-bottom: 4px;
}
.day-panel-subtitle {
    color: var(--muted);
    font-size: 12px;
}
.day-card {
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 12px 14px;
    background: var(--panel-3);
    margin-bottom: 10px;
}
.day-card:last-child {
    margin-bottom: 0;
}

.day-card-title {
    font-size: 17px;
    font-weight: 700;
    margin: 8px 0 4px 0;
}
.day-card-subtitle {
    color: var(--muted);
    font-size: 12px;
    margin-bottom: 10px;
}
.day-card-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(160px, 1fr));
    gap: 8px;
}
.day-card-field {
    border: 1px solid var(--line-soft);
    border-radius: 10px;
    padding: 8px 10px;
    background: rgba(17, 24, 39, 0.55);
}
.day-card-field-label {
    color: var(--muted);
    font-size: 10px;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.day-card-field-value {
    font-size: 14px;
    font-weight: 600;
}
.day-scroll-box {
    height: 760px;
    max-height: 760px;
    overflow-y: auto;
    padding-right: 6px;
}
.day-scroll-box::-webkit-scrollbar {
    width: 10px;
}
.day-scroll-box::-webkit-scrollbar-thumb {
    background: rgba(148, 163, 184, 0.35);
    border-radius: 99px;
}
.watchlist-box {
    max-height: 220px;
    overflow-y: auto;
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 10px 12px;
    background: var(--panel-3);
}
.watch-item {
    padding: 8px 0;
    border-bottom: 1px solid var(--line-soft);
}
.watch-item:last-child {
    border-bottom: none;
}
.empty-state {
    color: var(--muted);
    font-style: italic;
    padding: 8px 0;
}
.event-app-shell {
    width: 100%;
}
.event-app-shell button {
    background: var(--panel-2);
    color: var(--text);
    border: 1px solid var(--line);
    border-radius: 10px;
}
.event-app-shell button:hover {
    filter: brightness(1.06);
}
.event-app-shell .widget-text input,
.event-app-shell .widget-textarea textarea,
.event-app-shell .widget-select select,
.event-app-shell .widget-select-multiple select,
.event-app-shell .widget-dropdown select {
    background: var(--panel-3) !important;
    color: var(--text) !important;
    border: 1px solid var(--line) !important;
    border-radius: 10px !important;
}
.event-app-shell .event-list-select select {
    font-family: 'SFMono-Regular', Menlo, Monaco, Consolas, 'Liberation Mono', monospace !important;
    font-size: 10.75px !important;
    line-height: 1.42 !important;
    white-space: pre !important;
    letter-spacing: 0 !important;
}
.event-app-shell .event-list-select option {
    font-family: inherit !important;
}
.event-app-shell .widget-label {
    color: var(--muted) !important;
}
.event-app-shell .app-panel {
    background: var(--panel);
    border: 1px solid var(--line);
    border-radius: 16px;
    padding: 14px 16px;
    box-shadow: 0 8px 26px rgba(0, 0, 0, 0.2);
}
.event-app-shell .calendar-day-btn button {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
}
.event-app-shell .calendar-day-btn button {
    color: var(--text) !important;
    font-weight: 700 !important;
    font-size: 13px !important;
    padding: 0 !important;
    text-align: left !important;
    min-height: auto !important;
}
.event-app-shell .calendar-day-btn button:disabled {
    opacity: 1 !important;
}
.event-app-shell .section-spacer {
    height: 8px;
}
</style>
"""

DEFAULT_MARKET_CODE_MAP = {
    "US": "United States",
    "UK": "United Kingdom",
    "GE": "Germany",
    "FR": "France",
    "SP": "Spain",
    "IT": "Italy",
    "NO": "Norway",
    "SW": "Sweden",
    "CA": "Canada",
    "MX": "Mexico",
    "JP": "Japan",
    "CH": "China",
    "SI": "Singapore",
    "SK": "South Korea",
    "AU": "Australia",
    "NZ": "New Zealand",
}

DEFAULT_MARKET_HOLIDAY_TSV = """US\tUK\tGE\tFR\tSP\tIT\tNO\tSW\tCA\tMX\tJP\tCH\tSK\tSI\tAU\tNZ
4/3/2026\t5/8/2026\t4/3/2026\t4/3/2026\t4/3/2026\t4/3/2026\t4/3/2026\t4/3/2026\t4/3/2026\t4/3/2026\t4/3/2026\t5/1/2026\t4/3/2026\t5/1/2026\t4/3/2026\t4/3/2026
5/25/2026\t6/1/2026\t4/6/2026\t4/6/2026\t4/6/2026\t4/6/2026\t4/6/2026\t5/18/2026\t4/3/2026\t4/29/2026\t5/1/2026\t5/1/2026\t5/5/2026\t4/6/2026
# ... (full data table truncated for brevity)
"""

DEFAULT_ALWAYS_IMPORTANT_TSV = """Country\tKeyEvents
US\tWards Total Vehicle Sales
US\tU. of Mich. Sentiment
US\tS&P Global US Services PMI
US\tS&P Global US Manufacturing PMI
US\tPersonal Spending
US\tPersonal Income
US\tJOLTS Job Openings
US\tISM Services Index
US\tISM Manufacturing
US\tInitial Jobless Claims
US\tGDP Annualized QoQ
US\tFOMC Rate Decision (Upper Bound)
US\tFOMC Meeting Minutes
US\tDurable Goods Orders
US\tCore PCE Price Index MoM
US\tConf. Board Consumer Confidence
US\tChange in Nonfarm Payrolls
US\tChallenger Job Cuts YoY
US\tADP Employment Change
US\t3Y High Yield Rate
US\t5Y High Yield Rate
US\t7Y High Yield Rate
US\t10Y High Yield Rate
US\t20Y High Yield Rate
US\t30Y High Yield Rate
US\t2Y High Yield Rate
US\tEmpire Manufacturing
US\tCPI MoM
US\tPPI Final Demand YoY
US\tFed's External Communications Blackout
US\tEmployment Cost Index
US\tADP Weekly Employment Change
UK\tS&P Global UK Services PMI
UK\tS&P Global UK Manufacturing PMI
UK\tRetail Sales Inc Auto Fuel MoM
UK\tPublic-Sector Net Borrowing
UK\tGDP QoQ
UK\tEmployment Change 3M/3M
UK\tCPI YoY
UK\tBank of England Bank Rate
UK\tMonthly GDP (MoM)
SZ\tSNB Policy Rate
SZ\tCPI EU Harmonized YoY
SW\tRiksbank Policy Rate
SW\tCPI YoY
NZ\tRBNZ Official Cash Rate
NZ\tRBNZ Monetary Policy Review
NZ\tGDP YoY
NZ\tCPI QoQ
NZ\tEmployment Change QoQ
NO\tDeposit Rates
NO\tCPI YoY
JN\tTokyo CPI YoY
JN\tNatl CPI YoY
JN\tTankan Large Mfg Index
JN\tJapan to Sell 10-Year Bonds
JN\tHousehold Spending YoY
JN\tJapan to Sell 30-Year Bonds
JN\tReal Cash Earnings YoY
JN\tPPI YoY
JN\tJapan to Sell 20-Year Bonds
JN\tCapital Spending YoY
GE\tGDP SA QoQ
GE\tZEW Survey Expectations
GE\tIFO Business Climate
GE\tHCOB Germany Services PMI
GE\tHCOB Germany Manufacturing PMI
GE\tCPI Saxony MoM
GE\tCPI North Rhine Westphalia MoM
GE\tCPI Hesse MoM
GE\tCPI EU Harmonized YoY
GE\tCPI Brandenburg MoM
GE\tCPI Bavaria MoM
GE\tCPI Baden Wuerttemberg MoM
FR\tHCOB France Services PMI
FR\tHCOB France Manufacturing PMI
FR\tCPI EU Harmonized YoY
EC\tCPI Estimate YoY
EC\tCPI YoY
CH\tRatingDog China PMI Services
CH\tRatingDog China PMI Mfg
CH\tNon-manufacturing PMI
CH\tManufacturing PMI
CH\t5-Year Loan Prime Rate
CH\t1-Year Loan Prime Rate
CH\tGDP YoY
CH\tRetail Sales YoY
CH\tIndustrial Production YoY
CH\tFixed Assets Ex Rural YTD YoY
CA\tRetail Sales MoM
CA\tNet Change in Employment
CA\tGDP MoM
CA\tCPI YoY
CA\tBank of Canada Rate Decision
AU\tRBA Cash Rate Target
AU\tEmployment Change
AU\tCPI YoY
EC\tHCOB Eurozone Manufacturing PMI
EC\tHCOB Eurozone Services PMI
EC\tGDP SA QoQ
CA\tBoC Overall Business Outlook Survey
"""

CATEGORY_ORDER = {"Holiday": 0, "Economic": 1, "Earnings": 2, "Commodity": 3}
SEPARATOR_PREFIX = "_SEP_"
NBSP = "\u00A0"
ELLIPSIS = "\u2026"
LIST_CATEGORY_CODES = {"Economic": "ECON", "Earnings": "EARN", "Commodity": "CMDY", "Holiday": "HOL"}

DEFAULT_IMPORTANT_COUNTRY_ALIASES = {
    "US": ["US", "United States", "United States of America", "USA"],
    "UK": ["UK", "United Kingdom", "Britain", "Great Britain"],
    "SZ": ["SZ", "Switzerland", "Swiss"],
    "SW": ["SW", "Sweden", "Swedish"],
    "NZ": ["NZ", "New Zealand"],
    "NO": ["NO", "Norway", "Norwegian"],
    "JN": ["JN", "Japan", "Japanese"],
    "GE": ["GE", "Germany", "German"],
    "FR": ["FR", "France", "French"],
    "EC": ["EC", "Euro Area", "Eurozone", "Euro Zone", "European Union", "European Community"],
    "CH": ["CH", "China", "Chinese", "Mainland China"],
    "CA": ["CA", "Canada", "Canadian"],
    "AU": ["AU", "Australia", "Australian"],
}


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def is_missing(value):
    if value is None:
        return True
    try:
        missing = pd.isna(value)
    except Exception:
        return False
    try:
        return bool(missing)
    except Exception:
        return False


def has_value(value):
    if is_missing(value):
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def text_or_blank(value, strip=False):
    if is_missing(value):
        return ""
    text = str(value)
    return text.strip() if strip else text


def coalesce(*values, default=""):
    for value in values:
        if has_value(value):
            return value
    return default


def clean_join(values, sep=" | "):
    return sep.join(text_or_blank(value, strip=True) for value in values if has_value(value))


def safe_html(value):
    import html
    if is_missing(value):
        return ""
    return html.escape(str(value))


def fmt_value(value, decimals=2):
    if is_missing(value):
        return "\u2014"
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"{value:.{decimals}f}"
    return str(value)


def fmt_market_cap(value):
    if is_missing(value):
        return "\u2014"
    try:
        numeric_value = float(value)
    except Exception:
        return str(value)
    abs_value = abs(numeric_value)
    if abs_value >= 1_000_000_000_000:
        return f"${numeric_value / 1_000_000_000_000:.2f}T"
    if abs_value >= 1_000_000_000:
        return f"${numeric_value / 1_000_000_000:.2f}B"
    if abs_value >= 1_000_000:
        return f"${numeric_value / 1_000_000:.2f}M"
    return f"${numeric_value:.0f}"


def make_event_id(prefix, *parts):
    tokens = [prefix]
    for part in parts:
        token = "" if part is None else str(part).strip()
        tokens.append(token.replace("|", "/"))
    return "|".join(tokens)


def pick_column(columns, candidates):
    lowered = {str(c).lower(): c for c in columns}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    for col in columns:
        col_str = str(col).lower()
        for candidate in candidates:
            if candidate.lower() in col_str:
                return col
    return None


def as_bool(value):
    if isinstance(value, bool):
        return value
    if is_missing(value):
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "t", "yes", "y"}


def normalize_key_text(value):
    text = text_or_blank(value, strip=True).lower()
    text = re.sub(r"[^a-z0-9\s]", "", text)
    return re.sub(r" +", " ", text).strip()


def fmt_metric(value, scaling_factor=None, decimals=2):
    if is_missing(value):
        return "\u2014"
    metric_text = fmt_value(value, decimals=decimals)
    scale_text = text_or_blank(scaling_factor, strip=True)
    return f"{metric_text}{scale_text}".strip()


def fmt_time_text(value):
    return text_or_blank(value, strip=True) or "TBA"


def truncate_display_text(value, width, default="\u2014"):
    text = text_or_blank(value, strip=True)
    if not text:
        text = default
    if width <= 0:
        return ""
    if len(text) <= width:
        return text
    if width == 1:
        return text[:1]
    return text[:width - 1] + ELLIPSIS


def display_cell(value, width, align="left", default="\u2014"):
    text = truncate_display_text(value, width=width, default=default)
    pad = max(0, width - len(text))
    filler = NBSP * pad
    if align == "right":
        return f"{filler}{text}"
    return f"{text}{filler}"


def join_display_cells(cells):
    return f"{NBSP}|{NBSP}".join(cells)


# ---------------------------------------------------------------------------
# docx helper functions
# ---------------------------------------------------------------------------

def _docx_set_cell_shading(cell, fill):
    if OxmlElement is None or qn is None:
        return
    tc_pr = cell._tc.get_or_add_tcPr()
    shd = tc_pr.find(qn("w:shd"))
    if shd is None:
        shd = OxmlElement("w:shd")
        tc_pr.append(shd)
    shd.set(qn("w:val"), "clear")
    shd.set(qn("w:color"), "auto")
    shd.set(qn("w:fill"), fill)


def _docx_set_cell_margins(cell, top=60, start=85, bottom=60, end=85):
    if OxmlElement is None or qn is None:
        return
    tc_pr = cell._tc.get_or_add_tcPr()
    tc_mar = tc_pr.find(qn("w:tcMar"))
    if tc_mar is None:
        tc_mar = OxmlElement("w:tcMar")
        tc_pr.append(tc_mar)
    for side, value in [("top", top), ("start", start), ("bottom", bottom), ("end", end)]:
        node = tc_mar.find(qn(f"w:{side}"))
        if node is None:
            node = OxmlElement(f"w:{side}")
            tc_mar.append(node)
        node.set(qn("w:w"), str(value))
        node.set(qn("w:type"), "dxa")


def _docx_set_repeat_header(row):
    if OxmlElement is None or qn is None:
        return
    tr_pr = row._tr.get_or_add_trPr()
    tbl_header = OxmlElement("w:tblHeader")
    tbl_header.set(qn("w:val"), "true")
    tr_pr.append(tbl_header)


def _docx_write_cell(cell, text, font_name="Arial", font_size=7.5, bold=False,
                     color="111827", align="left", fill=None):
    cell.text = ""
    paragraph = cell.paragraphs[0]
    paragraph.paragraph_format.space_before = Pt(0) if Pt is not None else None
    paragraph.paragraph_format.space_after = Pt(0) if Pt is not None else None
    paragraph.paragraph_format.line_spacing = 1.05
    if WD_ALIGN_PARAGRAPH is not None:
        if align == "center":
            paragraph.alignment = WD_ALIGN_PARAGRAPH.CENTER
        elif align == "right":
            paragraph.alignment = WD_ALIGN_PARAGRAPH.RIGHT
        else:
            paragraph.alignment = WD_ALIGN_PARAGRAPH.LEFT
    run = paragraph.add_run("" if text is None else str(text))
    if Pt is not None:
        run.font.size = Pt(font_size)
    run.font.name = font_name
    run.font.bold = bold
    if RGBColor is not None:
        run.font.color.rgb = RGBColor.from_string(color)
    if WD_CELL_VERTICAL_ALIGNMENT is not None:
        cell.vertical_alignment = WD_CELL_VERTICAL_ALIGNMENT.CENTER
    _docx_set_cell_margins(cell)
    if fill:
        _docx_set_cell_shading(cell, fill)
class BloombergEventCalendarApp:

    def __init__(
        self,
        bql=None,
        bq=None,
        eco_universe=None,
        comdty_tickers=None,
        top_n_market_cap=250,
        state_path="event_annotations.csv",
        legacy_state_path="event_calendar_state.json",
        export_dir="event_calendar_exports",
        market_holidays_tsv=None,
        market_code_map=None,
        always_important_tsv=None,
        important_country_aliases=None,
        months_per_page_default=3,
    ):
        self.bql = bql
        if bq is not None:
            self.bq = bq
        elif self.bql is not None:
            try:
                self.bq = self.bql.Service()
            except Exception:
                self.bq = None
        else:
            self.bq = None

        self.eco_universe = eco_universe or [
            'US Country', 'EU Country', 'IN Country', 'CN Country', 'KR Country',
            'AU Country', 'NZ Country', 'BZ Country', 'MX Country', 'CA Country',
            'CL Country', 'CO Country',
        ]

        self.comdty_tickers = comdty_tickers or [
            'CL1 Comdty', 'CO1 Comdty', 'NG1 Comdty', 'GC1 Comdty', 'SI1 Comdty',
            'HG1 Comdty', 'ES1 Comdty', 'RTY1 Comdty', 'NQ1 Comdty', 'Z1 Comdty',
        ]

        self.top_n_market_cap = top_n_market_cap
        self.state_path = Path(state_path)
        self.legacy_state_path = Path(legacy_state_path) if legacy_state_path else None
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        self.market_code_map = dict(DEFAULT_MARKET_CODE_MAP)
        if market_code_map:
            self.market_code_map.update(market_code_map)
        self.market_holidays_tsv = market_holidays_tsv or DEFAULT_MARKET_HOLIDAY_TSV

        self.important_country_aliases = {
            code: list(values)
            for code, values in DEFAULT_IMPORTANT_COUNTRY_ALIASES.items()
        }
        if important_country_aliases:
            for code, values in important_country_aliases.items():
                if values is None:
                    continue
                if isinstance(values, (list, tuple, set)):
                    self.important_country_aliases[str(code).strip().upper()] = [str(v) for v in values]
                else:
                    self.important_country_aliases[str(code).strip().upper()] = [str(values)]
        self.always_important_tsv = always_important_tsv or DEFAULT_ALWAYS_IMPORTANT_TSV
        self.always_important_rules = self._parse_always_important_table(self.always_important_tsv)

        self.core_events = pd.DataFrame(columns=self.master_columns())
        self.all_holiday_events = self._build_all_holiday_events(self.market_holidays_tsv)
        self.events = pd.DataFrame(columns=self.master_columns())
        self.filtered_events = pd.DataFrame(columns=self.master_columns())
        self.state = {}
        self.current_event_id = None
        self.selected_day = None
        self.month_page_start = 0
        self._suspend_state_events = False
        self._suspend_month_events = False

        self._build_widgets(months_per_page_default)
        self._load_state_from_disk(silent=True)
        self._wire_events()

    @staticmethod
    def master_columns():
        return [
            'event_id', 'event_date', 'event_time', 'category', 'title', 'subtitle',
            'ticker', 'company', 'period', 'survey', 'actual', 'prior', 'revision',
            'scaling_factor', 'market_cap_usd', 'future_price', 'currency',
            'underlying_ticker', 'delivery_date', 'expire_group_id', 'orig_ids',
        ]

    def _series(self, df, col, default=None):
        if col is None or col not in df.columns:
            return pd.Series([default] * len(df), index=df.index)
        return df[col]

    def _build_widgets(self, months_per_page_default):
        default_start = pd.Timestamp.today().date()
        default_end = (pd.Timestamp.today() + pd.Timedelta(days=45)).date()

        common_desc = {'description_width': 'initial'}

        self.start_picker = widgets.DatePicker(description='Start', value=default_start, style=common_desc)
        self.end_picker = widgets.DatePicker(description='End', value=default_end, style=common_desc)
        self.refresh_button = widgets.Button(description='Refresh Data', button_style='primary', icon='refresh')
        self.apply_filters_button = widgets.Button(description='Apply Explorer Filters', icon='filter')
        self.reset_filters_button = widgets.Button(description='Reset Filters', icon='undo')

        self.search_box = widgets.Text(
            description='Search',
            placeholder='Event, ticker, company, country...',
            style=common_desc,
            layout=widgets.Layout(width='340px'),
        )

        self.category_filter = widgets.SelectMultiple(
            description='Explorer Categories',
            options=['Economic', 'Earnings', 'Commodity', 'Holiday'],
            value=('Economic', 'Earnings', 'Commodity', 'Holiday'),
            rows=5,
            style=common_desc,
            layout=widgets.Layout(width='260px'),
        )

        self.only_flagged = widgets.Checkbox(description='Flagged only', value=False, style=common_desc)
        self.only_watch = widgets.Checkbox(description='Important only', value=False, style=common_desc)

        self.months_per_page = widgets.BoundedIntText(
            description='Months / pane',
            value=max(1, int(months_per_page_default)),
            min=1,
            max=12,
            style=common_desc,
            layout=widgets.Layout(width='170px'),
        )

        self.prev_months_button = widgets.Button(description='Previous Pane', icon='arrow-left')
        self.next_months_button = widgets.Button(description='Next Pane', icon='arrow-right')
        self.month_page_label = widgets.HTML()
        self.month_page_note = widgets.HTML()

        self.export_csv_path = widgets.HTML(
            '<div class="event-app page-caption">CSV export preview will appear here after each export.</div>'
        )

        self.export_csv_preview = widgets.Textarea(
            description='CSV Preview',
            placeholder='After export, the CSV text will appear here for copy / paste.',
            layout=widgets.Layout(width='100%', height='220px'),
            style=common_desc,
        )

        self.persistence_note = widgets.HTML(
            f'<div class="event-app page-caption">'
            f'Annotations auto-save to <b>{safe_html(str(self.state_path))}</b>. '
            f'Rule-based key releases from the embedded watch list are always marked important, even after clearing user-selected important items.'
            f'</div>'
        )

        self.status_out = widgets.Output()
        self.summary_out = widgets.Output()
        self.monthly_out = widgets.Output()
        self.day_out = widgets.Output(layout=widgets.Layout(height='780px', overflow='auto'))
        self.detail_html = widgets.HTML()
        self.watchlist_out = widgets.Output()

        controls_title = widgets.HTML('<div class="event-app section-title">Controls</div>')
        controls_grid = widgets.VBox([
            widgets.HBox([self.start_picker, self.end_picker, self.refresh_button]),
            widgets.HBox([self.search_box, self.apply_filters_button, self.reset_filters_button]),
            widgets.HBox([self.category_filter, widgets.VBox([self.only_flagged, self.only_watch])]),
        ])
        self.controls_panel = widgets.VBox([controls_title, controls_grid])
        self.controls_panel.add_class('app-panel')

        summary_title = widgets.HTML('<div class="event-app section-title">Explorer Summary</div>')
        self.summary_panel = widgets.VBox([summary_title, self.summary_out])
        self.summary_panel.add_class('app-panel')

        monthly_title = widgets.HTML('<div class="event-app section-title">Monthly Page</div>')
        month_nav = widgets.HBox([
            self.months_per_page,
            self.prev_months_button,
            self.next_months_button,
            self.month_page_label,
        ])
        self.monthly_panel = widgets.VBox([monthly_title, month_nav, self.month_page_note, self.monthly_out])
        self.monthly_panel.add_class('app-panel')
        self.monthly_panel.layout = widgets.Layout(width='67%')

        day_title = widgets.HTML('<div class="event-app section-title">Selected Day</div>')
        self.day_panel = widgets.VBox([day_title, self.day_out])
        self.day_panel.add_class('app-panel')
        self.day_panel.layout = widgets.Layout(width='33%')

        list_layout = widgets.Layout(width='100%', height='540px')
        self.all_list = widgets.Select(description='', options=[], layout=list_layout)
        self.econ_list = widgets.Select(description='', options=[], layout=list_layout)
        self.earnings_list = widgets.Select(description='', options=[], layout=list_layout)
        self.comdty_list = widgets.Select(description='', options=[], layout=list_layout)
        self.holiday_list = widgets.Select(description='', options=[], layout=list_layout)
        self.watch_list = widgets.Select(description='', options=[], layout=list_layout)

        for selector in [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list]:
            selector.add_class('event-list-select')

        self.tabs = widgets.Tab(children=[
            self.all_list,
            self.econ_list,
            self.earnings_list,
            self.comdty_list,
            self.holiday_list,
            self.watch_list,
        ])
        for i, title in enumerate(['ALL', 'ECONOMIC', 'EARNINGS', 'COMMODITIES', 'HOLIDAYS', 'IMPORTANT']):
            self.tabs.set_title(i, title)

        self.flag_toggle = widgets.ToggleButton(
            description='Flag', icon='flag-o', value=False, layout=widgets.Layout(width='120px')
        )

        self.watch_toggle = widgets.ToggleButton(
            description='Important', icon='star-o', value=False, layout=widgets.Layout(width='150px')
        )

        self.note_area = widgets.Textarea(
            description='Notes',
            placeholder='Add custom notes, trading angles, reminders, scenario comments...',
            layout=widgets.Layout(width='100%', height='150px'),
            style=common_desc,
        )

        self.save_note_button = widgets.Button(description='Save Note', button_style='success', icon='save')
        self.clear_note_button = widgets.Button(description='Clear Note', icon='trash')
        self.export_state_button = widgets.Button(description='Sync CSV Now', icon='save')
        self.load_state_button = widgets.Button(description='Reload CSV', icon='refresh')
        self.clear_watchlist_button = widgets.Button(description='Clear Important List', icon='times')

        self.export_scope_dropdown = widgets.Dropdown(
            description='Export Scope',
            options=[
                ('Current visible pane', 'pane'),
                ('Full selected range', 'range'),
                ('Important / flagged only', 'highlights'),
            ],
            value='pane',
            style=common_desc,
            layout=widgets.Layout(width='330px'),
        )

        self.export_document_button = widgets.Button(
            description='Export View CSV', icon='download', button_style='primary'
        )

        self.export_note = widgets.HTML(
            '<div class="event-app page-caption">'
            'Creates a rich CSV snapshot for the active view. Copy the CSV text from the box below into a local file if notebook downloads are blocked, then run the '
            'companion local builder notebook to format the report into a Wall Street-style Excel workbook with an Event Brief, Data Snapshot, and Daily One-Pager.'
            '</div>'
        )

        list_caption = widgets.HTML(
            '<div class="event-app page-caption">Aligned fixed-width explorer rows. '
            'Symbols: ★ important, ⚑ flagged.</div>'
        )
        list_card = widgets.VBox([list_caption, self.tabs], layout=widgets.Layout(width='48%'))
        list_card.add_class('app-panel')

        detail_title = widgets.HTML('<div class="event-app section-title">Event Detail & Notes</div>')
        detail_controls = widgets.HBox([
            self.flag_toggle,
            self.watch_toggle,
            self.save_note_button,
            self.clear_note_button,
        ])
        detail_actions = widgets.HBox([self.export_state_button, self.load_state_button, self.clear_watchlist_button])
        export_controls = widgets.HBox([self.export_scope_dropdown, self.export_document_button])
        detail_card = widgets.VBox([
            detail_title,
            self.detail_html,
            detail_controls,
            self.note_area,
            detail_actions,
            self.persistence_note,
            export_controls,
            self.export_note,
            self.export_csv_path,
            self.export_csv_preview,
            self.watchlist_out,
            self.status_out,
        ], layout=widgets.Layout(width='52%'))
        detail_card.add_class('app-panel')

        header_html = widgets.HTML(
            APP_CSS +
            '''
            <div class="event-app">
            <div class="app-title">Market Event Calendar</div>
            <div class="app-subtitle">
            Dark-themed notebook calendar for economic releases, large-cap earnings, futures timing, market holidays, and embedded high-priority macro
            events.
            </div>
            <div class="app-footnote">
            Explorer filters control the event tabs. The monthly page keeps holidays and key-event importance visible, annotations auto-save to CSV, and the
            export workflow now produces copy-friendly CSV output for a separate local formatting notebook.
            </div>
            </div>
            '''
        )

        self.root = widgets.VBox([
            header_html,
            self.controls_panel,
            self.summary_panel,
            widgets.HBox([self.monthly_panel, self.day_panel]),
            widgets.HBox([list_card, detail_card]),
        ], layout=widgets.Layout(width='100%'))
        self.root.add_class('event-app-shell')

    def _wire_events(self):
        self.refresh_button.on_click(self.refresh_data)
        self.apply_filters_button.on_click(self.apply_filters)
        self.reset_filters_button.on_click(self.reset_filters)
        self.save_note_button.on_click(self.save_note)
        self.clear_note_button.on_click(self.clear_note)
        self.export_state_button.on_click(self.export_state)
        self.load_state_button.on_click(self.load_state)
        self.clear_watchlist_button.on_click(self.clear_watchlist)
        self.export_document_button.on_click(self.export_current_view_csv)
        self.prev_months_button.on_click(self.previous_month_pane)
        self.next_months_button.on_click(self.next_month_pane)
        self.months_per_page.observe(self.on_months_per_page_changed, names='value')

        for selector in [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list]:
            selector.observe(self.on_event_selected, names='value')

        self.tabs.observe(self.on_tab_change, names='selected_index')
        self.flag_toggle.observe(self.on_flag_toggled, names='value')
        self.watch_toggle.observe(self.on_watch_toggled, names='value')

    def display(self):
        display(self.root)

    def _validate_dates(self):
        if self.start_picker.value is None or self.end_picker.value is None:
            raise ValueError("Please choose both a start and an end date.")
        if self.start_picker.value > self.end_picker.value:
            raise ValueError("Start date must be on or before end date.")
        return pd.Timestamp(self.start_picker.value).normalize(), pd.Timestamp(self.end_picker.value).normalize()

    def log(self, message):
        with self.status_out:
            clear_output(wait=True)
            print(message)

    def _require_bql(self):
        if self.bq is None or self.bql is None:
            raise RuntimeError("BQL is not available in this environment. Load existing dataframes or use a Bloomberg-enabled notebook session.")

    def load_from_dataframes(self, eco_df=None, earnings_df=None, comdty_df=None):
        eco_df = eco_df.copy() if eco_df is not None else pd.DataFrame()
        earnings_df = earnings_df.copy() if earnings_df is not None else pd.DataFrame()
        comdty_df = comdty_df.copy() if comdty_df is not None else pd.DataFrame()

        self.core_events = self.build_core_events(eco_df, earnings_df, comdty_df)

        if not self.core_events.empty:
            inferred_start = self.core_events['event_date'].min().normalize()
            inferred_end = self.core_events['event_date'].max().normalize()
            self.start_picker.value = inferred_start.date()
            self.end_picker.value = inferred_end.date()
            self.selected_day = inferred_start.date()
        else:
            self.selected_day = self.start_picker.value

        self.month_page_start = 0
        self.apply_filters()
        holiday_count = len(self.prepare_holiday_events(*self._validate_dates()))
        self.log(
            f"Loaded {len(self.core_events)} live events and {holiday_count} market holiday markers "
            f"for {self.start_picker.value} to {self.end_picker.value}."
        )

    def refresh_data(self, _=None):
        try:
            self._require_bql()
            start_ts, end_ts = self._validate_dates()
            self.log(f"Refreshing data for {start_ts.date()} to {end_ts.date()} ...")
            eco_df = self.fetch_economic_calendar(start_ts, end_ts)
            earnings_df = self.fetch_earnings_calendar(start_ts, end_ts)
            comdty_df = self.fetch_commodity_calendar(start_ts, end_ts)
            self.core_events = self.build_core_events(eco_df, earnings_df, comdty_df)
            self.selected_day = start_ts.date()
            self.month_page_start = 0
            self.apply_filters()
            holiday_count = len(self.prepare_holiday_events(start_ts, end_ts))
            self.log(
                f"Loaded {len(self.core_events)} live events and {holiday_count} market holiday markers "
                f"for {start_ts.date()} to {end_ts.date()}."
            )
        except Exception as exc:
            self.log(f"Data refresh failed: {exc}")

    def fetch_economic_calendar(self, start_ts, end_ts):
        date_range = self.bq.func.range(start_ts.strftime('%Y-%m-%d'), end_ts.strftime('%Y-%m-%d'))
        data_item = self.bq.data.calendar(type='ECONOMIC_RELEASES', dates=date_range)
        request = self.bql.Request(self.eco_universe, data_item)
        response = self.bq.execute(request)
        if not response:
            return pd.DataFrame()
        df = pd.concat([item.df() for item in response], axis=1).reset_index()
        return df

    def fetch_earnings_calendar(self, start_ts, end_ts):
        universe = self.bq.univ.filter(
            self.bq.univ.equities('active', 'primary'),
            self.bq.func.grouprank(self.bq.data.cur_mkt_cap(currency='USD')) <= self.top_n_market_cap
        )
        fields = {
            'Market Cap (USD)': self.bq.func.groupsort(self.bq.data.cur_mkt_cap(currency='USD'), order='desc'),
            'Company Name': self.bq.data.name(),
            'Expected Report Date': self.bq.data.expected_report_dt()
        }
        request = self.bql.Request(universe, fields)
        response = self.bq.execute(request)
        if not response:
            return pd.DataFrame()
        df = pd.concat([item.df()[item.name] for item in response], axis=1).reset_index()
        dt_col = pick_column(df.columns, ['Expected Report Date'])
        if dt_col:
            df[dt_col] = pd.to_datetime(df[dt_col], errors='coerce')
            df = df[df[dt_col].between(start_ts, end_ts)]
        return df

    def fetch_commodity_calendar(self, start_ts, end_ts):
        universe = self.bq.univ.options(self.bq.univ.futures(self.comdty_tickers))

        grouped_underlying = self.bq.func.first(
            self.bq.func.group(
                self.bq.data.undl_ticker()['value'],
                self.bq.data.expire_dt()['value'],
            )
        )

        request = self.bql.Request(universe, {'Underlying Ticker': grouped_underlying})
        response = self.bq.execute(request)
        if not response:
            return pd.DataFrame()

        grouped_df = pd.concat([item.df() for item in response], axis=1).reset_index()

        id_col = pick_column(grouped_df.columns, ['ID'])
        if id_col:
            grouped_df = grouped_df[grouped_df[id_col].astype(str).ne('NullGroup')]
            grouped_df = grouped_df.rename(columns={id_col: 'EXPIRE_GROUP_ID'})

        undl_col = pick_column(grouped_df.columns, ['Underlying Ticker'])
        expire_col = pick_column(grouped_df.columns, ['EXPIRE_DT.VALUE', 'expire_dt', 'expire dt'])
        orig_id_col = pick_column(grouped_df.columns, ['ORIG_IDS', 'orig_ids'])

        if undl_col is None:
            return pd.DataFrame()

        underlying_tickers = grouped_df[undl_col].dropna().unique().tolist()
        if not underlying_tickers:
            return pd.DataFrame()

        fields = {
            'delivery_dates': self.bq.data.FUT_LAST_TRADE_DT(),
            'Future Price': self.bq.data.px_last()
        }

        request2 = self.bql.Request(underlying_tickers, fields)
        response2 = self.bq.execute(request2)
        if not response2:
            return pd.DataFrame()
        price_df = pd.concat([item.df() for item in response2], axis=1).reset_index()

        price_id_col = pick_column(price_df.columns, ['ID'])
        if price_id_col:
            price_df = price_df.rename(columns={price_id_col: 'Underlying Ticker'})

        keep_cols = ['Underlying Ticker']
        for col in ['EXPIRE_GROUP_ID', orig_id_col, expire_col]:
            if col and col not in keep_cols:
                keep_cols.append(col)

        merged = price_df.merge(
            grouped_df[keep_cols].drop_duplicates(),
            on='Underlying Ticker',
            how='left'
        )

        event_date_col = expire_col if expire_col in merged.columns else pick_column(merged.columns, ['delivery_dates'])
        if event_date_col:
            merged[event_date_col] = pd.to_datetime(merged[event_date_col], errors='coerce')
            merged = merged[merged[event_date_col].between(start_ts, end_ts)]
        return merged

    def build_core_events(self, eco_df, earnings_df, comdty_df):
        frames = [
            self.prepare_economic_events(eco_df),
            self.prepare_earnings_events(earnings_df),
            self.prepare_commodity_events(comdty_df),
        ]

        frames = [frame for frame in frames if frame is not None and not frame.empty]

        if not frames:
            return pd.DataFrame(columns=self.master_columns())

        events = pd.concat(frames, ignore_index=True)
        events['event_date'] = pd.to_datetime(events['event_date'], errors='coerce')
        events = events[events['event_date'].notna()].copy()
        events['sort_time'] = events['event_time'].fillna('99:99')
        events['category_order'] = events['category'].map(CATEGORY_ORDER).fillna(99)
        events = events.sort_values(['event_date', 'category_order', 'sort_time', 'title']).reset_index(drop=True)
        return events.drop(columns=['sort_time', 'category_order'])

    def compose_events_for_range(self, start_ts, end_ts):
        frames = []
        core_df = self.core_events.copy()
        if not core_df.empty:
            core_df['event_date'] = pd.to_datetime(core_df['event_date'], errors='coerce')
            core_df = core_df[core_df['event_date'].between(start_ts, end_ts)]
            frames.append(core_df)
        holiday_df = self.prepare_holiday_events(start_ts, end_ts)
        if not holiday_df.empty:
            frames.append(holiday_df)

        if not frames:
            return pd.DataFrame(columns=self.master_columns())

        events = pd.concat(frames, ignore_index=True)
        events['event_date'] = pd.to_datetime(events['event_date'], errors='coerce')
        events = events[events['event_date'].notna()].copy()
        events['sort_time'] = events['event_time'].fillna('99:99')
        events['category_order'] = events['category'].map(CATEGORY_ORDER).fillna(99)
        events = events.sort_values(['event_date', 'category_order', 'sort_time', 'title']).reset_index(drop=True)
        return events.drop(columns=['sort_time', 'category_order'])

    def _parse_market_holiday_table(self, tsv_text):
        if tsv_text is None or not str(tsv_text).strip():
            return pd.DataFrame(columns=['market_code', 'holiday_date', 'country_name'])

        lines = [line.strip() for line in str(tsv_text).strip().splitlines() if line.strip()]

        if not lines:
            return pd.DataFrame(columns=['market_code', 'holiday_date', 'country_name'])

        headers = [item.strip() for item in lines[0].split('\t')]
        rows = []
        for raw_line in lines[1:]:
            parts = raw_line.split('\t')
            if len(parts) < len(headers):
                parts = parts + [''] * (len(headers) - len(parts))
            elif len(parts) > len(headers):
                parts = parts[:len(headers)]

            for code, raw_value in zip(headers, parts):
                raw_value = str(raw_value).strip()
                if not raw_value:
                    continue
                dt = pd.to_datetime(raw_value, format='%m/%d/%Y', errors='coerce')
                if pd.isna(dt):
                    dt = pd.to_datetime(raw_value, errors='coerce')
                if pd.isna(dt):
                    continue
                rows.append({
                    'market_code': code,
                    'holiday_date': dt.normalize(),
                    'country_name': self.market_code_map.get(code, code),
                })

        out = pd.DataFrame(rows).drop_duplicates()
        if out.empty:
            return pd.DataFrame(columns=['market_code', 'holiday_date', 'country_name'])
        return out.sort_values(['holiday_date', 'market_code']).reset_index(drop=True)

    def _build_all_holiday_events(self, tsv_text):
        df = self._parse_market_holiday_table(tsv_text)
        if df.empty:
            return pd.DataFrame(columns=self.master_columns())

        out = pd.DataFrame(index=df.index)
        out['event_id'] = [
            make_event_id('HOL', code, holiday_date.date())
            for code, holiday_date in zip(df['market_code'], df['holiday_date'])
        ]
        out['event_date'] = pd.to_datetime(df['holiday_date'], errors='coerce')
        out['event_time'] = None
        out['category'] = 'Holiday'
        out['title'] = df['market_code'].astype(str) + ' - market closed'
        out['subtitle'] = df['country_name'].astype(str) + ' | National holiday / exchange closure'
        out['country'] = df['country_name']
        out['ticker'] = df['market_code']
        out['company'] = None
        out['period'] = 'Market Closed'
        out['survey'] = None
        out['actual'] = None
        out['prior'] = None
        out['revision'] = None
        out['scaling_factor'] = None
        out['market_cap_usd'] = None
        out['future_price'] = None
        out['currency'] = None
        out['underlying_ticker'] = None
        out['delivery_date'] = None
        out['expire_group_id'] = None
        out['orig_ids'] = None
        return out[self.master_columns()]

    def prepare_holiday_events(self, start_ts=None, end_ts=None):
        df = self.all_holiday_events.copy()
        if df.empty:
            return df
        if start_ts is not None:
            df = df[df['event_date'] >= pd.Timestamp(start_ts).normalize()]
        if end_ts is not None:
            df = df[df['event_date'] <= pd.Timestamp(end_ts).normalize()]
        return df.reset_index(drop=True)

    def prepare_economic_events(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=self.master_columns())

        release_date_col = pick_column(df.columns, ['RELEASE_DATE'])
        time_col = pick_column(df.columns, ['RELEASE_TIME'])
        country_col = pick_column(df.columns, ['COUNTRY_NAME'])
        event_col = pick_column(df.columns, ['EVENT_NAME'])
        period_col = pick_column(df.columns, ['PERIOD'])
        survey_col = pick_column(df.columns, ['SURVEY_MEDIAN'])
        actual_col = pick_column(df.columns, ['ACTUAL'])
        prior_col = pick_column(df.columns, ['PRIOR'])
        revision_col = pick_column(df.columns, ['REVISION'])
        scaling_col = pick_column(df.columns, ['SCALING_FACTOR'])
        ticker_col = pick_column(df.columns, ['CALENDAR'])

        country_series = self._series(df, country_col, None)
        period_series = self._series(df, period_col, None)
        title_series = self._series(df, event_col, None)
        ticker_series = self._series(df, ticker_col, None)

        subtitle_parts = []
        for country_value, period_value, ticker_value in zip(country_series, period_series, ticker_series):
            parts = []
            country_text = text_or_blank(country_value, strip=True)
            period_text = text_or_blank(period_value, strip=True)
            ticker_text = text_or_blank(ticker_value, strip=True)
            if country_text:
                parts.append(country_text)
            if period_text:
                parts.append(f"Period {period_text}")
            if ticker_text:
                parts.append(f"Ticker {ticker_text}")
            subtitle_parts.append(' | '.join(parts) or None)

        out = pd.DataFrame(index=df.index)
        out['event_id'] = [
            make_event_id('ECO', c, t, p, d, tm)
            for c, t, p, d, tm in zip(
                country_series,
                title_series,
                period_series,
                pd.to_datetime(self._series(df, release_date_col), errors='coerce'),
                self._series(df, time_col, ''),
            )
        ]
        out['event_date'] = pd.to_datetime(self._series(df, release_date_col), errors='coerce')
        out['event_time'] = self._series(df, time_col, None)
        out['category'] = 'Economic'
        out['title'] = title_series
        out['subtitle'] = pd.Series(subtitle_parts, index=df.index).replace('', pd.NA)
        out['country'] = country_series.replace('', pd.NA)
        out['ticker'] = ticker_series
        out['company'] = None
        out['period'] = period_series.replace('', pd.NA)
        out['survey'] = self._series(df, survey_col, None)
        out['actual'] = self._series(df, actual_col, None)
        out['prior'] = self._series(df, prior_col, None)
        out['revision'] = self._series(df, revision_col, None)
        out['scaling_factor'] = self._series(df, scaling_col, None)
        out['market_cap_usd'] = None
        out['future_price'] = None
        out['currency'] = None
        out['underlying_ticker'] = None
        out['delivery_date'] = None
        out['expire_group_id'] = None
        out['orig_ids'] = None
        out = out[out['event_date'].notna()]
        return out[self.master_columns()]

    def prepare_commodity_events(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=self.master_columns())

        undl_col = pick_column(df.columns, ['Underlying Ticker'])
        delivery_col = pick_column(df.columns, ['delivery_dates'])
        price_col = pick_column(df.columns, ['Future Price'])
        currency_col = pick_column(df.columns, ['CURRENCY', 'currency'])
        expire_group_col = pick_column(df.columns, ['EXPIRE_GROUP_ID'])
        orig_id_col = pick_column(df.columns, ['ORIG_IDS', 'orig_ids'])
        expire_col = pick_column(df.columns, ['EXPIRE_DT.VALUE', 'expire_dt', 'expire dt'])

        event_date_source = expire_col or delivery_col
        if event_date_source is None:
            return pd.DataFrame(columns=self.master_columns())

        undl_series = self._series(df, undl_col, 'Unknown future')

        out = pd.DataFrame(index=df.index)
        out['event_id'] = [
            make_event_id('CMDTY', und, dt)
            for und, dt in zip(undl_series, pd.to_datetime(self._series(df, event_date_source), errors='coerce'))
        ]
        out['event_date'] = pd.to_datetime(self._series(df, event_date_source), errors='coerce')
        out['event_time'] = None
        out['category'] = 'Commodity'
        out['title'] = undl_series.astype(str) + ' expiry'
        out['subtitle'] = 'Commodity / index futures timing'
        out['country'] = None
        out['ticker'] = undl_series
        out['company'] = None
        out['period'] = None
        out['survey'] = None
        out['actual'] = None
        out['prior'] = None
        out['revision'] = None
        out['scaling_factor'] = None
        out['market_cap_usd'] = None
        out['future_price'] = self._series(df, price_col, None)
        out['currency'] = self._series(df, currency_col, None)
        out['underlying_ticker'] = undl_series
        out['delivery_date'] = pd.to_datetime(self._series(df, delivery_col), errors='coerce')
        out['expire_group_id'] = self._series(df, expire_group_col, None)
        out['orig_ids'] = self._series(df, orig_id_col, None)
        out = out[out['event_date'].notna()]
        return out[self.master_columns()]

    def prepare_earnings_events(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=self.master_columns())

        id_col = pick_column(df.columns, ['ID'])
        mkt_cap_col = pick_column(df.columns, ['Market Cap (USD)'])
        company_col = pick_column(df.columns, ['Company Name'])
        dt_col = pick_column(df.columns, ['Expected Report Date'])

        if id_col is not None:
            ticker_series = self._series(df, id_col, 'Unknown ticker')
            if getattr(df.index, 'name', None) is not None:
                ticker_series = pd.Series(df.index.astype(str), index=df.index)
        else:
            ticker_series = pd.Series(['Unknown ticker'] * len(df), index=df.index)

        company_series = self._series(df, company_col, 'Unknown company')

        out = pd.DataFrame(index=df.index)
        out['event_id'] = [
            make_event_id('ERN', ticker, dt)
            for ticker, dt in zip(ticker_series, pd.to_datetime(self._series(df, dt_col), errors='coerce'))
        ]
        out['event_date'] = pd.to_datetime(self._series(df, dt_col), errors='coerce')
        out['event_time'] = None
        out['category'] = 'Earnings'
        out['title'] = company_series.astype(str)
        out['subtitle'] = ticker_series.astype(str)
        out['country'] = None
        out['ticker'] = ticker_series
        out['company'] = company_series
        out['period'] = None
        out['survey'] = None
        out['actual'] = None
        out['prior'] = None
        out['revision'] = None
        out['scaling_factor'] = None
        out['market_cap_usd'] = self._series(df, mkt_cap_col, None)
        out['future_price'] = None
        out['currency'] = None
        out['underlying_ticker'] = None
        out['delivery_date'] = None
        out['expire_group_id'] = None
        out['orig_ids'] = None
        out = out[out['event_date'].notna() & out['title'].notna()]
        return out[self.master_columns()]

    def _parse_always_important_table(self, tsv_text):
        if tsv_text is None or not str(tsv_text).strip():
            return pd.DataFrame(columns=['country_code', 'event_name', 'normalized_event'])
        try:
            df = pd.read_csv(StringIO(str(tsv_text)), sep='\t')
        except Exception:
            return pd.DataFrame(columns=['country_code', 'event_name', 'normalized_event'])
        if df.empty:
            return pd.DataFrame(columns=['country_code', 'event_name', 'normalized_event'])
        country_col = pick_column(df.columns, ['Country'])
        event_col = pick_column(df.columns, ['Key Events', 'Key Event', 'Event', 'Event Name'])
        if country_col is None or event_col is None:
            return pd.DataFrame(columns=['country_code', 'event_name', 'normalized_event'])
        out = pd.DataFrame({
            'country_code': df[country_col].fillna('').astype(str).str.strip().str.upper(),
            'event_name': df[event_col].fillna('').astype(str).str.strip(),
        })
        out['normalized_event'] = out['event_name'].map(normalize_key_text)
        out = out[(out['country_code'] != '') & (out['normalized_event'] != '')]
        return out.drop_duplicates().reset_index(drop=True)

    def _country_code_from_value(self, value):
        raw_text = text_or_blank(value, strip=True)
        normalized = normalize_key_text(raw_text)
        if not normalized:
            return ''
        # Build flat lookup cache on first call
        if not hasattr(self, '_alias_lookup_cache'):
            self._alias_lookup_cache = {}
            for code, aliases in self.important_country_aliases.items():
                self._alias_lookup_cache[normalize_key_text(code)] = code
                for item in aliases:
                    self._alias_lookup_cache[normalize_key_text(item)] = code
        if normalized in self._alias_lookup_cache:
            return self._alias_lookup_cache[normalized]
        raw_upper = raw_text.upper()
        return raw_upper if len(raw_upper) <= 4 else ''

    def _row_country_code(self, row):
        if row is None:
            return ''
        category = text_or_blank(row.get('category'), strip=True)
        candidates = []
        if category == 'Holiday':
            candidates = [row.get('ticker'), row.get('country')]
        else:
            candidates = [row.get('country'), row.get('ticker')]
        for candidate in candidates:
            code = self._country_code_from_value(candidate)
            if code:
                return code
        return ''

    def _row_matches_important_rule(self, row):
        if row is None:
            return False
        if text_or_blank(row.get('category'), strip=True) != 'Economic':
            return False
        if self.always_important_rules is None or self.always_important_rules.empty:
            return False
        country_code = self._row_country_code(row)
        if not country_code:
            return False
        rule_df = self.always_important_rules[self.always_important_rules['country_code'] == country_code]
        if rule_df.empty:
            return False
        normalized_title = normalize_key_text(row.get('title'))
        if not normalized_title:
            return False
        for key in rule_df['normalized_event'].tolist():
            if not key:
                continue
            if normalized_title == key:
                return True
            if normalized_title.startswith(key) or key.startswith(normalized_title):
                return True
            if len(key) >= 12 and (key in normalized_title or normalized_title in key):
                return True
        return False

    def is_auto_important_event(self, event_id, row=None):
        row = row if row is not None else self._find_event_row(event_id)
        if row is None:
            return False
        if hasattr(row, 'to_dict'):
            row = row.to_dict()
        return self._row_matches_important_rule(row)

    def is_event_important(self, event_id, row=None):
        state = self.state.get(event_id, {})
        if as_bool(state.get('watch', False)):
            return True
        return self.is_auto_important_event(event_id, row=row)

    def important_source(self, event_id, row=None):
        state = self.state.get(event_id, {})
        if as_bool(state.get('watch', False)):
            return 'User'
        if self.is_auto_important_event(event_id, row=row):
            return 'Key Event Rule'
        return ''

    def _is_separator_value(self, value):
        return isinstance(value, str) and value.startswith(SEPARATOR_PREFIX)

    def _list_separator_label(self, event_date):
        ts = pd.Timestamp(event_date)
        return f'--- {ts.strftime("%Y-%m-%d (%a)")} ' + '=' * 42

    def _row_flagged(self, event_id):
        return as_bool(self.state.get(event_id, {}).get('flagged', False))

    def _row_marker_text(self, event_id, row=None):
        row = row if row is not None else self._find_event_row(event_id)
        important = self.is_event_important(event_id, row=row)
        flagged = self._row_flagged(event_id)
        if important and flagged:
            return '★⚑'
        if important:
            return '★'
        if flagged:
            return '⚑'
        return ''

    def _list_category_code(self, category):
        category_text = text_or_blank(category, strip=True)
        return LIST_CATEGORY_CODES.get(category_text, category_text[:4].upper() or 'EVT')

    def _display_country_code(self, row):
        code = self._row_country_code(row)
        if code:
            return code
        country = text_or_blank(row.get('country'), strip=True)
        if not country:
            return '-'
        if len(country) <= 4:
            return country.upper()
        return country[:4].upper()

    def _list_metrics_text(self, row):
        survey = fmt_metric(row.get('survey'), row.get('scaling_factor'))
        prior = fmt_metric(row.get('prior'), row.get('scaling_factor'))
        actual = fmt_metric(row.get('actual'), row.get('scaling_factor'))
        return f"S {survey} - P {prior} - A {actual}"

    def _list_detail_text(self, row):
        category = text_or_blank(row.get('category'), strip=True)
        if category == 'Economic':
            country = self._display_country_code(row)
            ticker = text_or_blank(row.get('ticker'), strip=True) or '-'
            period = text_or_blank(row.get('period'), strip=True) or '-'
            return f"{country} | {ticker} | {self._list_metrics_text(row)}"
        if category == 'Earnings':
            ticker = text_or_blank(row.get('ticker'), strip=True) or '-'
            market_cap = fmt_market_cap(row.get('market_cap_usd'))
            return f"{ticker} | Mkt Cap {market_cap}"
        if category == 'Commodity':
            ticker = text_or_blank(coalesce(row.get('underlying_ticker'), row.get('ticker')), strip=True) or '-'
            delivery = fmt_value(row.get('delivery_date'))
            price = fmt_value(row.get('future_price'))
            currency = text_or_blank(row.get('currency'), strip=True)
            px_text = f"{price} {currency}".strip() if price != '-' else '-'
            return f"{ticker} | Last trade {delivery} | Px {px_text}"
        market_text = self._display_country_code(row)
        return f"{market_text} | Market Closed"

    def _list_header_label(self, view_name):
        view_name = str(view_name or 'ALL').upper()
        if view_name == 'ECONOMIC':
            cells = [
                display_cell('MARK', 4, default=''),
                display_cell('TIME', 5, default=''),
                display_cell('EVENT', 24, default=''),
                display_cell('CTRY', 4, default=''),
                display_cell('TICKER', 8, default=''),
                display_cell('PER', 6, default=''),
                display_cell('METRICS', 18, default=''),
            ]
        elif view_name == 'EARNINGS':
            cells = [
                display_cell('MARK', 4, default=''),
                display_cell('TIME', 5, default=''),
                display_cell('EVENT', 26, default=''),
                display_cell('TICKER', 10, default=''),
                display_cell('MKT CAP', 14, default=''),
            ]
        elif view_name == 'COMMODITIES':
            cells = [
                display_cell('MARK', 4, default=''),
                display_cell('TIME', 5, default=''),
                display_cell('EVENT', 24, default=''),
                display_cell('TICKER', 10, default=''),
                display_cell('DELV', 10, default=''),
                display_cell('PX', 10, default=''),
            ]
        elif view_name == 'HOLIDAYS':
            cells = [
                display_cell('MARK', 4, default=''),
                display_cell('TIME', 5, default=''),
                display_cell('EVENT', 30, default=''),
                display_cell('MKT', 8, default=''),
                display_cell('STATUS', 12, default=''),
            ]
        else:
            cells = [
                display_cell('MARK', 4, default=''),
                display_cell('TIME', 5, default=''),
                display_cell('TYPE', 5, default=''),
                display_cell('EVENT', 22, default=''),
                display_cell('DETAILS', 34, default=''),
            ]

        return join_display_cells(cells)

    def _row_list_label(self, row, view_name='ALL'):
        event_id = row['event_id']
        marker = self._row_marker_text(event_id, row=row)
        category = text_or_blank(row.get('category'), strip=True) or 'Event'
        event_time = fmt_time_text(row.get('event_time'))
        title = text_or_blank(row.get('title'), strip=True) or 'Untitled'
        view_name = str(view_name or 'ALL').upper()

        if view_name == 'ECONOMIC':
            metrics = self._list_metrics_text(row)
            cells = [
                display_cell(marker, 4, default=''),
                display_cell(event_time, 5),
                display_cell(title, 24),
                display_cell(self._display_country_code(row), 4),
                display_cell(text_or_blank(row.get('ticker'), strip=True) or '-', 8),
                display_cell(text_or_blank(row.get('period'), strip=True) or '-', 6),
                display_cell(metrics, 18),
            ]
            return join_display_cells(cells)

        if view_name == 'EARNINGS':
            cells = [
                display_cell(marker, 4, default=''),
                display_cell(event_time, 5),
                display_cell(title, 26),
                display_cell(text_or_blank(row.get('ticker'), strip=True) or '-', 10),
                display_cell(fmt_market_cap(row.get('market_cap_usd')), 14, align='right'),
            ]
            return join_display_cells(cells)

        if view_name == 'COMMODITIES':
            ticker = text_or_blank(coalesce(row.get('underlying_ticker'), row.get('ticker')), strip=True) or '-'
            delivery = fmt_value(row.get('delivery_date'))
            price = fmt_value(row.get('future_price'))
            cells = [
                display_cell(marker, 4, default=''),
                display_cell(event_time, 5),
                display_cell(title, 24),
                display_cell(ticker, 10),
                display_cell(delivery, 10),
                display_cell(price, 10, align='right'),
            ]
            return join_display_cells(cells)

        if view_name == 'HOLIDAYS':
            cells = [
                display_cell(marker, 4, default=''),
                display_cell(event_time, 5),
                display_cell(title, 30),
                display_cell(self._display_country_code(row), 8),
                display_cell('MARKET CLOSED', 12),
            ]
            return join_display_cells(cells)

        detail_text = self._list_detail_text(row)
        cells = [
            display_cell(marker, 4, default=''),
            display_cell(event_time, 5),
            display_cell(self._list_category_code(category), 5),
            display_cell(title, 22),
            display_cell(detail_text, 34),
        ]
        return join_display_cells(cells)

    def apply_filters(self, _=None):
        try:
            start_ts, end_ts = self._validate_dates()
        except Exception as exc:
            self.log(str(exc))
            return

        self.events = self.compose_events_for_range(start_ts, end_ts)

        categories = set(self.category_filter.value)
        search_text = self.search_box.value.strip().lower()

        df = self.events.copy()

        if categories:
            df = df[df['category'].isin(categories)]

        if self.only_flagged.value:
            df = df[df['event_id'].map(self._row_flagged)]

        if self.only_watch.value:
            important_ids = {eid for eid in df['event_id'] if self.is_event_important(eid)}
            df = df[df['event_id'].isin(important_ids)]

        if search_text:
            searchable = (
                df[['title', 'subtitle', 'country', 'ticker', 'company', 'period']]
                .fillna('')
                .astype(str)
                .agg(' '.join, axis=1)
                .str.lower()
            )
            df = df[searchable.str.contains(search_text, na=False)]

        self.filtered_events = df.reset_index(drop=True)

        valid_months = self._all_months_in_range(start_ts, end_ts)
        if not valid_months:
            self.month_page_start = 0
        else:
            max_start = max(0, len(valid_months) - max(1, int(self.months_per_page.value)))
            self.month_page_start = min(self.month_page_start, max_start)

        self._ensure_selected_day(start_ts, end_ts)
        self.render_dashboard()
        self.refresh_lists()

    def reset_filters(self, _=None):
        self.search_box.value = ''
        self.category_filter.value = ('Economic', 'Earnings', 'Commodity', 'Holiday')
        self.only_flagged.value = False
        self.only_watch.value = False
        self.month_page_start = 0
        self.apply_filters()

    def _ensure_selected_day(self, start_ts, end_ts):
        start_day = start_ts.date()
        end_day = end_ts.date()
        available_days = []
        if not self.events.empty:
            available_days = sorted(self.events['event_date'].dt.date.unique())

        if self.selected_day is None or self.selected_day < start_day or self.selected_day > end_day:
            if available_days:
                self.selected_day = available_days[0]
            else:
                self.selected_day = start_day

    def _all_months_in_range(self, start_ts, end_ts):
        return list(pd.period_range(start_ts.to_period('M'), end_ts.to_period('M'), freq='M'))

    def on_months_per_page_changed(self, change):
        if change.get('new') == change.get('old'):
            return
        self.month_page_start = 0
        self.render_monthly_page()

    def previous_month_pane(self, _=None):
        page_size = max(1, int(self.months_per_page.value))
        self.month_page_start = max(0, self.month_page_start - page_size)
        self.render_monthly_page()

    def next_month_pane(self, _=None):
        try:
            start_ts, end_ts = self._validate_dates()
        except Exception:
            return
        months = self._all_months_in_range(start_ts, end_ts)
        page_size = max(1, int(self.months_per_page.value))
        max_start = max(0, len(months) - page_size)
        self.month_page_start = min(max_start, self.month_page_start + page_size)
        self.render_monthly_page()

    def refresh_lists(self):
        preferred = self.current_event_id

        econ_df = self.filtered_events[self.filtered_events['category'] == 'Economic']
        earnings_df = self.filtered_events[self.filtered_events['category'] == 'Earnings']
        comdty_df = self.filtered_events[self.filtered_events['category'] == 'Commodity']
        holiday_df = self.filtered_events[self.filtered_events['category'] == 'Holiday']
        watch_df = self.filtered_events[
            self.filtered_events['event_id'].map(lambda x: self.is_event_important(x))
        ]

        self._set_list_options(self.all_list, self.filtered_events, preferred, view_name='ALL')
        self._set_list_options(self.econ_list, econ_df, preferred, view_name='ECONOMIC')
        self._set_list_options(self.earnings_list, earnings_df, preferred, view_name='EARNINGS')
        self._set_list_options(self.comdty_list, comdty_df, preferred, view_name='COMMODITIES')
        self._set_list_options(self.holiday_list, holiday_df, preferred, view_name='HOLIDAYS')
        self._set_list_options(self.watch_list, watch_df, preferred, view_name='IMPORTANT')

    def _set_list_options(self, widget, df, preferred=None, view_name='ALL'):
        if df.empty:
            widget.options = [('No events in this view.', None)]
            widget.value = None
            return

        options = []
        last_date = None
        for row in df.to_dict('records'):
            row_date = pd.Timestamp(row['event_date']).normalize()
            if last_date is None or row_date != last_date:
                sep_value = f"{SEPARATOR_PREFIX}DATE|{view_name}|{row_date.strftime('%Y-%m-%d')}"
                header_value = f"{SEPARATOR_PREFIX}HEAD|{view_name}|{row_date.strftime('%Y-%m-%d')}"
                options.append((self._list_separator_label(row_date), sep_value))
                options.append((self._list_header_label(view_name), header_value))
                last_date = row_date
            options.append((self._row_list_label(row, view_name=view_name), row['event_id']))

        widget.options = options
        values = [value for _, value in options if value is not None and not self._is_separator_value(value)]
        widget.value = preferred if preferred in values else (values[0] if values else None)

    def get_active_list_widget(self):
        idx = self.tabs.selected_index or 0
        return [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list][idx]

    def on_tab_change(self, change):
        widget = self.get_active_list_widget()
        values = [
            value for _, value in widget.options
            if value is not None and not self._is_separator_value(value)
        ]
        if self.current_event_id in values:
            widget.value = self.current_event_id
        elif values:
            widget.value = values[0]
        else:
            if self.current_event_id:
                self.render_detail(self.current_event_id)
            else:
                self.render_detail(None)

    def on_event_selected(self, change):
        new_value = change['new']
        if new_value is None or self._is_separator_value(new_value):
            if self.current_event_id:
                self.render_detail(self.current_event_id)
            else:
                self.render_detail(None)
            return
        self.render_detail(new_value)

    def render_dashboard(self):
        df = self.filtered_events
        total = len(df)
        event_days = int(df['event_date'].dt.normalize().nunique()) if not df.empty else 0
        econ = len(df[df["category"] == "Economic"]) if not df.empty else 0
        earnings = len(df[df["category"] == "Earnings"]) if not df.empty else 0
        comdty = len(df[df["category"] == "Commodity"]) if not df.empty else 0
        holidays = len(df[df["category"] == "Holiday"]) if not df.empty else 0
        flagged = int(sum(self._row_flagged(eid) for eid in df.get("event_id", [])))
        watch = int(sum(self.is_event_important(eid) for eid in df.get("event_id", [])))

        summary_html = f"""
        <div class="event-app">
        <div class="summary-grid">
        <div class="summary-item"><div class="summary-label">Visible Events</div><div class="summary-value">{total}</div></div>
        <div class="summary-item"><div class="summary-label">Event Days</div><div class="summary-value">{event_days}</div></div>
        <div class="summary-item"><div class="summary-label">Economic</div><div class="summary-value">{econ}</div></div>
        <div class="summary-item"><div class="summary-label">Earnings</div><div class="summary-value">{earnings}</div></div>
        <div class="summary-item"><div class="summary-label">Commodities</div><div class="summary-value">{comdty}</div></div>
        <div class="summary-item"><div class="summary-label">Holidays</div><div class="summary-value">{holidays}</div></div>
        <div class="summary-item"><div class="summary-label">Flagged</div><div class="summary-value">{flagged}</div></div>
        <div class="summary-item"><div class="summary-label">Important</div><div class="summary-value">{watch}</div></div>
        </div>
        </div>
        """
        with self.summary_out:
            clear_output(wait=True)
            display(HTML(summary_html))

        self.render_monthly_page()
        self.render_selected_day_detail()

    def _daily_summary_maps(self, df):
        if df is None or df.empty:
            empty = {}
            return {
                "counts": empty,
                "flagged": empty,
                "watch": empty,
                "econ": empty,
                "earnings": empty,
                "commodity": empty,
                "holiday": empty,
                "holiday_codes": empty,
            }

        dates = df["event_date"].dt.date
        counts = df.groupby(dates).size().to_dict()

        flagged_subset = df[df["event_id"].map(self._row_flagged)]
        flagged_counts = flagged_subset.groupby(flagged_subset["event_date"].dt.date).size().to_dict() if not flagged_subset.empty else {}

        watch_subset = df[df["event_id"].map(lambda x: self.is_event_important(x))]
        watch_counts = watch_subset.groupby(watch_subset["event_date"].dt.date).size().to_dict() if not watch_subset.empty else {}

        econ_subset = df[df["category"] == "Economic"]
        earnings_subset = df[df["category"] == "Earnings"]
        commodity_subset = df[df["category"] == "Commodity"]
        holiday_subset = df[df["category"] == "Holiday"]

        holiday_codes = {}
        if not holiday_subset.empty:
            holiday_codes = (
                holiday_subset.groupby(holiday_subset["event_date"].dt.date)["ticker"]
                .apply(lambda s: sorted(pd.Series(s).dropna().astype(str).unique().tolist()))
                .to_dict()
            )

        return {
            "counts": counts,
            "flagged": flagged_counts,
            "watch": watch_counts,
            "econ": econ_subset.groupby(econ_subset["event_date"].dt.date).size().to_dict() if not econ_subset.empty else {},
            "earnings": earnings_subset.groupby(earnings_subset["event_date"].dt.date).size().to_dict() if not earnings_subset.empty else {},
            "commodity": commodity_subset.groupby(commodity_subset["event_date"].dt.date).size().to_dict() if not commodity_subset.empty else {},
            "holiday": holiday_subset.groupby(holiday_subset["event_date"].dt.date).size().to_dict() if not holiday_subset.empty else {},
            "holiday_codes": holiday_codes,
        }

    def render_monthly_page(self):
        try:
            start_ts, end_ts = self._validate_dates()
        except Exception:
            return

        months = self._all_months_in_range(start_ts, end_ts)
        page_size = max(1, int(self.months_per_page.value or 1))

        if not months:
            self.month_page_start = 0
            self.prev_months_button.disabled = True
            self.next_months_button.disabled = True
            self.month_page_label.value = '<div class="event-app page-indicator">No months in range.</div>'
            with self.monthly_out:
                clear_output(wait=True)
                display(HTML('<div class="event-app empty-state">No monthly range available.</div>'))
            return

        max_start = max(0, len(months) - page_size)
        self.month_page_start = min(max(self.month_page_start, 0), max_start)
        visible_months = months[self.month_page_start:self.month_page_start + page_size]

        page_number = (self.month_page_start // page_size) + 1
        total_pages = max(1, math.ceil(len(months) / page_size))
        self.month_page_label.value = (
            f'<div class="event-app page-indicator">Page {page_number} of {total_pages} '
            f'- {visible_months[0].strftime("%b %Y")} to {visible_months[-1].strftime("%b %Y")}</div>'
        )

        self.prev_months_button.disabled = self.month_page_start == 0
        self.next_months_button.disabled = self.month_page_start + page_size >= len(months)

        summary = self._daily_summary_maps(self.events)
        month_widgets = [
            self._build_single_month_widget(period, start_ts.date(), end_ts.date(), summary)
            for period in visible_months
        ]

        grid = widgets.GridBox(
            month_widgets,
            layout=widgets.Layout(
                width="100%",
                grid_template_columns=f"repeat({len(month_widgets)}, minmax(260px, 1fr))",
                gap="12px",
            ),
        )

        with self.monthly_out:
            clear_output(wait=True)
            display(grid)

    def _build_single_month_widget(self, period, start_day, end_day, summary):
        cal = calendar.Calendar(firstweekday=0)
        header = widgets.HTML(
            f'<div class="event-app" style="font-weight:700;font-size:15px;margin-bottom:5px">{period.strftime("%B %Y")}</div>'
        )

        weekday_labels = [widgets.HTML(f'<div class="event-app weekday-label">{label}</div>') for label in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]]
        cells = weekday_labels[:]

        for week in cal.monthdatescalendar(period.year, period.month):
            for day in week:
                cells.append(self._build_day_box(day, period.month, start_day, end_day, summary))

        month_grid = widgets.GridBox(
            cells,
            layout=widgets.Layout(
                width="100%",
                grid_template_columns="repeat(7, minmax(0, 1fr))",
                gap="6px",
            ),
        )

        card = widgets.VBox([header, month_grid])
        card.add_class("app-panel")
        return card

    def _build_day_box(self, day, current_month, start_day, end_day, summary):
        in_range = start_day <= day <= end_day
        in_month = day.month == current_month
        total = summary["counts"].get(day, 0)
        econ = summary["econ"].get(day, 0)
        ern = summary["earnings"].get(day, 0)
        cmd = summary["commodity"].get(day, 0)
        hol = summary["holiday"].get(day, 0)
        hol_codes = summary["holiday_codes"].get(day, [])

        day_button = widgets.Button(
            description=str(day.day),
            layout=widgets.Layout(width="100%"),
            disabled=not in_range,
            tooltip=f"{day.isoformat()}"
        )
        day_button.add_class("calendar-day-btn")
        if in_range:
            day_button.on_click(lambda _, selected_day=day: self.on_day_clicked(selected_day))

        meta_lines = []
        if total:
            meta_lines.append(f"<strong>{total} evt</strong>")
        bucket_parts = []
        if econ:
            bucket_parts.append(f"ECO {econ}")
        if ern:
            bucket_parts.append(f"ERN {ern}")
        if cmd:
            bucket_parts.append(f"CMD {cmd}")
        if hol:
            bucket_parts.append(f"HOL {hol}")
        if bucket_parts:
            meta_lines.append(" | ".join(bucket_parts))

        chip_text = ""
        if hol_codes:
            shown = hol_codes[:4]
            chips = " ".join(f'<span class="holiday-chip">{safe_html(code)}</span>' for code in shown)
            if len(hol_codes) > 4:
                chips += f'<span class="holiday-chip">+{len(hol_codes) - 4}</span>'
            chip_text = chips

        meta_html = widgets.HTML(
            f'<div class="event-app day-meta">{"<br>".join(meta_lines)}{"<div>" + chip_text + "</div>" if chip_text else ""}</div>'
        )

        box = widgets.VBox([day_button, meta_html], layout=widgets.Layout(width="100%"))
        box.add_class("calendar-day-box")
        if not in_month:
            box.add_class("out-month")
        if total:
            box.add_class("has-events")
        if summary["holiday"].get(day, 0):
            box.add_class("holiday-day")
        if summary["watch"].get(day, 0):
            box.add_class("watch-day")
        if summary["flagged"].get(day, 0):
            box.add_class("flagged-day")
        if self.selected_day == day:
            box.add_class("selected-day")
        if not in_range:
            box.add_class("no-click")
        return box

    def on_day_clicked(self, selected_day):
        self.selected_day = pd.Timestamp(selected_day).date()

        day_events = self.events[self.events["event_date"].dt.date == self.selected_day]
        if not day_events.empty:
            first_event_id = day_events.iloc[0]["event_id"]
            self.current_event_id = first_event_id
            if not self.filtered_events.empty and first_event_id in set(self.filtered_events["event_id"]):
                self.tabs.selected_index = 0
            try:
                self.all_list.value = first_event_id
            except Exception:
                self.render_detail(first_event_id)
            else:
                self.render_detail(first_event_id)
        else:
            self.render_detail(None)
            self.render_monthly_page()
            self.render_selected_day_detail()

    def render_selected_day_detail(self):
        if self.selected_day is None:
            with self.day_out:
                clear_output(wait=True)
                display(HTML('<div class="event-app day-scroll-box"><div class="empty-state">Click a day to inspect the full state.</div></div>'))
            return

        day_df = self.events[self.events["event_date"].dt.date == self.selected_day].copy()
        if day_df.empty:
            html = (
                f'<div class="event-app day-scroll-box">'
                f'<div class="day-panel-header"><div class="day-panel-title">{safe_html(self.selected_day)}</div></div>'
                f'<div class="day-panel-subtitle">No events or holiday markers for this day.</div></div>'
                f'<div class="empty-state">No events in the selected range for this date.</div>'
            )
            with self.day_out:
                clear_output(wait=True)
                display(HTML(html))
            return

        day_df["category_order"] = day_df["category"].map(CATEGORY_ORDER).fillna(99)
        day_df["sort_time"] = day_df["event_time"].fillna("99:99")
        day_df = day_df.sort_values(["category_order", "sort_time", "title"]).drop(columns=["category_order", "sort_time"])

        holiday_codes = day_df.loc[day_df["category"] == "Holiday", "ticker"].dropna().astype(str).unique().tolist()
        header_subtitle = f"{len(day_df)} total events"
        if holiday_codes:
            header_subtitle += " - Markets closed: " + ", ".join(holiday_codes)

        cards = []
        for row in day_df.to_dict("records"):
            state = self.state.get(row["event_id"], {})
            fields = self._event_fields(row)
            fields_html = "".join(
                f'<div class="day-card-field"><div class="day-card-field-label">{safe_html(label)}</div><div class="day-card-field-'
                f'value">{safe_html(value)}</div></div>'
                for label, value in fields
            )

            note_text = text_or_blank(state.get("note"), strip=True)
            note_html = f'<div class="note-preview"><b>Saved note</b><br>{safe_html(note_text)}</div>' if note_text else ""

            subtitle = text_or_blank(row.get("subtitle"), strip=True)
            event_time = text_or_blank(row.get("event_time"), strip=True)
            if event_time and event_time not in subtitle:
                subtitle = f"{subtitle} | {event_time}" if subtitle else event_time

            cards.append(
                f'<div class="day-card">'
                f'{self._event_badges_html(row, state)}'
                f'<div class="day-card-title">{safe_html(row["title"])}</div>'
                f'<div class="day-card-subtitle">{safe_html(subtitle)}</div>'
                f'<div class="day-card-grid">{fields_html}</div>'
                f'{note_html}'
                f'</div>'
            )

        html = (
            f'<div class="event-app day-scroll-box">'
            f'<div class="day-panel-header">'
            f'<div class="day-panel-title">{safe_html(self.selected_day)}</div>'
            f'<div class="day-panel-subtitle">{safe_html(header_subtitle)}</div>'
            f'</div>'
            + "".join(cards) +
            f'</div>'
        )

        with self.day_out:
            clear_output(wait=True)
            display(HTML(html))

    def render_detail(self, event_id):
        self.current_event_id = event_id
        if not event_id:
            self.detail_html.value = (
                '<div class="event-app"><div class="empty-state">'
                "'Select an event to view details, notes, flags, and important status.'"
                '</div></div>'
            )

            self._suspend_state_events = True
            self.note_area.value = ""
            self.flag_toggle.value = False
            self.watch_toggle.value = False
            self.flag_toggle.icon = "flag-o"
            self.watch_toggle.icon = "star-o"
            self.watch_toggle.disabled = False
            self.watch_toggle.tooltip = ""
            self._suspend_state_events = False
            return

        base_df = self.events if not self.events.empty else self.filtered_events
        row_df = base_df[base_df["event_id"] == event_id]
        if row_df.empty:
            self.detail_html.value = '<div class="event-app"><div class="empty-state">Selected event is no longer in the active date range.</div></div>'
            return

        row = row_df.iloc[0]
        row_dict = row.to_dict()
        if event_id in self.state:
            self._sync_state_metadata(event_id, row=row, touch_timestamp=False)
        state = self.state.get(event_id, {})
        auto_important = self.is_auto_important_event(event_id, row=row_dict)
        important_value = self.is_event_important(event_id, row=row_dict)

        self._suspend_state_events = True
        self.flag_toggle.value = self._row_flagged(event_id)
        self.watch_toggle.value = important_value
        self.note_area.value = state.get("note", "")
        self.flag_toggle.icon = "flag" if self.flag_toggle.value else "flag-o"
        self.watch_toggle.icon = "star" if self.watch_toggle.value else "star-o"
        self.watch_toggle.disabled = auto_important
        self.watch_toggle.tooltip = "Always important: matched embedded key-event rule." if auto_important else ""
        self._suspend_state_events = False

        badge_html = self._event_badges_html(row_dict, state)
        subtitle = text_or_blank(row.get("subtitle"), strip=True)
        event_time = text_or_blank(row.get("event_time"), strip=True)
        if event_time and event_time not in subtitle:
            subtitle = f"{subtitle} | {event_time}" if subtitle else event_time

        fields = self._event_fields(row_dict)
        field_html = "".join(
            f'<div class="field"><div class="field-label">{safe_html(label)}</div><div class="field-value">{safe_html(value)}</div></div>'
            for label, value in fields
        )

        note_preview = state.get("note", "").strip()
        note_html = ""
        if note_preview:
            note_html = f'<div class="note-preview"><b>Saved note</b><br>{safe_html(note_preview)}</div>'
        important_note = ""
        if auto_important:
            important_note = '<div class="note-preview"><b>Rule-based important event</b><br>This release match the embedded key-event list and is always highlighted as important.</div>'

        self.detail_html.value = f"""
        <div class="event-app">
        <div>{badge_html}</div>
        <div class="detail-title">{safe_html(row["title"])}</div>
        <div class="detail-subtitle">{safe_html(subtitle)}</div>
        <div class="field-grid">{field_html}</div>
        {important_note}
        {note_html}
        </div>
        """

    def save_note(self, _=None):
        if not self.current_event_id:
            self.log("Select an event before saving a note.")
            return

        state = self.state.setdefault(self.current_event_id, {})
        state["note"] = self.note_area.value.strip()
        self._sync_state_metadata(self.current_event_id)
        state["updated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")
        self._prune_state_entry(self.current_event_id)
        self._save_state_to_disk()
        self.render_detail(self.current_event_id)
        self.render_watchlist_box()
        self.render_monthly_page()
        self.render_selected_day_detail()
        self.log("Note saved and synced to the annotations CSV")

    def clear_note(self, _=None):
        if not self.current_event_id:
            self.log("Select an event before clearing a note.")
            return

        state = self.state.setdefault(self.current_event_id, {})
        state["note"] = ""
        self._sync_state_metadata(self.current_event_id)
        state["updated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")
        self.note_area.value = ""
        self._prune_state_entry(self.current_event_id)
        self._save_state_to_disk()
        self.render_detail(self.current_event_id)
        self.render_watchlist_box()
        self.render_selected_day_detail()
        self.render_monthly_page()
        self.log("Note cleared.")

    def on_flag_toggled(self, change):
        if self._suspend_state_events or self.current_event_id is None:
            return
        state = self.state.setdefault(self.current_event_id, {})
        state["flagged"] = bool(change["new"])
        self.flag_toggle.icon = "flag" if change["new"] else "flag-o"
        self._sync_state_metadata(self.current_event_id)
        state["updated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")
        self._prune_state_entry(self.current_event_id)
        self._save_state_to_disk()
        self.render_detail(self.current_event_id)
        self.render_dashboard()
        self.refresh_lists()

    def on_watch_toggled(self, change):
        if self._suspend_state_events or self.current_event_id is None:
            return
        if self.is_auto_important_event(self.current_event_id):
            self._suspend_state_events = True
            self.watch_toggle.value = True
            self.watch_toggle.icon = "star"
            self._suspend_state_events = False
            self.log("This event is auto-marked important by the embedded key-event rules.")
            return
        state = self.state.setdefault(self.current_event_id, {})
        state["watch"] = bool(change["new"])
        self.watch_toggle.icon = "star" if change["new"] else "star-o"
        self._sync_state_metadata(self.current_event_id)
        state["updated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")
        self._prune_state_entry(self.current_event_id)
        self._save_state_to_disk()
        self.render_detail(self.current_event_id)
        self.render_dashboard()
        self.refresh_lists()

    def export_state(self, _=None):
        self._save_state_to_disk()
        self.log(f"Annotations synced to {self.state_path.resolve()}")

    def load_state(self, _=None):
        self._load_state_from_disk(silent=False)
        self.apply_filters()
        if self.current_event_id:
            self.render_detail(self.current_event_id)

    def _find_event_row(self, event_id):
        for df in [self.filtered_events, self.events, self.core_events, self.all_holiday_events]:
            if df is None or df.empty or "event_id" not in df.columns:
                continue
            match = df[df["event_id"] == event_id]
            if not match.empty:
                return match.iloc[0]
        return None

    def _sync_state_metadata(self, event_id, row=None, touch_timestamp=True):
        if not event_id:
            return
        row = row if row is not None else self._find_event_row(event_id)
        if row is None:
            return
        state = self.state.setdefault(event_id, {})
        metadata_keys = [
            "event_date", "event_time", "category", "title", "subtitle", "country", "ticker",
            "company", "period", "survey", "prior", "actual", "revision", "scaling_factor"
        ]
        for key in metadata_keys:
            value = row[key] if key in row.index else None
            if key == "event_date":
                state[key] = "" if pd.isna(value) else pd.Timestamp(value).strftime("%Y-%m-%d")
            elif key == "event_time":
                state[key] = text_or_blank(value, strip=True)
            else:
                if pd.isna(value):
                    state[key] = ""
                elif isinstance(value, (pd.Timestamp, datetime)):
                    state[key] = value.strftime("%Y-%m-%d")
                elif isinstance(value, date):
                    state[key] = value.isoformat()
                else:
                    state[key] = str(value)
        if touch_timestamp:
            state["updated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")

    def _prune_state_entry(self, event_id):
        value = self.state.get(event_id)
        if value is None:
            return
        note_text = text_or_blank(value.get("note"), strip=True)
        if not as_bool(value.get("flagged")) and not as_bool(value.get("watch")) and not note_text:
            self.state.pop(event_id, None)

    def _state_metadata_events(self):
        rows = []
        for event_id, value in self.state.items():
            event_date = pd.to_datetime(value.get("event_date"), errors="coerce")
            if pd.isna(event_date):
                continue
            rows.append({
                "event_id": event_id,
                "event_date": event_date,
                "event_time": value.get("event_time") or None,
                "category": value.get("category") or None,
                "title": value.get("title") or None,
                "subtitle": value.get("subtitle") or None,
                "country": value.get("country") or None,
                "ticker": value.get("ticker") or None,
                "company": value.get("company") or None,
                "period": value.get("period") or None,
                "survey": value.get("survey") or None,
                "actual": value.get("actual") or None,
                "prior": value.get("prior") or None,
                "revision": value.get("revision") or None,
                "scaling_factor": value.get("scaling_factor") or None,
                "market_cap_usd": None,
                "future_price": None,
                "currency": None,
                "underlying_ticker": None,
                "delivery_date": None,
                "expire_group_id": None,
                "orig_ids": None,
            })
        if not rows:
            return pd.DataFrame(columns=self.master_columns)
        out = pd.DataFrame(rows)
        for col in self.master_columns:
            if col not in out.columns:
                out[col] = None
        return out[self.master_columns].drop_duplicates(subset=["event_id"])

    def _current_pane_bounds(self):
        start_ts, end_ts = self._validate_dates()
        months = self._all_months_in_range(start_ts, end_ts)
        if not months:
            return start_ts, end_ts
        page_size = max(1, int(self.months_per_page.value or 1))
        max_start = max(0, len(months) - page_size)
        pane_start_idx = min(max(0, self.month_page_start), max_start)
        visible_months = months[pane_start_idx:pane_start_idx + page_size]
        pane_start = visible_months[0].to_timestamp(how="start").normalize()
        pane_end = visible_months[-1].to_timestamp(how="end").normalize()
        pane_start = max(pane_start, start_ts)
        pane_end = min(pane_end, end_ts)
        return pane_start, pane_end

    def _export_view_frame(self, scope=None):
        scope = scope or self.export_scope_dropdown.value
        if self.filtered_events is None or self.filtered_events.empty:
            return pd.DataFrame(columns=self.master_columns), None, None, scope

        if scope == "range":
            start_ts, end_ts = self._validate_dates()
        else:
            start_ts, end_ts = self._current_pane_bounds()

        df = self.filtered_events.copy()
        df = df[(df["event_date"] >= start_ts.normalize()) & (df["event_date"] <= end_ts.normalize())]

        if scope == "highlights":
            df = df[df["event_id"].map(lambda x: self._row_flagged(x) or self.is_event_important(x))]

        if df.empty:
            return df.reset_index(drop=True), start_ts, end_ts, scope

        df["category_order"] = df["category"].map(CATEGORY_ORDER).fillna(99)
        df["sort_time"] = df["event_time"].fillna("99:99")
        df = df.sort_values(["event_date", "category_order", "sort_time", "title"]).drop(columns=["category_order", "sort_time"])
        return df.reset_index(drop=True), start_ts, end_ts, scope

    def _export_snapshot_dataframe(self, df):
        rows = []
        for row in df.to_dict("records"):
            state = self.state.get(row["event_id"], {})
            important = self.is_event_important(row["event_id"], row=row)
            flagged = self._row_flagged(row["event_id"])
            important_source = self.important_source(row["event_id"], row=row)

            scale_value = text_or_blank(row.get("scaling_factor"), strip=True)
            category = text_or_blank(row.get("category"), strip=True)
            event_title = text_or_blank(row.get("title"), strip=True)
            event_time = fmt_time_text(row.get("event_time"))
            country = text_or_blank(row.get("country"), strip=True)
            ticker = text_or_blank(row.get("ticker"), strip=True)
            period = text_or_blank(row.get("period"), strip=True)
            subtitle = text_or_blank(row.get("subtitle"), strip=True)

            if category == "Economic":
                market_value = country or ticker or "--"
                period_detail = period or "--"
                survey = fmt_metric(row.get("survey"), scale_value)
                prior = fmt_metric(row.get("prior"), scale_value)
                actual = fmt_metric(row.get("actual"), scale_value)
                revision = fmt_metric(row.get("revision"), scale_value)
            elif category == "Earnings":
                market_value = ticker or "--"
                period_detail = fmt_market_cap(row.get("market_cap_usd"))
                survey = prior = actual = revision = "--"
            elif category == "Commodity":
                ticker = text_or_blank(coalesce(row.get("underlying_ticker"), row.get("ticker")), strip=True)
                market_value = ticker or "--"
                period_detail = fmt_value(row.get("delivery_date"))
                survey = prior = actual = revision = "--"
            else:
                market_value = text_or_blank(row.get("ticker"), row.get("country")) or text_or_blank(row.get("country"), strip=True) or "--"
                period_detail = "Market Closed"
                survey = prior = actual = revision = "--"

            note_text = text_or_blank(state.get("note"), strip=True).replace("\n", " | ")
            if len(note_text) > 280:
                note_text = note_text[:277] + "..."

            status_tags = []
            if important:
                status_tags.append("IMPORTANT")
            if flagged:
                status_tags.append("FLAGGED")
            if category == "Holiday":
                status_tags.append("CLOSED")
            status_text = " / ".join(status_tags) if status_tags else "--"
            notes_and_tags = " | ".join(status_tags + ([note_text] if note_text else [])) if (status_tags or note_text) else "--"

            rows.append({
                "Event ID": row["event_id"],
                "Date": pd.Timestamp(row["event_date"]).strftime("%Y-%m-%d"),
                "Day": pd.Timestamp(row["event_date"]).strftime("%a"),
                "Time": event_time,
                "Category": category,
                "Event": event_title,
                "Country": country or "--",
                "Ticker": ticker or "--",
                "Market / Ticker": market_value,
                "Period": period or ("Market Closed" if category == "Holiday" else "--"),
                "Period / Detail": period_detail,
                "Survey / Expected": survey,
                "Survey": survey,
                "Prior": prior,
                "Actual": actual,
                "Revision": revision,
                "Scaling": scale_value or "--",
                "Status": status_text,
                "Important": bool(important),
                "Flagged": bool(flagged),
                "Important Source": important_source,
                "Subtitle": subtitle or "--",
                "Notes": note_text or "--",
                "Notes / Tags": notes_and_tags,
            })
        if not rows:
            return pd.DataFrame(columns=[
                "Event ID", "Date", "Day", "Time", "Category", "Event", "Country", "Ticker",
                "Market / Ticker", "Period", "Period / Detail", "Survey / Expected", "Survey",
                "Prior", "Actual", "Revision", "Scaling", "Status", "Important", "Flagged",
                "Important Source", "Subtitle", "Notes", "Notes / Tags"
            ])
        return pd.DataFrame(rows)

    def _build_export_excel(self, export_df, xlsx_path, start_ts, end_ts, scope):
        if Workbook is None:
            raise RuntimeError("openpyxl is not available in this environment!")

        scope_label_map = {
            "pane": "Current visible pane",
            "range": "Full selected range",
            "highlights": "Important / flagged only",
        }

        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        filters_text = f"Categories: {', '.join(self.category_filter.value) if self.category_filter.value else 'ALL'}"
        search_text = self.search_box.value.strip()
        if search_text:
            filters_text += f" | Search: {search_text}"

        counts = export_df["Category"].value_counts().to_dict() if not export_df.empty else {}
        summary_text = (
            f"{len(export_df):} events | "
            f"Economic {counts.get('Economic', 0):} | "
            f"Earnings {counts.get('Earnings', 0):} | "
            f"Commodity {counts.get('Commodity', 0):} | "
            f"Holiday {counts.get('Holiday', 0):} | "
            f"Important {int(export_df['Important'].sum()) if 'Important' in export_df else 0:} | "
            f"Flagged {int(export_df['Flagged'].sum()) if 'Flagged' in export_df else 0:}"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Print Brief"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A8"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        try:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        except Exception:
            pass

        if PageMargins is not None:
            ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.55, bottom=0.45, header=0.20, footer=0.20)
        ws.oddFooter.left.text = f"&8{scope_label_map.get(scope, scope)}"
        ws.oddFooter.right.text = "&8Page &P of &N"

        brief_columns = [
            ("Date", 12, "center"),
            ("Day", 8, "center"),
            ("Time", 10, "center"),
            ("Category", 14, "center"),
            ("Event", 38, "left"),
            ("Market / Ticker", 18, "left"),
            ("Period / Detail", 18, "left"),
            ("Survey", 12, "right"),
            ("Prior", 12, "right"),
            ("Actual", 12, "right"),
            ("Revision", 12, "right"),
            ("Status", 18, "center"),
            ("Notes", 36, "left"),
        ]
        last_col = len(brief_columns)
        last_col_letter = get_column_letter(last_col)

        title_fill = PatternFill("solid", fgColor="102A43")
        info_fill = PatternFill("solid", fgColor="EEF3F8")
        header_fill = PatternFill("solid", fgColor="1F3A5F")
        month_fill = PatternFill("solid", fgColor="D9E2F3")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        stripe_fill = PatternFill("solid", fgColor="F7FAFC")
        important_fill = PatternFill("solid", fgColor="FFF2CC")
        flagged_fill = PatternFill("solid", fgColor="FDE9E7")
        both_fill = PatternFill("solid", fgColor="FCE4D6")
        status_important_fill = PatternFill("solid", fgColor="FFD667")
        status_flagged_fill = PatternFill("solid", fgColor="F4CCCC")
        status_both_fill = PatternFill("solid", fgColor="FAB183")
        status_closed_fill = PatternFill("solid", fgColor="FCE4D6")

        title_font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
        subtitle_font = Font(name="Arial", size=9, color="4A5568")
        header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        month_font = Font(name="Arial", size=10, bold=True, color="102A43")
        body_font = Font(name="Arial", size=9, color="111827")
        status_font = Font(name="Arial", size=9, bold=True, color="102A43")

        title_border = Border(bottom=Side(style="medium", color="102A43"))
        info_border = Border(bottom=Side(style="thin", color="DBE1EB"))
        header_border = Border(top=Side(style="medium", color="102A43"), bottom=Side(style="medium", color="102A43"))
        month_border = Border(top=Side(style="medium", color="102A43"), bottom=Side(style="thin", color="DBE1EB"))
        body_border = Border(bottom=Side(style="thin", color="E2E8F0"))

        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        category_font_map = {
            "Economic": Font(name="Arial", size=9, bold=True, color="0F766E"),
            "Earnings": Font(name="Arial", size=9, bold=True, color="6D28D9"),
            "Commodity": Font(name="Arial", size=9, bold=True, color="BE123C"),
            "Holiday": Font(name="Arial", size=9, bold=True, color="B45309"),
        }

        def write_cell(cell, value, font=None, fill=None, alignment=None, border=None):
            cell.value = value
            if font is not None:
                cell.font = font
            if fill is not None:
                cell.fill = fill
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border

        def row_fill(record, striped=False):
            if bool(record.get("Important")) and bool(record.get("Flagged")):
                return both_fill
            if bool(record.get("Important")):
                return important_fill
            if bool(record.get("Flagged")):
                return flagged_fill
            return stripe_fill if striped else white_fill

        def status_fill(record):
            status_text = text_or_blank(record.get("Status"), strip=True)
            if bool(record.get("Important")) and bool(record.get("Flagged")):
                return status_both_fill
            if bool(record.get("Important")):
                return status_important_fill
            if bool(record.get("Flagged")):
                return status_flagged_fill
            if "CLOSED" in status_text:
                return status_closed_fill
            return info_fill

        for idx, (label, width, _) in enumerate(brief_columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        write_cell(ws["A1"], "MARKET EVENT BRIEF", font=title_font, fill=title_fill, alignment=align_left, border=title_border)
        ws.row_dimensions[1].height = 26

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        write_cell(
            ws["A2"],
            f"{scope_label_map.get(scope, scope)} | {start_ts.date()} to {end_ts.date()} | Generated {generated_at}",
            font=subtitle_font,
            fill=info_fill,
            alignment=align_left,
            border=info_border,
        )
        ws.row_dimensions[2].height = 18

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
        write_cell(ws["A3"], filters_text, font=subtitle_font, fill=info_fill, alignment=align_left, border=info_border)
        ws.row_dimensions[3].height = 18

        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
        write_cell(ws["A4"], summary_text, font=Font(name="Arial", size=9, bold=True, color="102A43"), fill=info_fill, alignment=align_left, border=info_border)
        ws.row_dimensions[4].height = 18

        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=last_col)
        write_cell(
            ws["A5"],
            "Highlight legend: gold = important, rose = flagged, amber = both important and flagged, orange status = market closed.",
            font=subtitle_font,
            fill=info_fill,
            alignment=align_left,
            border=info_border,
        )
        ws.row_dimensions[5].height = 18

        header_row = 7
        for idx, (label, _, align) in enumerate(brief_columns, start=1):
            write_cell(
                ws.cell(header_row, idx),
                label.upper(),
                font=header_font,
                fill=header_fill,
                alignment={"left": align_left, "center": align_center, "right": align_right}[align],
                border=header_border,
            )
        ws.row_dimensions[header_row].height = 22

        work_df = export_df.copy()
        work_df["_parsed_date"] = pd.to_datetime(work_df["Date"], errors="coerce")
        work_df["Day"] = work_df["_parsed_date"].dt.strftime("%a").fillna("")
        if "Notes" not in work_df.columns:
            work_df["Notes"] = work_df.get("Notes / Tags", "--")
        if "Status" not in work_df.columns:
            work_df["Status"] = "--"

        row_cursor = header_row + 1
        current_month = None
        striped = False

        for record in work_df.to_dict("records"):
            month_label = record["_parsed_date"].strftime("%B %Y").upper() if pd.notna(record["_parsed_date"]) else "UNSCHEDULED"
            if month_label != current_month:
                ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=last_col)
                write_cell(ws.cell(row_cursor, 1), month_label, font=month_font, fill=month_fill, alignment=align_left, border=month_border)
                ws.row_dimensions[row_cursor].height = 20
                row_cursor += 1
                current_month = month_label
                striped = False

            fl = row_fill(record, striped)
            striped = not striped
            status_text = text_or_blank(record.get("Status"), strip=True) or "--"
            note_text = text_or_blank(record.get("Notes"), strip=True) or "--"

            values = [
                record.get("Date", ""),
                record.get("Day", ""),
                record.get("Time", ""),
                record.get("Category", ""),
                record.get("Event", ""),
                record.get("Market / Ticker", ""),
                record.get("Period / Detail", ""),
                record.get("Survey", ""),
                record.get("Prior", ""),
                record.get("Actual", ""),
                record.get("Revision", ""),
                status_text,
                note_text,
            ]

            for idx, (label, _, align) in enumerate(brief_columns, start=1):
                font = body_font
                if label == "Category":
                    font = category_font_map.get(text_or_blank(record.get("Category"), strip=True), Font(name="Arial", size=9, bold=True, color="102A43"))
                elif label == "Status":
                    font = status_font

                write_cell(
                    ws.cell(row_cursor, idx),
                    values[idx - 1],
                    font=font,
                    fill=status_fill(record) if label == "Status" else fl,
                    alignment={"left": align_left, "center": align_center, "right": align_right}[align],
                    border=body_border,
                )

            event_length = len(text_or_blank(record.get("Event"), strip=True))
            note_length = len(note_text)
            detail_length = len(text_or_blank(record.get("Period / Detail"), strip=True))
            line_units = max(event_length / 40, note_length / 42, detail_length / 24, 1)
            ws.row_dimensions[row_cursor].height = min(18 * (math.ceil(line_units) - 1) * 10, 68)
            row_cursor += 1

        ws.print_title_rows = f"1:{header_row}"
        ws.print_area = f"A1:{last_col_letter}{row_cursor - 1}"

        snapshot = wb.create_sheet("Data Snapshot")
        snapshot.sheet_view.showGridLines = False
        snapshot.freeze_panes = "A2"
        snapshot.page_setup.orientation = "landscape"
        snapshot.page_setup.fitToWidth = 1
        snapshot.page_setup.fitToHeight = 0
        if PageMargins is not None:
            snapshot.page_margins = PageMargins(left=0.30, right=0.30, top=0.45, bottom=0.35, header=0.20, footer=0.20)

        snapshot_columns = [
            ("Date", 12, "center"),
            ("Time", 10, "center"),
            ("Category", 14, "center"),
            ("Event", 36, "left"),
            ("Market / Ticker", 18, "left"),
            ("Period / Detail", 18, "left"),
            ("Survey", 12, "right"),
            ("Prior", 12, "right"),
            ("Actual", 12, "right"),
            ("Revision", 12, "right"),
            ("Status", 18, "center"),
            ("Important", 11, "center"),
            ("Flagged", 11, "center"),
            ("Notes", 34, "left"),
            ("Event ID", 28, "left"),
        ]

        for idx, (label, width, align) in enumerate(snapshot_columns, start=1):
            snapshot.column_dimensions[get_column_letter(idx)].width = width

            write_cell(
                snapshot.cell(1, idx),
                label.upper(),
                font=header_font,
                fill=header_fill,
                alignment={"left": align_left, "center": align_center, "right": align_right}[align],
                border=header_border,
            )

        snapshot.row_dimensions[1].height = 22

        snap_row = 2
        striped = False
        for record in work_df.to_dict("records"):
            fl = row_fill(record, striped)
            striped = not striped
            status_text = text_or_blank(record.get("Status"), strip=True) or "--"
            note_text = text_or_blank(record.get("Notes"), strip=True) or "--"

            values = [
                record.get("Date", ""),
                record.get("Time", ""),
                record.get("Category", ""),
                record.get("Event", ""),
                record.get("Market / Ticker", ""),
                record.get("Period / Detail", ""),
                record.get("Survey", ""),
                record.get("Prior", ""),
                record.get("Actual", ""),
                record.get("Revision", ""),
                status_text,
                "Y" if bool(record.get("Important")) else "",
                "Y" if bool(record.get("Flagged")) else "",
                note_text,
                record.get("Event ID", ""),
            ]

            for idx, (label, _, align) in enumerate(snapshot_columns, start=1):
                font = body_font
                if label == "Category":
                    font = category_font_map.get(text_or_blank(record.get("Category"), strip=True), Font(name="Arial", size=9, bold=True, color="102A43"))
                elif label == "Status":
                    font = status_font

                write_cell(
                    snapshot.cell(snap_row, idx),
                    values[idx - 1],
                    font=font,
                    fill=status_fill(record) if label == "Status" else fl,
                    alignment={"left": align_left, "center": align_center, "right": align_right}[align],
                    border=body_border,
                )

            snapshot.row_dimensions[snap_row].height = min(18 * (math.ceil(max(len(note_text) / 42, 1)) - 1) * 8, 54)
            snap_row += 1

        snapshot.auto_filter.ref = f"A1:{get_column_letter(len(snapshot_columns))}{max(snap_row - 1, 1)}"
        snapshot.print_title_rows = "1:1"
        snapshot.column_dimensions[get_column_letter(len(snapshot_columns))].hidden = True

        wb.save(xlsx_path)

    def export_current_view_csv(self, _=None):
        self._save_state_to_disk()
        try:
            export_base_df, start_ts, end_ts, scope = self._export_view_frame(self.export_scope_dropdown.value)
        except Exception as exc:
            self.log(f"Could not build export view: {exc}")
            return

        if export_base_df.empty:
            self.log("No events are available in the selected export scope.")
            return

        try:
            export_df = self._export_snapshot_dataframe(export_base_df)
            export_df = export_df.sort_values(["Date", "Time", "Category", "Event"]).reset_index(drop=True)
        except Exception as exc:
            self.log(f"Could not normalize export data: {exc}")
            return

        self.export_dir.mkdir(parents=True, exist_ok=True)
        file_stub = f"market_event_export_{start_ts.strftime('%Y%m%d')}_{end_ts.strftime('%Y%m%d')}_{scope}"
        csv_path = self.export_dir / f"{file_stub}.csv"

        try:
            export_df.to_csv(csv_path, index=False)
            csv_text = export_df.to_csv(index=False)
            self.export_csv_preview.value = csv_text
            self.export_csv_path.value = (
                f'<div class="event-app page-caption">'
                f'CSV saved to <b>{safe_html(str(csv_path.resolve()))}</b>. '
                f'Copy the raw CSV text from the preview box below if direct notebook downloads are unavailable.'
                f'</div>'
            )

            self.log(f"Exported CSV snapshot to {csv_path.resolve()}")
        except Exception as exc:
            self.log(f"Export failed: {exc}")

    def export_current_view_document(self, _=None):
        return self.export_current_view_csv()

    def clear_watchlist(self, _=None):
        for key in list(self.state.keys()):
            self.state[key]["watch"] = False
            self.state[key]["updated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")
            self._prune_state_entry(key)
        self._save_state_to_disk()
        self.apply_filters()
        self.log("Cleared user-selected important markers. Rule-based key events remain important.")

    def _watch_source_events(self):
        frames = []
        if self.core_events is not None and not self.core_events.empty:
            frames.append(self.core_events.copy())
        if self.all_holiday_events is not None and not self.all_holiday_events.empty:
            frames.append(self.all_holiday_events.copy())
        state_events = self._state_metadata_events()
        if state_events is not None and not state_events.empty:
            frames.append(state_events.copy())
        if not frames:
            return pd.DataFrame(columns=self.master_columns)
        df = pd.concat(frames, ignore_index=True)
        df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce")
        df = df.dropna(subset=["event_date"]).sort_values(["event_date", "title"])
        return df.drop_duplicates(subset=["event_id"], keep="first")

    def render_watchlist_box(self):
        base_df = self._watch_source_events()
        if base_df.empty:
            with self.watchlist_out:
                clear_output(wait=True)
                display(HTML('<div class="event-app"><div class="empty-state">No important items yet.</div></div>'))
            return

        watch_df = base_df[base_df["event_id"].map(lambda eid: self.is_event_important(eid))].sort_values(["event_date", "title"])

        if watch_df.empty:
            html = '<div class="event-app"><div class="watchlist-box"><div class="empty-state">No important events selected.</div></div></div>'
        else:
            items = []
            for row in watch_df.to_dict("records"):
                note = text_or_blank(self.state.get(row["event_id"], {}).get("note"), strip=True)
                note_html = f'<div class="day-meta">{safe_html(note[:180])}</div>' if note else ""
                source = self.important_source(row["event_id"], row=row)
                source_html = f'<div class="day-meta">Source: {safe_html(source)}</div>' if source else ""
                items.append(
                    f'<div class="watch-item"><b>{safe_html(row["title"])}</b><br>'
                    f'<span style="color:#94a3b8">{pd.Timestamp(row["event_date"]).date()} | {safe_html(row["category"])}</span>'
                    f'{source_html}{note_html}</div>'
                )
            html = f'<div class="event-app"><div class="watchlist-box">{"".join(items)}</div></div>'

        with self.watchlist_out:
            clear_output(wait=True)
            display(HTML(html))

    def _save_state_to_disk(self):
        self.state_path.parent.mkdir(parents=True, exist_ok=True)

        clean_records = []
        clean_state = {}
        for key in sorted(self.state.keys()):
            value = self.state.get(key, {})
            flagged = bool(value.get("flagged", False))
            watch = bool(value.get("watch", False))
            note = text_or_blank(value.get("note"), strip=True)

            if not flagged and not watch and not note:
                continue

            record = {
                "event_id": key,
                "event_date": value.get("event_date", ""),
                "event_time": value.get("event_time", ""),
                "category": value.get("category", ""),
                "title": value.get("title", ""),
                "subtitle": value.get("subtitle", ""),
                "country": value.get("country", ""),
                "ticker": value.get("ticker", ""),
                "company": value.get("company", ""),
                "period": value.get("period", ""),
                "flagged": flagged,
                "important": watch,
                "note": note,
                "updated_at": value.get("updated_at", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")),
            }
            clean_records.append(record)

            clean_state[key] = {
                "flagged": flagged,
                "watch": watch,
                "note": note,
                "updated_at": record["updated_at"],
                "event_date": record["event_date"],
                "event_time": record["event_time"],
                "category": record["category"],
                "title": record["title"],
                "subtitle": record["subtitle"],
                "country": record["country"],
                "ticker": record["ticker"],
                "company": record["company"],
                "period": record["period"],
            }

        self.state = clean_state
        if self.state_path.suffix.lower() == ".json":
            self.state_path.write_text(json.dumps(clean_state, indent=2))
            return

        pd.DataFrame(clean_records, columns=[
            "event_id", "event_date", "event_time", "category", "title", "subtitle",
            "country", "ticker", "company", "period", "flagged", "important", "note", "updated_at"
        ]).to_csv(self.state_path, index=False)

    def _load_state_from_disk(self, silent=True):
        self.state = {}

        def _log(msg):
            if not silent:
                self.log(msg)

        def _load_csv(csv_path):
            if not csv_path.exists():
                return False
            df = pd.read_csv(csv_path)
            loaded = {}
            for row in df.fillna("").to_dict("records"):
                event_id = str(row.get("event_id", "")).strip()
                if not event_id:
                    continue
                loaded[event_id] = {
                    "flagged": as_bool(row.get("flagged")),
                    "watch": as_bool(row.get("important", row.get("watch", False))),
                    "note": text_or_blank(row.get("note"), strip=True),
                    "updated_at": text_or_blank(row.get("updated_at"), strip=True),
                    "event_date": text_or_blank(row.get("event_date"), strip=True),
                    "event_time": text_or_blank(row.get("event_time"), strip=True),
                    "category": text_or_blank(row.get("category"), strip=True),
                    "title": text_or_blank(row.get("title"), strip=True),
                    "subtitle": text_or_blank(row.get("subtitle"), strip=True),
                    "country": text_or_blank(row.get("country"), strip=True),
                    "ticker": text_or_blank(row.get("ticker"), strip=True),
                    "company": text_or_blank(row.get("company"), strip=True),
                    "period": text_or_blank(row.get("period"), strip=True),
                }
            self.state = loaded
            _log(f"Loaded annotations from {csv_path.resolve()}")
            return True

        def _load_json(json_path):
            if not json_path.exists():
                return False
            raw = json.loads(json_path.read_text())
            loaded = {}
            for event_id, value in raw.items():
                loaded[event_id] = {
                    "flagged": bool(value.get("flagged", False)),
                    "watch": bool(value.get("watch", value.get("important", False))),
                    "note": text_or_blank(value.get("note"), strip=True),
                    "updated_at": text_or_blank(value.get("updated_at"), strip=True),
                    "event_date": text_or_blank(value.get("event_date"), strip=True),
                    "event_time": text_or_blank(value.get("event_time"), strip=True),
                    "category": text_or_blank(value.get("category"), strip=True),
                    "title": text_or_blank(value.get("title"), strip=True),
                    "subtitle": text_or_blank(value.get("subtitle"), strip=True),
                    "country": text_or_blank(value.get("country"), strip=True),
                    "ticker": text_or_blank(value.get("ticker"), strip=True),
                    "company": text_or_blank(value.get("company"), strip=True),
                    "period": text_or_blank(value.get("period"), strip=True),
                }
            self.state = loaded
            _log(f"Loaded legacy annotations from {json_path.resolve()}")
            return True

        try:
            if self.state_path.exists():
                if self.state_path.suffix.lower() == ".json":
                    _load_json(self.state_path)
                else:
                    _load_csv(self.state_path)
                return

            if self.legacy_state_path and self.legacy_state_path.exists():
                if self.legacy_state_path.suffix.lower() == ".csv":
                    loaded = _load_csv(self.legacy_state_path)
                else:
                    loaded = _load_json(self.legacy_state_path)
                if loaded and self.state_path.suffix.lower() != ".json":
                    self._save_state_to_disk()
                    _log(f"Migrated annotations to {self.state_path.resolve()}")
                return

            _log(f"No annotations file found at {self.state_path.resolve()}")
        except Exception as exc:
            _log(f"Could not load annotations: {exc}")

    def _event_badges_html(self, row, state):
        badge_map = {
            "Economic": "badge-econ",
            "Earnings": "badge-earn",
            "Commodity": "badge-cmdty",
            "Holiday": "badge-holiday",
        }

        html = f'<span class="badge {badge_map.get(row["category"], "badge-econ")}">{safe_html(row["category"])}</span>'
        if self.is_event_important(row["event_id"], row=row):
            source = self.important_source(row["event_id"], row=row)
            label = "Important" if not source else f"Important - {source}"
            html += f'<span class="badge badge-watch">{safe_html(label)}</span>'
        if self._row_flagged(row["event_id"]):
            html += '<span class="badge badge-flag">Flagged</span>'
        return html

    def _event_fields(self, row):
        if row["category"] == "Economic":
            scale = row.get("scaling_factor")
            return [
                ("Date", fmt_value(row["event_date"])),
                ("Time", fmt_time_text(row.get("event_time"))),
                ("Country", fmt_value(row.get("country"), 0)),
                ("Ticker", fmt_value(row.get("ticker"), 0)),
                ("Period", fmt_value(row.get("period"), 0)),
                ("Survey / Expected", fmt_metric(row.get("survey"), scale)),
                ("Prior", fmt_metric(row.get("prior"), scale)),
                ("Actual", fmt_metric(row.get("actual"), scale)),
                ("Revision", fmt_metric(row.get("revision"), scale)),
            ]

        if row["category"] == "Earnings":
            return [
                ("Date", fmt_value(row["event_date"])),
                ("Company", fmt_value(row["company"], 0)),
                ("Ticker", fmt_value(row["ticker"], 0)),
                ("Market Cap", fmt_market_cap(row["market_cap_usd"])),
            ]

        if row["category"] == "Commodity":
            return [
                ("Event Date", fmt_value(row["event_date"])),
                ("Underlying", fmt_value(row["underlying_ticker"], 0)),
                ("Future Price", fmt_value(row["future_price"])),
                ("Currency", fmt_value(row["currency"], 0)),
                ("Future Last Trade", fmt_value(row["delivery_date"])),
                ("Expiry Group", fmt_value(row["expire_group_id"], 0)),
                ("Original IDs", fmt_value(row["orig_ids"], 0)),
            ]

        return [
            ("Date", fmt_value(row["event_date"])),
            ("Market", fmt_value(row["ticker"], 0)),
            ("Country", fmt_value(row["country"], 0)),
            ("Status", "Market closed"),
            ("Reason", "National holiday / exchange closure"),
        ]
# ---------------------------------------------------------------------------
# Section 04 -- State, patches, alarm monitor, and fast tabular explorer
# ---------------------------------------------------------------------------

# ---- Additional CSS (appended from earlier APP_CSS) -----------------------
# NOTE: The APP_CSS variable is assumed to already exist from prior sections.
#       Each block below appends to it with  APP_CSS = APP_CSS + """..."""

APP_CSS = APP_CSS + """
<style>
.row-chip-flag {
    color: #fff;
    background: rgba(251, 113, 133, 0.16);
    border-color: rgba(251, 113, 133, 0.32);
}
.row-chip-note {
    color: #c0e3ff;
    background: rgba(93, 168, 255, 0.15);
    border-color: rgba(93, 168, 255, 0.24);
}
.outcome-chip {
    min-width: 54px;
}
.outcome-beat {
    color: #4ff3c8;
    background: rgba(52, 211, 153, 0.16);
    border-color: rgba(52, 211, 153, 0.30);
}
.outcome-miss {
    color: #ffc6c6;
    background: rgba(239, 68, 68, 0.16);
    border-color: rgba(239, 68, 68, 0.32);
}
.outcome-match {
    color: #d8c8ff;
    background: rgba(148, 163, 184, 0.16);
    border-color: rgba(148, 163, 184, 0.26);
}
.outcome-na {
    color: #c6d3e6;
    background: rgba(148, 163, 184, 0.12);
    border-color: rgba(148, 163, 184, 0.22);
}
.day-scroll-box {
    display: flex;
    flex-direction: column;
    gap: 10px;
}
.day-event-card {
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 10px 12px;
    background: linear-gradient(180deg, rgba(16, 24, 40, 0.96) 0%, rgba(12, 18, 31, 0.98) 100%);
}
.day-event-card.selected-day-event-card {
    box-shadow: inset 0 0 0 1px rgba(93, 168, 255, 0.72);
    background: linear-gradient(180deg, rgba(17, 32, 58, 0.96) 0%, rgba(12, 18, 31, 0.99) 100%);
}
.day-card-top {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 10px;
    margin-bottom: 8px;
}
.day-card-badges {
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
    align-items: center;
}
.day-card-title-btn button {
    width: 100%;
    background: transparent;
    color: var(--text);
    border: none;
    border-radius: 0;
    text-align: left;
    padding: 0;
    font-size: 14px;
    font-weight: 700;
    box-shadow: none;
}
.day-card-title-btn button:hover {
    color: #fff;
}
.day-card-subline {
    color: var(--muted);
    font-size: 12px;
    margin-bottom: 8px;
}
.day-card-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 8px;
}
.day-card-field {
    border: 1px solid rgba(36, 50, 74, 0.92);
    border-radius: 10px;
    padding: 8px 9px;
    background: rgba(11, 16, 32, 0.6);
}
.day-card-field-label {
    color: var(--muted);
    font-size: 10px;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 4px;
}
.day-card-field-value {
    color: var(--text);
    font-size: 12px;
    font-weight: 600;
    font-variant-numeric: tabular-nums;
}
.calendar-day-box.watch-day {
    background: rgba(250, 204, 21, 0.18);
    border-color: rgba(250, 204, 21, 0.34);
    box-shadow: inset 0 0 0 1px rgba(250, 204, 21, 0.38);
}
.calendar-day-box.flagged-day {
    background: rgba(251, 113, 133, 0.18);
    border-color: rgba(251, 113, 133, 0.35);
    box-shadow: inset 0 0 0 1px rgba(251, 113, 133, 0.40);
}
.calendar-day-box.flagged-day.watch-day {
    background: linear-gradient(180deg, rgba(251, 113, 133, 0.18) 0%, rgba(250, 204, 21, 0.16) 100%);
}
.calendar-day-box.watch-day .day-meta strong,
.calendar-day-box.flagged-day .day-meta strong {
    color: #fff;
}
.calendar-day-box.selected-day {
    outline: 2px solid rgba(93, 168, 255, 0.54);
}
</style>
"""

# ---------------------------------------------------------------------------
# Release-outcome pattern lists
# ---------------------------------------------------------------------------

_RELEASE_HIGHER_BETTER_PATTERNS = [normalize_key_text(x) for x in (
    "gdp", "pmi", "manufacturing", "services", "sentiment", "confidence", "job openings",
    "employment change", "change in nonfarm payrolls", "adp employment", "weekly employment change",
    "personal spending", "personal income", "retail sales", "durable goods", "vehicle sales",
    "industrial production", "household spending", "capital spending", "business climate",
    "zew survey expectations", "ifo business climate", "tankan", "real cash earnings",
    "fixed assets", "payrolls", "sales", "gdp sa qoq", "gdp qoq", "gdp yoy", "monthly gdp",
    "services index", "manufacturing index", "empire manufacturing", "consumer confidence",
)]

_RELEASE_LOWER_BETTER_PATTERNS = [normalize_key_text(x) for x in (
    "cpi", "ppi", "core price index", "core pce", "core pce price index", "inflation",
    "price index", "jobless claims", "claims", "claimant count", "unemployment", "job cuts",
    "rate decision", "policy rate", "bank rate", "cash rate", "deposit rate", "loan prime rate",
    "yield rate", "high yield rate", "yield", "public sector net borrowing", "borrowing",
    "employment cost index", "consumer price", "producer price", "harmonized yoy", "loan rate",
    "external communications blackout", "blackout",
)]

# ---------------------------------------------------------------------------
# Display / table constants
# ---------------------------------------------------------------------------

_DISPLAY_CATEGORY_ORDER = {"Economic": 1, "Earnings": 2, "Commodity": 3, "Holiday": 4}

_TABLE_COLUMN_SPECS = {
    "ALL": [
        ("Time", "74px", "time"),
        ("Type", "74px", "category"),
        ("Key", "54px", "key"),
        ("Flag", "56px", "flag"),
        ("Note", "60px", "note"),
        ("Event", "308px", "event"),
        ("Mkt", "90px", "market"),
        ("Ticker", "138px", "ticker"),
        ("Period / Info", "126px", "period"),
        ("Survey", "94px", "survey"),
        ("Prior", "94px", "prior"),
        ("Actual", "94px", "actual"),
        ("B / M", "88px", "outcome"),
    ],
    "ECONOMIC": [
        ("Time", "74px", "time"),
        ("Key", "54px", "key"),
        ("Flag", "56px", "flag"),
        ("Note", "60px", "note"),
        ("Event", "340px", "event"),
        ("Mkt", "96px", "market"),
        ("Ticker", "152px", "ticker"),
        ("Period", "104px", "period"),
        ("Survey", "96px", "survey"),
        ("Prior", "96px", "prior"),
        ("Actual", "96px", "actual"),
        ("B / M", "88px", "outcome"),
    ],
    "EARNINGS": [
        ("Time", "70px", "time"),
        ("Key", "54px", "key"),
        ("Flag", "56px", "flag"),
        ("Note", "60px", "note"),
        ("Company", "330px", "event"),
        ("Ticker", "142px", "ticker"),
        ("Mkt Cap", "118px", "period"),
        ("B / M", "88px", "outcome"),
        ("Market", "96px", "market"),
    ],
    "COMMODITIES": [
        ("Time", "70px", "time"),
        ("Key", "54px", "key"),
        ("Flag", "56px", "flag"),
        ("Note", "60px", "note"),
        ("Event", "320px", "event"),
        ("Mkt", "94px", "market"),
        ("Ticker", "142px", "ticker"),
        ("Period / Info", "138px", "period"),
        ("Survey", "96px", "survey"),
        ("Prior", "96px", "prior"),
        ("Actual", "96px", "actual"),
        ("B / M", "88px", "outcome"),
    ],
    "HOLIDAYS": [
        ("Time", "70px", "time"),
        ("Key", "54px", "key"),
        ("Flag", "56px", "flag"),
        ("Note", "60px", "note"),
        ("Holiday / Closure", "340px", "event"),
        ("Market", "110px", "ticker"),
        ("Geo", "118px", "market"),
        ("Status", "116px", "period"),
        ("B / M", "88px", "outcome"),
    ],
    "IMPORTANT": [
        ("Time", "74px", "time"),
        ("Type", "74px", "category"),
        ("Key", "54px", "key"),
        ("Flag", "56px", "flag"),
        ("Note", "60px", "note"),
        ("Event", "320px", "event"),
        ("Mkt", "94px", "market"),
        ("Ticker", "142px", "ticker"),
        ("Period / Info", "126px", "period"),
        ("Survey", "96px", "survey"),
        ("Prior", "96px", "prior"),
        ("Actual", "96px", "actual"),
        ("B / M", "88px", "outcome"),
    ],
}

# ---------------------------------------------------------------------------
# Helper functions for release outcome
# ---------------------------------------------------------------------------


def _strip_country_word(value):
    text = text_or_blank(value, strip=True)
    if not text:
        return ""
    return re.sub(r"\s+Country$", "", text).strip()


def _safe_float(value):
    if is_missing(value):
        return None
    try:
        return float(value)
    except Exception:
        try:
            return float(str(value).replace(",", "").strip())
        except Exception:
            return None


def _signed_metric(value, scale=None):
    numeric = _safe_float(value)
    if numeric is None:
        return "--"
    sign = "+" if numeric > 0 else ""
    return f"{sign}{fmt_metric(numeric, scale)}"


def _infer_release_direction(row):
    category = text_or_blank(row.get("category"), strip=True)
    title = normalize_key_text(row.get("title"))
    if category == "Earnings":
        return 1
    if category != "Economic":
        return 0
    for pattern in _RELEASE_LOWER_BETTER_PATTERNS:
        if pattern and pattern in title:
            return -1
    for pattern in _RELEASE_HIGHER_BETTER_PATTERNS:
        if pattern and pattern in title:
            return 1
    return 1


def release_outcome(row):
    category = text_or_blank(row.get("category"), strip=True)
    if category not in ("Economic", "Earnings"):
        return {
            "label": "N/A",
            "css": "outcome-na",
            "delta": None,
            "delta_text": "--",
            "title": "No survey/actual comparison is available for this event type.",
        }

    actual = _safe_float(row.get("actual"))
    survey = _safe_float(row.get("survey"))
    scale = row.get("scaling_factor")
    if actual is None or survey is None:
        return {
            "label": "N/A",
            "css": "outcome-na",
            "delta": None,
            "delta_text": "--",
            "title": "Survey and actual data are not both available yet.",
        }

    delta = actual - survey
    if math.isclose(actual, survey, rel_tol=1e-09, abs_tol=max(abs(actual), abs(survey), 1.0) * 1e-08):
        label = "Match"
        css = "outcome-match"
    else:
        direction = _infer_release_direction(row)
        if direction == 0:
            label = "N/A"
            css = "outcome-na"
        else:
            positive = delta * direction > 0
            label = "Beat" if positive else "Miss"
            css = "outcome-beat" if positive else "outcome-miss"

    return {
        "label": label,
        "css": css,
        "delta": delta,
        "delta_text": _signed_metric(delta, scale),
        "title": f"Actual - Survey: {_signed_metric(delta, scale)}",
    }


def outcome_badge_html(row):
    outcome = release_outcome(row)
    return (
        f'<span class="row-chip outcome-chip outcome-{outcome["css"]}" title="{safe_html(outcome["title"])}">'
        f'{safe_html(outcome["label"])}</span>'
    )


def _category_chip_html(row):
    category = text_or_blank(row.get("category"), strip=True)
    mapping = {
        "Economic": ("ECON", "row-chip-econ"),
        "Earnings": ("ERN", "row-chip-earn"),
        "Commodity": ("CMDTY", "row-chip-cmdty"),
        "Holiday": ("HOL", "row-chip-holiday"),
    }
    text, css = mapping.get(category, (category[0:5].upper() or "--", "row-chip-econ"))
    return f'<span class="row-chip row-chip-type {css}">{safe_html(text)}</span>'


def _marker_chip_html(self, row, kind):
    if kind == "key":
        if self.is_event_important(row["event_id"], row=row):
            return '<span class="row-chip row-chip-key">KEY</span>'
        return '<span class="row-chip row-chip-empty">--</span>'
    if kind == "flag":
        if self._row_flagged(row["event_id"]):
            return '<span class="row-chip row-chip-flag">FLAG</span>'
        note_text = text_or_blank(self.state.get(row["event_id"], {}).get("note"), strip=True)
        if note_text:
            return '<span class="row-chip row-chip-note">NOTE</span>'
        return '<span class="row-chip row-chip-empty">--</span>'
    if kind == "note":
        note_text = text_or_blank(self.state.get(row["event_id"], {}).get("note", ""), strip=True)
        if note_text:
            return '<span class="row-chip row-chip-note">NOTE</span>'
        return '<span class="row-chip row-chip-empty">---</span>'


def _row_market_text(self, row):
    category = text_or_blank(row.get("category"), strip=True)
    if category == "Commodity":
        return text_or_blank(coalesce(row.get("underlying_ticker"), row.get("ticker")), strip=True) or "--"
    if category == "Earnings":
        country = _strip_country_word(row.get("country"))
        return country or "--"
    if category == "Holiday":
        country = _strip_country_word(row.get("country"))
        return country or text_or_blank(row.get("ticker"), strip=True) or "--"
    country = _strip_country_word(row.get("country"))
    code = self._row_country_code(row) or ""
    if country:
        return country
    if code:
        return code
    return "--"


def _row_ticker_text(row):
    category = text_or_blank(row.get("category"), strip=True)
    if category == "Commodity":
        return text_or_blank(coalesce(row.get("underlying_ticker"), row.get("ticker")), strip=True) or "--"
    return text_or_blank(row.get("ticker"), strip=True) or "--"


def _row_period_text(row):
    category = text_or_blank(row.get("category"), strip=True)
    if category == "Economic":
        return text_or_blank(row.get("period"), strip=True) or "--"
    if category == "Earnings":
        return fmt_market_cap(row.get("market_cap_usd"))
    if category == "Commodity":
        detail = fmt_value(row.get("delivery_date"))
        if detail == "--":
            detail = fmt_value(row.get("event_date"))
        return detail
    return "Market Closed"


def _row_actual_text(row):
    category = text_or_blank(row.get("category"), strip=True)
    if category == "Commodity":
        price = fmt_value(row.get("future_price"))
        currency = text_or_blank(row.get("currency"), strip=True)
        if price == "--":
            return "--"
        return f"{price} ({currency})".strip()
    return fmt_metric(row.get("actual"), row.get("scaling_factor"))


def _row_cell_value(self, row, key):
    if key == "time":
        return fmt_time_text(row.get("event_time")) or "--"
    if key == "category":
        return _category_chip_html(row)
    if key == "key":
        return _marker_chip_html(self, row, "key")
    if key == "flag":
        return _marker_chip_html(self, row, "flag")
    if key == "note":
        return _marker_chip_html(self, row, "note")
    if key == "event":
        return text_or_blank(row.get("title"), strip=True) or "Untitled event"
    if key == "market":
        return _row_market_text(self, row)
    if key == "ticker":
        return _row_ticker_text(row)
    if key == "period":
        return _row_period_text(row)
    if key == "survey":
        return fmt_metric(row.get("survey"), row.get("scaling_factor"))
    if key == "prior":
        return fmt_metric(row.get("prior"), row.get("scaling_factor"))
    if key == "actual":
        return _row_actual_text(row)
    if key == "outcome":
        return outcome_badge_html(row)
    return "--"


def _table_html_cell(content, width, align="left", html_content=False, extra_classes=""):
    if content is None:
        content = "--"
    if not html_content:
        title_text = text_or_blank(content, strip=True) or ""
        inner = safe_html(title_text) if title_text else "&ndash;"
    else:
        title_text = ""
        inner = content
    align_class = {"left": "", "center": "list-cell-center", "right": "list-cell-right"}.get(align, "")
    layout = widgets.Layout(width=width, min_width=width)
    widget = widgets.HTML(
        f'<div class="event-app list-cell-html {align_class} {extra_classes}"'
        + (f' title="{safe_html(title_text)}"' if title_text else "")
        + f">{inner}</div>",
        layout=layout,
    )
    return widget


def _event_button_widget(self, row, width):
    title_text = text_or_blank(row.get("title"), strip=True) or "Untitled event"
    btn = widgets.Button(
        description=truncate_display_text(title_text, 54),
        tooltip=title_text,
        layout=widgets.Layout(width=width, min_width=width),
    )
    btn.add_class("list-event-button")
    if row["event_id"] == self.current_event_id:
        btn.add_class("list-event-button-selected")
    btn.on_click(lambda _, event_id=row["event_id"]: _activate_event_selection(self, event_id))
    return btn


def _group_header_widget(label):
    widget = widgets.HTML(f'<div class="event-app list-group-row">{safe_html(label)}</div>')
    return widget


def _row_box_widget(self, row, columns, stripe_even=False):
    cells = []
    for _, width, key in columns:
        if key == "event":
            cells.append(_event_button_widget(self, row, width))
        else:
            value = _row_cell_value(self, row, key)
            html_content = key in ("category", "key", "flag", "note", "outcome")
            align = "left"
            if key in ("time", "category", "key", "flag", "note", "outcome"):
                align = "center"
            elif key in ("survey", "prior", "actual"):
                align = "right"
            cells.append(_table_html_cell(value, width, align=align, html_content=html_content))
    row_box = widgets.HBox(cells, layout=widgets.Layout(width="1260px"))
    row_box.add_class("list-data-row")
    if stripe_even:
        row_box.add_class("stripe-even")
    if self.is_event_important(row["event_id"], row=row):
        row_box.add_class("important-row")
    if self._row_flagged(row["event_id"]):
        row_box.add_class("flagged-row")
    if row["event_id"] == self.current_event_id:
        row_box.add_class("selected-row")
    return row_box


def _header_row_widget(columns):
    header_cells = []
    for label, width, key in columns:
        align = "center" if key in ("time", "category", "key", "flag", "note", "outcome") else "left"
        if key in ("survey", "prior", "actual"):
            align = "right"
        header_cells.append(
            _table_html_cell(label, width, align=align, html_content=False, extra_classes="list-header-cell")
        )
    row_box = widgets.HBox(header_cells, layout=widgets.Layout(width="1260px"))
    row_box.add_class("list-header-row")
    return row_box


def _render_event_table(self, output_widget, df, view_name="ALL"):
    columns = _TABLE_COLUMN_SPECS.get(view_name, _TABLE_COLUMN_SPECS["ALL"])
    with output_widget:
        clear_output(wait=True)
        if df is None or df.empty:
            display(HTML('<div class="event-app empty-state">No events in this view.</div>'))
            return

        render_df = df.copy()
        render_df["event_date"] = pd.to_datetime(render_df["event_date"], errors="coerce")
        render_df["category_order"] = render_df["category"].map(_DISPLAY_CATEGORY_ORDER).fillna(99)
        render_df["sort_time"] = render_df["event_time"].fillna("99:99")
        render_df = render_df.sort_values(["event_date", "category_order", "sort_time", "title"]).drop(columns=["category_order", "sort_time"])

        table_children = [_header_row_widget(columns)]
        stripe_even = False

        for day_value, day_group in render_df.groupby(render_df["event_date"].dt.date, sort=True):
            total = len(day_group)
            key_count = int(sum(self.is_event_important(row["event_id"], row=row) for row in day_group.to_dict("records")))
            flag_count = int(sum(self._row_flagged(event_id) for event_id in day_group["event_id"]))
            day_label = pd.Timestamp(day_value).strftime("%a %Y-%m-%d")
            header_bits = [f"{day_label} \u00b7 {total} event{'s' if total != 1 else ''}"]
            if key_count:
                header_bits.append(f"KEY ({key_count})")
            if flag_count:
                header_bits.append(f"FLAG ({flag_count})")
            holiday_codes = (
                day_group.loc[day_group["category"] == "Holiday", "ticker"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
            if holiday_codes:
                header_bits.append("Closed: " + ", ".join(sorted(holiday_codes)))
            table_children.append(_group_header_widget(" \u00b7 ".join(header_bits)))
            for row in day_group.to_dict("records"):
                table_children.append(_row_box_widget(self, row, columns, stripe_even=stripe_even))
                stripe_even = not stripe_even

        table = widgets.VBox(table_children, layout=widgets.Layout(width="1260px"))
        table.add_class("list-table")
        scroller = widgets.Box([table], layout=widgets.Layout(width="100%", height="560px", overflow="auto"))
        scroller.add_class("list-scroller")
        shell = widgets.Box([scroller], layout=widgets.Layout(width="100%"))
        shell.add_class("list-shell")
        display(shell)


def _activate_event_selection(self, event_id):
    row = self._find_event_row(event_id)
    self.current_event_id = event_id
    if row is not None:
        if hasattr(row, "to_dict"):
            row_dict = row.to_dict()
        else:
            row_dict = dict(row)
        event_date = row_dict.get("event_date")
        if not is_missing(event_date):
            self.selected_day = pd.Timestamp(event_date).date()
        self.render_detail(event_id)
        self.refresh_lists()
        self.render_selected_day_detail()
        self.render_monthly_page()


# ---------------------------------------------------------------------------
# Patched build_widgets
# ---------------------------------------------------------------------------

def _patched_build_widgets(self, months_per_page_default):
    default_start = pd.Timestamp.today().date()
    default_end = (pd.Timestamp.today() + pd.Timedelta(days=45)).date()

    common_desc = {"description_width": "initial"}

    self.start_picker = widgets.DatePicker(description="Start", value=default_start, style=common_desc)
    self.end_picker = widgets.DatePicker(description="End", value=default_end, style=common_desc)
    self.refresh_button = widgets.Button(description="Refresh Data", button_style="primary", icon="refresh")
    self.apply_filters_button = widgets.Button(description="Apply Explorer Filters", icon="filter")
    self.reset_filters_button = widgets.Button(description="Reset Filters", icon="undo")

    self.search_box = widgets.Text(
        description="Search",
        placeholder="Event, ticker, company, country...",
        style=common_desc,
        layout=widgets.Layout(width="340px"),
    )

    self.category_filter = widgets.SelectMultiple(
        description="Explorer Categories",
        options=["Economic", "Earnings", "Commodity", "Holiday"],
        value=("Economic", "Earnings", "Commodity", "Holiday"),
        rows=5,
        style=common_desc,
        layout=widgets.Layout(width="260px"),
    )

    self.only_flagged = widgets.Checkbox(description="Flagged only", value=False, style=common_desc)
    self.only_watch = widgets.Checkbox(description="Important only", value=False, style=common_desc)

    self.months_per_page = widgets.BoundedIntText(
        description="Months / pane",
        value=max(1, int(months_per_page_default)),
        min=1,
        max=12,
        style=common_desc,
        layout=widgets.Layout(width="170px"),
    )

    self.prev_months_button = widgets.Button(description="Previous Pane", icon="arrow-left")
    self.next_months_button = widgets.Button(description="Next Pane", icon="arrow-right")
    self.month_page_label = widgets.HTML()
    self.month_page_note = widgets.HTML(
        '<div class="event-app page-caption">'
        "Monthly view keeps market calendar and now color-highlights KEY days in yellow and FLAG days in red. "
        "Click any day to open a clickable side-state with shared notes, beat / miss status, ticker, period, and release metrics."
        "</div>"
    )

    self.status_out = widgets.Output()
    self.summary_out = widgets.Output()
    self.monthly_out = widgets.Output()
    self.day_out = widgets.Output(layout=widgets.Layout(height="780px", overflow="auto"))
    self.detail_html = widgets.HTML()
    self.watchlist_out = widgets.HTML()

    list_layout = widgets.Layout(width="100%", height="560px", overflow="auto")
    self.all_list = widgets.Output(layout=list_layout)
    self.econ_list = widgets.Output(layout=list_layout)
    self.earnings_list = widgets.Output(layout=list_layout)
    self.comdty_list = widgets.Output(layout=list_layout)
    self.holiday_list = widgets.Output(layout=list_layout)
    self.watch_list = widgets.Output(layout=list_layout)

    self.tabs = widgets.Tab(children=[
        self.all_list,
        self.econ_list,
        self.earnings_list,
        self.comdty_list,
        self.holiday_list,
        self.watch_list,
    ])

    for i, title in enumerate(["ALL", "ECONOMIC", "EARNINGS", "COMMODITIES", "HOLIDAYS", "IMPORTANT"]):
        self.tabs.set_title(i, title)

    self.flag_toggle = widgets.ToggleButton(
        description="Flag", icon="flag-o", value=False, layout=widgets.Layout(width="120px"),
    )

    self.watch_toggle = widgets.ToggleButton(
        description="Important", icon="star-o", value=False, layout=widgets.Layout(width="150px"),
    )

    self.note_area = widgets.Textarea(
        description="Notes",
        placeholder="Add custom notes, trading angles, reminders, scenario comments...",
        layout=widgets.Layout(width="100%", height="150px"),
        style=common_desc,
    )

    self.save_note_button = widgets.Button(description="Save Note", button_style="success", icon="save")
    self.clear_note_button = widgets.Button(description="Clear Note", icon="trash")
    self.export_state_button = widgets.Button(description="Sync CSV Now", icon="save")
    self.load_state_button = widgets.Button(description="Reload CSV", icon="refresh")
    self.clear_watchlist_button = widgets.Button(description="Clear Important List", icon="times")

    self.export_scope_dropdown = widgets.Dropdown(
        description="Export Scope",
        options=[
            ("Current visible pane", "pane"),
            ("Full selected range", "range"),
            ("Important / flagged only", "highlights"),
        ],
        value="pane",
        style=common_desc,
        layout=widgets.Layout(width="330px"),
    )

    self.export_document_button = widgets.Button(
        description="Export View CSV", icon="download", button_style="primary",
    )

    self.export_note = widgets.HTML(
        '<div class="event-app page-caption">'
        "Create a rich CSV snapshot for the active view, now including beat / miss status. "
        "Copy the CSV text from the box below into a local file if notebook downloads are blocked, "
        "then run the companion local builder notebook to format the report into a Wall Street-style Excel workbook."
        "</div>"
    )

    self.export_csv_path = widgets.HTML(
        '<div class="event-app page-caption">CSV export preview will appear here after each export.</div>'
    )

    self.export_csv_preview = widgets.Textarea(
        description="CSV Preview",
        placeholder="After export, the CSV text will appear here for copy / paste.",
        layout=widgets.Layout(width="100%", height="220px"),
        style=common_desc,
    )

    self.persistence_note = widgets.HTML(
        '<div class="event-app page-caption">'
        f'Annotations auto-save to <b>{safe_html(str(self.state_path))}</b>.'
        "Rule-based key releases from the embedded watch list stay highlighted even after clearing user-selected important items."
        "</div>"
    )

    controls_title = widgets.HTML('<div class="event-app section-title">Controls</div>')
    controls_grid = widgets.VBox([
        widgets.HBox([self.start_picker, self.end_picker, self.refresh_button]),
        widgets.HBox([self.search_box, self.apply_filters_button, self.reset_filters_button]),
        widgets.HBox([self.category_filter, widgets.VBox([self.only_flagged, self.only_watch])]),
    ])

    self.controls_panel = widgets.VBox([controls_title, controls_grid])
    self.controls_panel.add_class("app-panel")

    summary_title = widgets.HTML('<div class="event-app section-title">Explorer Summary</div>')
    self.summary_panel = widgets.VBox([summary_title, self.summary_out])
    self.summary_panel.add_class("app-panel")

    monthly_title = widgets.HTML('<div class="event-app section-title">Monthly Page</div>')
    month_nav = widgets.HBox([
        self.months_per_page,
        self.prev_months_button,
        self.next_months_button,
        self.month_page_label,
    ])

    self.monthly_panel = widgets.VBox([monthly_title, month_nav, self.month_page_note, self.monthly_out])
    self.monthly_panel.add_class("app-panel")
    self.monthly_panel.layout = widgets.Layout(width="66%")

    day_title = widgets.HTML('<div class="event-app section-title">Selected Day</div>')
    self.day_panel = widgets.VBox([day_title, self.day_out])
    self.day_panel.add_class("app-panel")
    self.day_panel.layout = widgets.Layout(width="34%")

    list_caption = widgets.HTML(
        '<div class="event-app list-panel-caption">'
        "Table-based explorer. Click any event title to sync the detail pane and notes. "
        "KEY = important, FLAG = flagged, NOTE = saved annotation, B / M = beat / miss versus survey."
        "</div>"
    )

    list_card = widgets.VBox([list_caption, self.tabs], layout=widgets.Layout(width="50%"))
    list_card.add_class("app-panel")

    detail_title = widgets.HTML('<div class="event-app section-title">Event Detail & Notes</div>')
    detail_controls = widgets.HBox([
        self.flag_toggle,
        self.watch_toggle,
        self.save_note_button,
        self.clear_note_button,
    ])

    detail_actions = widgets.HBox([self.export_state_button, self.load_state_button, self.clear_watchlist_button])
    export_controls = widgets.HBox([self.export_scope_dropdown, self.export_document_button])
    detail_card = widgets.VBox([
        detail_title,
        self.detail_html,
        detail_controls,
        self.note_area,
        detail_actions,
        self.persistence_note,
        export_controls,
        self.export_note,
        self.export_csv_path,
        self.export_csv_preview,
        self.watchlist_out,
        self.status_out,
    ], layout=widgets.Layout(width="40%"))
    detail_card.add_class("app-panel")

    header_html = widgets.HTML(
        APP_CSS +
        """
        <div class="event-app">
            <div class="app-title">Market Event Calendar</div>
            <div class="app-subtitle">
                Dark-themed notebook calendar for economic releases, large-cap earnings, futures timing, market holidays, and embedded high-priority macro
                events.
            </div>
            <div class="app-footnote">
                The explorer now uses a professional table layout, selected-day cards are clickable and note-aware, and beat / miss tags track economic releases and
                any earnings rows that later gain survey versus actual fields.
            </div>
        </div>
        """
    )

    self.root = widgets.VBox([
        header_html,
        self.controls_panel,
        self.summary_panel,
        widgets.HBox([self.monthly_panel, self.day_panel]),
        widgets.HBox([list_card, detail_card]),
    ], layout=widgets.Layout(width="100%"))
    self.root.add_class("event-app-shell")


# ---------------------------------------------------------------------------
# Patched wire_events
# ---------------------------------------------------------------------------

def _patched_wire_events(self):
    self.refresh_button.on_click(self.refresh_data)
    self.apply_filters_button.on_click(self.apply_filters)
    self.reset_filters_button.on_click(self.reset_filters)
    self.save_note_button.on_click(self.save_note)
    self.clear_note_button.on_click(self.clear_note)
    self.export_state_button.on_click(self.export_state)
    self.load_state_button.on_click(self.load_state)
    self.clear_watchlist_button.on_click(self.clear_watchlist)
    self.export_document_button.on_click(self.export_current_view_csv)
    self.prev_months_button.on_click(self.previous_month_pane)
    self.next_months_button.on_click(self.next_month_pane)
    self.months_per_page.observe(self.on_months_per_page_changed, names="value")
    self.tabs.observe(self.on_tab_change, names="selected_index")
    self.flag_toggle.observe(self.on_flag_toggled, names="value")
    self.watch_toggle.observe(self.on_watch_toggled, names="value")


# ---------------------------------------------------------------------------
# Patched event_badges_html
# ---------------------------------------------------------------------------

def _patched_event_badges_html(self, row, state):
    badge_map = {
        "Economic": "badge-econ",
        "Earnings": "badge-earn",
        "Commodity": "badge-cmdty",
        "Holiday": "badge-holiday",
    }
    html = f'<span class="badge {badge_map.get(row["category"], "badge-econ")}">{safe_html(row["category"])}</span>'
    html += outcome_badge_html(row)
    if self.is_event_important(row["event_id"], row=row):
        source = self.important_source(row["event_id"], row=row)
        label = "Important" if not source else f"Important \u00b7 {source}"
        html += f'<span class="badge badge-watch">{safe_html(label)}</span>'
    if self._row_flagged(row["event_id"]):
        html += '<span class="badge badge-flag">Flagged</span>'
    note_text = text_or_blank(state.get("note"), strip=True)
    if note_text:
        html += '<span class="badge" style="color:#c0e3ff;background:rgba(93,168,255,0.14);border-color:rgba(93,168,255,0.28)">Notes</span>'
    return html


# ---------------------------------------------------------------------------
# Patched event_fields
# ---------------------------------------------------------------------------

def _patched_event_fields(self, row):
    outcome = release_outcome(row)
    if row["category"] == "Economic":
        scale = row.get("scaling_factor")
        return [
            ("Date", fmt_value(row["event_date"])),
            ("Time", fmt_time_text(row.get("event_time"))),
            ("Market", _row_market_text(self, row)),
            ("Ticker", _row_ticker_text(row)),
            ("Period", text_or_blank(row.get("period"), strip=True) or "--"),
            ("Survey / Expected", fmt_metric(row.get("survey"), scale)),
            ("Prior", fmt_metric(row.get("prior"), scale)),
            ("Actual", fmt_metric(row.get("actual"), scale)),
            ("Revision", fmt_metric(row.get("revision"), scale)),
            ("Beat / Miss", outcome["label"]),
            ("Actual - Survey", outcome["delta_text"]),
        ]

    if row["category"] == "Earnings":
        return [
            ("Date", fmt_value(row["event_date"])),
            ("Company", fmt_value(row.get("company", ""))),
            ("Ticker", _row_ticker_text(row)),
            ("Market Cap", fmt_market_cap(row["market_cap_usd"])),
            ("Beat / Miss", outcome["label"]),
            ("Actual - Survey", outcome["delta_text"]),
        ]

    if row["category"] == "Commodity":
        return [
            ("Event Date", fmt_value(row["event_date"])),
            ("Underlying", fmt_value(row["underlying_ticker"], 0)),
            ("Ticker", _row_ticker_text(row)),
            ("Future Price", fmt_value(row["future_price"])),
            ("Currency", fmt_value(row["currency"], 0)),
            ("Future Last Trade", fmt_value(row["delivery_date"])),
            ("Expiry Group", fmt_value(row["expire_group_id"], 0)),
            ("Original IDs", fmt_value(row["orig_ids"], 0)),
            ("Beat / Miss", outcome["label"]),
        ]

    return [
        ("Date", fmt_value(row["event_date"])),
        ("Market", fmt_value(row["ticker"], 0)),
        ("Geo", _strip_country_word(row.get("country")) or "--"),
        ("Status", "Market closed"),
        ("Reason", "National holiday / exchange closure"),
        ("Beat / Miss", outcome["label"]),
    ]


# ---------------------------------------------------------------------------
# Patched refresh_lists
# ---------------------------------------------------------------------------

def _patched_refresh_lists(self):
    econ_df = self.filtered_events[self.filtered_events["category"] == "Economic"]
    earnings_df = self.filtered_events[self.filtered_events["category"] == "Earnings"]
    comdty_df = self.filtered_events[self.filtered_events["category"] == "Commodity"]
    holiday_df = self.filtered_events[self.filtered_events["category"] == "Holiday"]
    watch_df = self.filtered_events[
        self.filtered_events["event_id"].map(lambda x: self.is_event_important(x))
    ]

    _render_event_table(self, self.all_list, self.filtered_events, view_name="ALL")
    _render_event_table(self, self.econ_list, econ_df, view_name="ECONOMIC")
    _render_event_table(self, self.earnings_list, earnings_df, view_name="EARNINGS")
    _render_event_table(self, self.comdty_list, comdty_df, view_name="COMMODITIES")
    _render_event_table(self, self.holiday_list, holiday_df, view_name="HOLIDAYS")
    _render_event_table(self, self.watch_list, watch_df, view_name="IMPORTANT")

    self.render_watchlist_box()

    if self.current_event_id:
        self.render_detail(self.current_event_id)
        if self.filtered_events is not None and not self.filtered_events.empty:
            preferred_df = self.filtered_events.copy()
            preferred_df["flag_sort"] = preferred_df["event_id"].map(lambda x: 0 if self._row_flagged(x) else 1)
            preferred_df["important_sort"] = preferred_df["event_id"].map(lambda x: 0 if self.is_event_important(x) else 1)
            preferred_df["category_order"] = preferred_df["category"].map(_DISPLAY_CATEGORY_ORDER).fillna(99)
            preferred_df["sort_time"] = preferred_df["event_time"].fillna("99:99")
            preferred_df = preferred_df.sort_values(["event_date", "flag_sort", "important_sort", "category_order", "sort_time", "title"])
            self.current_event_id = preferred_df.iloc[0]["event_id"]
            self.render_detail(self.current_event_id)
    else:
        self.render_detail(None)


def _patched_get_active_list_widget(self):
    idx = self.tabs.selected_index or 0
    return [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list][idx]


def _patched_on_tab_change(self, change):
    self.refresh_lists()


def _patched_on_event_selected(self, change):
    return


# ---------------------------------------------------------------------------
# Patched build_day_box
# ---------------------------------------------------------------------------

def _patched_build_day_box(self, day, current_month, start_day, end_day, summary):
    in_range = start_day <= day <= end_day
    in_month = day.month == current_month
    total = summary["counts"].get(day, 0)
    econ = summary["econ"].get(day, 0)
    ern = summary["earnings"].get(day, 0)
    cmd = summary["commodity"].get(day, 0)
    hol = summary["holiday"].get(day, 0)
    key_count = summary["watch"].get(day, 0)
    flag_count = summary["flagged"].get(day, 0)
    hol_codes = summary["holiday_codes"].get(day, [])

    day_button = widgets.Button(
        description=str(day.day),
        layout=widgets.Layout(width="100%"),
        disabled=not in_range,
        tooltip=f"{day.isoformat()}",
    )

    day_button.add_class("calendar-day-btn")
    if in_range:
        day_button.on_click(lambda _, selected_day=day: self.on_day_clicked(selected_day))

    meta_lines = []
    if total:
        meta_lines.append(f"<strong>{total} evt</strong>")
    bucket_parts = []
    if econ:
        bucket_parts.append(f"ECO ({econ})")
    if ern:
        bucket_parts.append(f"ERN ({ern})")
    if cmd:
        bucket_parts.append(f"CMD ({cmd})")
    if hol:
        bucket_parts.append(f"HOL ({hol})")
    if bucket_parts:
        meta_lines.append(" | ".join(bucket_parts))

    chips = []
    if key_count:
        chips.append(f'<span class="row-chip row-chip-key">KEY ({key_count})</span>')
    if flag_count:
        chips.append(f'<span class="row-chip row-chip-flag">FLAG ({flag_count})</span>')
    if hol_codes:
        for code in hol_codes[:3]:
            chips.append(f'<span class="holiday-chip">{safe_html(code)}</span>')
        if len(hol_codes) > 3:
            chips.append(f'<span class="holiday-chip">+{len(hol_codes) - 3}</span>')

    chips_block = ""
    if chips:
        chips_block = (
            '<div style="margin-top:6px;display:flex;flex-wrap:wrap;gap:4px;">'
            + "".join(chips)
            + "</div>"
        )

    meta_html = widgets.HTML(
        f'<div class="event-app day-meta">'
        f'{"<br>".join(meta_lines) if meta_lines else "&nbsp;"}'
        f"{chips_block}"
        f"</div>"
    )

    box = widgets.VBox([day_button, meta_html], layout=widgets.Layout(width="100%"))
    box.add_class("calendar-day-box")
    if not in_month:
        box.add_class("out-month")
    if total:
        box.add_class("has-events")
    if summary["holiday"].get(day, 0):
        box.add_class("holiday-day")
    if key_count:
        box.add_class("watch-day")
    if flag_count:
        box.add_class("flagged-day")
    if self.selected_day == day:
        box.add_class("selected-day")
    if not in_range:
        box.add_class("no-click")
    return box


# ---------------------------------------------------------------------------
# Patched day_card_widget
# ---------------------------------------------------------------------------

def _day_card_widget(self, row):
    state = self.state.get(row["event_id"], {})
    subtitle_bits = []
    time_text = fmt_time_text(row.get("event_time"))
    market_text = _row_market_text(self, row)
    ticker_text = _row_ticker_text(row)
    period_text = _row_period_text(row)
    if time_text and time_text != "--":
        subtitle_bits.append(time_text)
    if market_text and market_text != "--":
        subtitle_bits.append(market_text)
    if ticker_text and ticker_text != "--":
        subtitle_bits.append(ticker_text)
    if row.get("category") == "Economic" and period_text and period_text != "--":
        subtitle_bits.append(f"Period ({period_text})")

    fields = self._event_fields(row)
    shown_fields = []
    for label, value in fields:
        if label in ("Date", "Event Date"):
            continue
        if value in ("--", None, ""):
            if label not in ("Beat / Miss",):
                continue
        shown_fields.append(
            f'<div class="day-card-field">'
            f'<div class="day-card-field-label">{safe_html(label)}</div>'
            f'<div class="day-card-field-value">{safe_html(value)}</div>'
            f"</div>"
        )

    title_btn = widgets.Button(
        description=truncate_display_text(text_or_blank(row.get("title"), strip=True) or "Untitled event", 70),
        tooltip=text_or_blank(row.get("title"), strip=True) or "Untitled event",
        layout=widgets.Layout(width="100%"),
    )
    title_btn.add_class("day-card-title-btn")
    title_btn.on_click(lambda _, event_id=row["event_id"]: _activate_event_selection(self, event_id))

    badge_html = widgets.HTML(
        f'<div class="event-app day-card-badges">{self._event_badges_html(row, state)}</div>'
    )

    _sub_text = " \u00b7 ".join(subtitle_bits) if subtitle_bits else text_or_blank(row.get("subtitle"), strip=True) or "--"
    subtitle_html = widgets.HTML(
        f'<div class="event-app day-card-subline">{_sub_text}</div>'
    )

    grid_html = widgets.HTML(
        f'<div class="event-app day-card-grid">{"".join(shown_fields)}</div>'
    )

    pieces = [
        widgets.HBox([title_btn], layout=widgets.Layout(width="100%")),
        badge_html,
        subtitle_html,
        grid_html,
    ]

    note_text = text_or_blank(state.get("note"), strip=True)
    if note_text:
        pieces.append(widgets.HTML(f'<div class="event-app note-preview">Saved note:<b><br>{safe_html(note_text)}</b></div>'))

    card = widgets.VBox(pieces)
    card.add_class("day-event-card")
    if row["event_id"] == self.current_event_id:
        card.add_class("selected-day-event-card")
    return card


# ---------------------------------------------------------------------------
# Patched render_selected_day_detail
# ---------------------------------------------------------------------------

def _patched_render_selected_day_detail(self):
    with self.day_out:
        clear_output(wait=True)
        if self.selected_day is None:
            display(HTML('<div class="event-app day-scroll-box"><div class="empty-state">Click a day to inspect the full state.</div></div>'))
            return

        day_df = self.events[self.events["event_date"].dt.date == self.selected_day].copy()
        if day_df.empty:
            html = (
                '<div class="event-app day-scroll-box">'
                f'<div class="day-panel-header"><div class="day-panel-title">{safe_html(self.selected_day)}</div>'
                f'<div class="day-panel-subtitle">No events or holiday markers for this day.</div></div>'
                '<div class="empty-state">No events in the selected range for this date.</div></div>'
            )
            display(HTML(html))
            return

        day_df["flag_sort"] = day_df["event_id"].map(lambda x: 0 if self._row_flagged(x) else 1)
        day_df["important_sort"] = day_df["event_id"].map(lambda x: 0 if self.is_event_important(x) else 1)
        day_df["category_order"] = day_df["category"].map(_DISPLAY_CATEGORY_ORDER).fillna(99)
        day_df["sort_time"] = day_df["event_time"].fillna("99:99")
        day_df = day_df.sort_values(["flag_sort", "important_sort", "category_order", "sort_time", "title"]).drop(columns=["flag_sort", "important_sort", "category_order", "sort_time"])

        holiday_codes = day_df.loc[day_df["category"] == "Holiday", "ticker"].dropna().astype(str).unique().tolist()
        header_bits = [f"{len(day_df)} total events"]
        key_count = int(sum(self.is_event_important(row["event_id"], row=row) for row in day_df.to_dict("records")))
        flag_count = int(sum(self._row_flagged(event_id) for event_id in day_df["event_id"]))
        if key_count:
            header_bits.append(f"KEY ({key_count})")
        if flag_count:
            header_bits.append(f"FLAG ({flag_count})")
        if holiday_codes:
            header_bits.append("Markets closed: " + ", ".join(sorted(holiday_codes)))

        _hdr_sub = " \u00b7 ".join(header_bits)
        header = widgets.HTML(
            f'<div class="event-app day-panel-header">'
            f'<div class="day-panel-title">{safe_html(str(self.selected_day))}</div>'
            f'<div class="day-panel-subtitle">{safe_html(_hdr_sub)}</div>'
            f"</div>"
        )

        cards = [header] + [_day_card_widget(self, row) for row in day_df.to_dict("records")]
        display(widgets.VBox(cards, layout=widgets.Layout(width="100%")))


# ---------------------------------------------------------------------------
# Patched on_day_clicked
# ---------------------------------------------------------------------------

def _patched_on_day_clicked(self, selected_day):
    self.selected_day = pd.Timestamp(selected_day).date()
    day_events = self.events[self.events["event_date"].dt.date == self.selected_day].copy()
    if not day_events.empty:
        day_events["flag_sort"] = day_events["event_id"].map(lambda x: 0 if self._row_flagged(x) else 1)
        day_events["important_sort"] = day_events["event_id"].map(lambda x: 0 if self.is_event_important(x) else 1)
        day_events["category_order"] = day_events["category"].map(_DISPLAY_CATEGORY_ORDER).fillna(99)
        day_events["sort_time"] = day_events["event_time"].fillna("99:99")
        day_events = day_events.sort_values(["flag_sort", "important_sort", "category_order", "sort_time", "title"]).drop(columns=["flag_sort", "important_sort", "category_order", "sort_time"])
        self.current_event_id = day_events.iloc[0]["event_id"]
        self.render_detail(self.current_event_id)
    else:
        self.current_event_id = None
        self.render_detail(None)
    self.refresh_lists()
    self.render_monthly_page()
    self.render_selected_day_detail()


# ---------------------------------------------------------------------------
# Patched export_snapshot_dataframe
# ---------------------------------------------------------------------------

def _patched_export_snapshot_dataframe(self, df):
    rows = []
    for row in df.to_dict("records"):
        state = self.state.get(row["event_id"], {})
        important = self.is_event_important(row["event_id"], row=row)
        flagged = self._row_flagged(row["event_id"])
        important_source = self.important_source(row["event_id"], row=row) or ""
        outcome = release_outcome(row)

        scale_value = text_or_blank(row.get("scaling_factor"), strip=True)
        category = text_or_blank(row.get("category"), strip=True)
        event_title = text_or_blank(row.get("title"), strip=True)
        event_time = fmt_time_text(row.get("event_time"))
        country = _strip_country_word(row.get("country")) or "--"
        ticker = text_or_blank(_row_ticker_text(row), strip=True) or "--"
        period = text_or_blank(row.get("period"), strip=True)
        subtitle = text_or_blank(row.get("subtitle"), strip=True)

        if category == "Economic":
            market_value = _row_market_text(self, row)
            period_detail = period or "--"
            survey = fmt_metric(row.get("survey"), scale_value)
            prior = fmt_metric(row.get("prior"), scale_value)
            actual = fmt_metric(row.get("actual"), scale_value)
            revision = fmt_metric(row.get("revision"), scale_value)
        elif category == "Earnings":
            market_value = ticker or "--"
            period_detail = fmt_market_cap(row.get("market_cap_usd"))
            survey = fmt_metric(row.get("survey"), scale_value)
            prior = fmt_metric(row.get("prior"), scale_value)
            actual = fmt_metric(row.get("actual"), scale_value)
            revision = fmt_metric(row.get("revision"), scale_value)
        elif category == "Commodity":
            market_value = text_or_blank(coalesce(row.get("underlying_ticker"), row.get("ticker")), strip=True) or "--"
            period_detail = fmt_value(row.get("delivery_date"))
            survey = "--"
            prior = "--"
            actual = _row_actual_text(row)
            revision = "--"
        else:
            market_value = clean_join([row.get("ticker"), _strip_country_word(row.get("country"))]) or country or "--"
            period_detail = "Market Closed"
            survey = "--"
            prior = "--"
            actual = "--"
            revision = "--"

        note_text = text_or_blank(state.get("note"), strip=True).replace("\n", " | ")
        if len(note_text) > 280:
            note_text = note_text[:277] + "..."

        status_tags = []
        if important:
            status_tags.append("IMPORTANT")
        if flagged:
            status_tags.append("FLAGGED")
        if category == "Holiday":
            status_tags.append("CLOSED")
        status_text = " / ".join(status_tags) if status_tags else "--"
        notes_and_tags = " \u00b7 ".join([x for x in (status_tags + ([note_text] if note_text else [])) if x]) or "--"

        rows.append({
            "Event ID": row["event_id"],
            "Date": pd.Timestamp(row["event_date"]).strftime("%Y-%m-%d"),
            "Day": pd.Timestamp(row["event_date"]).strftime("%a"),
            "Time": event_time,
            "Category": category,
            "Event": event_title,
            "Country": country,
            "Ticker": ticker,
            "Market / Ticker": market_value or "--",
            "Period": period or ("Market Closed" if category == "Holiday" else "--"),
            "Period / Detail": period_detail,
            "Survey / Expected": survey,
            "Survey": survey,
            "Prior": prior,
            "Actual": actual,
            "Actual - Survey": outcome["delta_text"],
            "Beat / Miss": outcome["label"],
            "Revision": revision,
            "Scaling": scale_value or "--",
            "Status": status_text,
            "Important": bool(important),
            "Flagged": bool(flagged),
            "Important Source": important_source,
            "Subtitle": subtitle or "--",
            "Notes": note_text or "--",
            "Notes / Tags": notes_and_tags,
        })

    if not rows:
        return pd.DataFrame(columns=[
            "Event ID", "Date", "Day", "Time", "Category", "Event", "Country", "Ticker",
            "Market / Ticker", "Period", "Period / Detail", "Survey / Expected", "Survey",
            "Prior", "Actual", "Actual - Survey", "Beat / Miss", "Revision", "Scaling",
            "Status", "Important", "Flagged", "Important Source", "Subtitle", "Notes", "Notes / Tags",
        ])
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# save_note / clear_note patches (with refresh)
# ---------------------------------------------------------------------------

_original_save_note = BloombergEventCalendarApp.save_note


def _patched_save_note(self, _=None):
    _original_save_note(self, _)
    try:
        self.refresh_lists()
    except Exception:
        pass


_original_clear_note = BloombergEventCalendarApp.clear_note


def _patched_clear_note(self, _=None):
    _original_clear_note(self, _)
    try:
        self.refresh_lists()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Monkey-patch assignments (module level)
# ---------------------------------------------------------------------------

BloombergEventCalendarApp.build_widgets = _patched_build_widgets
BloombergEventCalendarApp._wire_events = _patched_wire_events
BloombergEventCalendarApp._event_badges_html = _patched_event_badges_html
BloombergEventCalendarApp._event_fields = _patched_event_fields
BloombergEventCalendarApp.refresh_lists = _patched_refresh_lists
BloombergEventCalendarApp.get_active_list_widget = _patched_get_active_list_widget
BloombergEventCalendarApp.on_tab_change = _patched_on_tab_change
BloombergEventCalendarApp.on_event_selected = _patched_on_event_selected
BloombergEventCalendarApp.build_day_box = _patched_build_day_box
BloombergEventCalendarApp.render_selected_day_detail = _patched_render_selected_day_detail
BloombergEventCalendarApp.on_day_clicked = _patched_on_day_clicked
BloombergEventCalendarApp.export_snapshot_dataframe = _patched_export_snapshot_dataframe
BloombergEventCalendarApp.save_note = _patched_save_note
BloombergEventCalendarApp.clear_note = _patched_clear_note

# ---------------------------------------------------------------------------
# Patch update: central-bank calendar support and flagged-event alarm monitor
# ---------------------------------------------------------------------------

APP_CSS = APP_CSS + """
<style>
.alarm-section-title {
    color: var(--muted);
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    margin: 8px 0 6px 0;
}
.alarm-panel-box {
    border: 1px solid var(--line);
    border-radius: 12px;
    background: var(--panel-3);
    padding: 10px 12px;
}
.alarm-status {
    color: var(--text);
    font-size: 12px;
    line-height: 1.45;
    margin-bottom: 8px;
}
.alarm-feed {
    max-height: 172px;
    overflow-y: auto;
    border: 1px solid var(--line-soft);
    border-radius: 10px;
    background: rgba(10, 18, 32, 0.62);
    padding: 6px 10px;
}
.alarm-feed-item {
    padding: 7px 0;
    border-bottom: 1px solid rgba(36, 50, 74, 0.8);
    font-size: 12px;
    line-height: 1.35;
}
.alarm-feed-item:last-child {
    border-bottom: none;
}
.alarm-feed-item.alarm-yellow {
    color: #fde68a;
}
.alarm-feed-item.alarm-red {
    color: #fecaca;
}
.alarm-feed-item.alarm-green {
    color: #86efac;
}
</style>
"""


def _patched_fetch_economic_calendar_with_central_banks(self, start_ts, end_ts):
    data_range = self.bq.func.range(start_ts.strftime("%Y-%m-%d"), end_ts.strftime("%Y-%m-%d"))
    calendar_types = getattr(
        self,
        "eco_calendar_types",
        globals().get("CUSTOM_ECO_CALENDAR_TYPES", ("central_banks", "ECONOMIC_RELEASES")),
    )
    if isinstance(calendar_types, str):
        calendar_types = [calendar_types]

    data_item = self.bq.data.calendar(type=list(calendar_types), dates=data_range)
    request = self.bq.Request(self.eco_universe, data_item)
    response = self.bq.execute(request)
    if not response:
        return pd.DataFrame()
    return pd.concat([item.df() for item in response], axis=1).reset_index()


def _patched_prepare_economic_events_with_calendar_alias(self, df):
    if df is None or df.empty:
        return pd.DataFrame(columns=self.master_columns)

    release_date_col = pick_column(df.columns, ["RELEASE_DATE", "Release Date", "DATE", "Date"])
    time_col = pick_column(df.columns, ["RELEASE_TIME", "Release Time", "TIME", "Time"])
    country_col = pick_column(df.columns, ["COUNTRY_NAME", "Country Name", "COUNTRY", "Country"])
    event_col = pick_column(df.columns, ["EVENT_NAME", "EventName", "EVENT", "Event"])
    period_col = pick_column(df.columns, ["PERIOD", "Period"])
    survey_col = pick_column(df.columns, ["SURVEY_MEDIAN", "Survey Median", "SURVEY", "Survey", "Expected"])
    actual_col = pick_column(df.columns, ["ACTUAL", "Actual"])
    prior_col = pick_column(df.columns, ["PRIOR", "Prior"])
    revision_col = pick_column(df.columns, ["REVISION", "Revision"])
    scaling_col = pick_column(df.columns, ["SCALING_FACTOR", "Scaling Factor", "SCALING", "Scaling"])
    ticker_col = pick_column(df.columns, ["Calendar", "CALENDAR_ID"])

    country_series = self._series(df, country_col, None)
    period_series = self._series(df, period_col, None)
    title_series = self._series(df, event_col, None)
    ticker_series = self._series(df, ticker_col, None)

    subtitle_parts = []
    for country_value, period_value, ticker_value in zip(country_series, period_series, ticker_series):
        parts = []
        country_text = text_or_blank(country_value, strip=True)
        period_text = text_or_blank(period_value, strip=True)
        ticker_text = text_or_blank(ticker_value, strip=True)
        if country_text:
            parts.append(country_text)
        if period_text:
            parts.append(f"Period ({period_text})")
        if ticker_text:
            parts.append(f"Ticker ({ticker_text})")
        subtitle_parts.append(" | ".join(parts) or None)

    release_dates = pd.to_datetime(self._series(df, release_date_col), errors="coerce")
    time_values = self._series(df, time_col, "")
    out = pd.DataFrame(index=df.index)
    out["event_id"] = [
        make_event_id("ECO", c, t, p, d, tm)
        for c, t, p, d, tm in zip(country_series, title_series, period_series, release_dates, time_values)
    ]
    out["event_date"] = release_dates
    out["event_time"] = self._series(df, time_col, None)
    out["category"] = "Economic"
    out["title"] = title_series
    out["subtitle"] = pd.Series(subtitle_parts, index=df.index).replace("", pd.NA)
    out["country"] = country_series.replace("", pd.NA)
    out["ticker"] = ticker_series
    out["company"] = None
    out["period"] = period_series.replace("", pd.NA)
    out["survey"] = self._series(df, survey_col, None)
    out["actual"] = self._series(df, actual_col, None)
    out["prior"] = self._series(df, prior_col, None)
    out["revision"] = self._series(df, revision_col, None)
    out["scaling_factor"] = self._series(df, scaling_col, None)
    out["market_cap_usd"] = None
    out["future_price"] = None
    out["currency"] = None
    out["underlying_ticker"] = None
    out["delivery_date"] = None
    out["expire_group_id"] = None
    out["orig_ids"] = None
    out = out[out["event_date"].notna() & out["title"].notna()]
    return out[self.master_columns]


# ---------------------------------------------------------------------------
# Alarm timestamp parsing
# ---------------------------------------------------------------------------

def _alarm_parse_timestamp(self, row):
    event_date = pd.to_datetime(row.get("event_date"), errors="coerce")
    if pd.isna(event_date):
        return pd.NaT

    time_text = text_or_blank(row.get("event_time"), strip=True)
    if not time_text:
        return pd.NaT

    cleaned = re.sub(r"\s+", " ", time_text.upper()).strip()
    if cleaned in ("", "N/A", "NAT", "NONE", "TBD", "TENTATIVE", "ALL DAY", "ALL-DAY", "DAY"):
        return pd.NaT

    cleaned = re.sub(r"[^A-Z0-9:\s]", "", cleaned).strip()

    for fmt in ("%H:%M:%S", "%H:%M%S", "%H:%M %p", "%I:%M %p", "%H:%M", "%I %p"):
        try:
            parsed = datetime.strptime(cleaned, fmt)
            return event_date.normalize() + pd.Timedelta(
                hours=parsed.hour,
                minutes=parsed.minute,
                seconds=parsed.second,
            )
        except Exception:
            continue

    compact = re.fullmatch(r"(\d{1,2})(\d{2})", cleaned)
    if compact:
        hour = int(compact.group(1))
        minute = int(compact.group(2))
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return event_date.normalize() + pd.Timedelta(hours=hour, minutes=minute)

    return pd.NaT


# ---------------------------------------------------------------------------
# Alarm candidates
# ---------------------------------------------------------------------------

def _alarm_candidates(self):
    base_df = self.events if self.events is not None and not self.events.empty else self.filtered_events
    if base_df is None or base_df.empty:
        return pd.DataFrame(columns=list(self.master_columns()) + ["alarm_timestamp"])

    flagged_df = base_df[base_df["event_id"].map(self._row_flagged)].copy()
    if flagged_df.empty:
        return pd.DataFrame(columns=list(self.master_columns()) + ["alarm_timestamp"])

    flagged_df["alarm_timestamp"] = flagged_df.apply(lambda row: self._alarm_parse_timestamp(row), axis=1)
    flagged_df = flagged_df[flagged_df["alarm_timestamp"].notna()].copy()
    if flagged_df.empty:
        return flagged_df

    now_ts = pd.Timestamp.now().floor("s") - pd.Timedelta(minutes=1)
    flagged_df = flagged_df[flagged_df["alarm_timestamp"] >= now_ts].copy()
    if flagged_df.empty:
        return flagged_df

    flagged_df = flagged_df.sort_values(["alarm_timestamp", "category", "title"]).reset_index(drop=True)
    return flagged_df


# ---------------------------------------------------------------------------
# Render alarm monitor
# ---------------------------------------------------------------------------

def _render_alarm_monitor(self):
    if not hasattr(self, "alarm_out"):
        return

    enabled = bool(getattr(self, "alarm_enabled_toggle", None).value) if hasattr(self, "alarm_enabled_toggle") else True
    sound_enabled = bool(getattr(self, "alarm_sound_checkbox", None).value) if hasattr(self, "alarm_sound_checkbox") else True
    poll_seconds = max(3, int(getattr(self, "alarm_poll_seconds", globals().get("ALARM_CHECK_INTERVAL_SECONDS", 5))))
    trigger_window = max(poll_seconds * 2, int(globals().get("ALARM_TRIGGER_WINDOW_SECONDS", 15)))
    key = getattr(self, "_alarm_monitor_key", "market-event-calendar-flag-alerts")
    candidates = self._alarm_candidates()

    schedule = []
    preview_items = []
    for row in candidates.to_dict("records"):
        event_ts = row.get("alarm_timestamp")
        if pd.isna(event_ts):
            continue
        event_iso = pd.Timestamp(event_ts).strftime("%Y-%m-%dT%H:%M:%S")
        event_label = text_or_blank(row.get("title"), strip=True) or "Untitled event"
        country_code = self._display_country_code(row) if hasattr(self, "_display_country_code") else text_or_blank(row.get("country"), strip=True)
        time_text = text_or_blank(row.get("event_time"), strip=True)
        category_text = text_or_blank(row.get("category"), strip=True)
        note_text = text_or_blank(self.state.get(row["event_id"], {}).get("note"), strip=True)
        if len(note_text) > 120:
            note_text = note_text[:117] + "..."
        important = bool(self.is_event_important(row["event_id"], row=row))
        preview_badges = []
        if important:
            preview_badges.append("KEY")
        preview_badges.append("FLAG")
        badge_text = " ".join(preview_badges)
        preview_sub = " \u00b7 ".join([v for v in [pd.Timestamp(event_ts).strftime("%Y-%m-%d %H:%M"), country_code, category_text, badge_text] if v])

        preview_html = (
            '<div class="alarm-feed-item alarm-green">'
            f"<b>{safe_html(event_label)}</b><br>"
            f'<span style="color:#94a3b8">{safe_html(preview_sub)}</span>'
        )
        if note_text:
            preview_html += f'<br><span style="color:#c8ddea">{safe_html(note_text)}</span>'
        preview_html += "</div>"
        preview_items.append(preview_html)

        schedule.append({
            "event_id": row["event_id"],
            "event_iso": event_iso,
            "title": event_label,
            "country": country_code,
            "time_text": time_text,
            "category": category_text,
            "important": important,
            "note": note_text,
        })

    if not preview_items:
        preview_html = '<div class="empty-state">No future flagged events with precise times are currently armed.</div>'
    else:
        preview_html = "".join(preview_items[:12])

    if hasattr(self, "alarm_status_html"):
        if schedule:
            self.alarm_status_html.value = (
                '<div class="event-app page-caption">'
                f'Alarm monitor {"armed" if enabled else "paused"} for <b>{len(schedule)}</b> flagged timed event'
                f'{"" if len(schedule) == 1 else "s"}. Warnings fire at <b>T-5:00</b> and <b>T-1:30</b>. '
                "Timeless or TBD events are also skipped. Alarm timing uses the notebook/browser local clock against the displayed event date and time."
                "</div>"
            )
        else:
            self.alarm_status_html.value = (
                '<div class="event-app page-caption">'
                "Alarm monitor watches only flagged events with a concrete date and time. "
                "Flag an event to arm alerts; timeless or TBD events are skipped."
                "</div>"
            )

    schedule_json = json.dumps(schedule, ensure_ascii=False).replace("<", "\\u003c")
    html = f"""
<div class="event-app alarm-panel-box">
    <div id="__KEY__-status" class="alarm-status"></div>
    <div class="alarm-section-title">Triggered Alerts</div>
    <div id="__KEY__-alerts" class="alarm-feed">
        <div class="empty-state">No alerts have fired in this session.</div>
    </div>
    <div class="alarm-section-title">Armed Flagged Events</div>
    <div id="__KEY__-upcoming" class="alarm-feed">__PREVIEW__</div>
</div>
<script>
(function(){{
    var key = "__KEY__";
    var enabled = __ENABLED__;
    var soundEnabled = __SOUND_ENABLED__;
    var pollMs = __POLL_MS__;
    var triggerWindow = __TRIGGER_WINDOW__;
    var schedule = __SCHEDULE__;
    var storeRoot = window.__marketEventAlarmStore = window.__marketEventAlarmStore || {{}};
    var existing = storeRoot[key] || {{}};
    if (existing.timer) {{
        window.clearInterval(existing.timer);
    }}

    var store = {{
        timer: null,
        fired: existing.fired || {{}}
    }};
    storeRoot[key] = store;
    var statusEl = document.getElementById(key + "-status");
    var alertsEl = document.getElementById(key + "-alerts");
    var upcomingEl = document.getElementById(key + "-upcoming");

    function esc(value) {{
        var span = document.createElement("span");
        span.textContent = value === null ? "" : String(value);
        return span.innerHTML;
    }}

    function beep(count) {{
        if (!soundEnabled) {{
            return;
        }}
        try {{
            var AudioContextCtor = window.AudioContext || window.webkitAudioContext;
            if (!AudioContextCtor) {{
                return;
            }}
            var ctx = new AudioContextCtor();
            for (var i = 0; i < count; i += 1) {{
                (function(idx) {{
                    var osc = ctx.createOscillator();
                    var gain = ctx.createGain();
                    osc.type = "sine";
                    osc.frequency.value = idx === 0 ? 880 : 1320;
                    gain.gain.value = 0.0001;
                    osc.connect(gain);
                    gain.connect(ctx.destination);
                    var start = ctx.currentTime + (idx * 0.22);
                    gain.gain.exponentialRampToValueAtTime(0.11, start + 0.02);
                    gain.gain.exponentialRampToValueAtTime(0.001, start + 0.18);
                    osc.start(start);
                    osc.stop(start + 0.20);
                }})(i);
            }}
        }} catch (err) {{
            try {{
                console.log("Market event alarm beep unavailable:", err);
            }} catch (ignore) {{}}
        }}
    }}

    function fmtRemaining(seconds) {{
        if (seconds <= 0) {{
            return "now";
        }}
        var mins = Math.floor(seconds / 60);
        var secs = seconds % 60;
        if (mins >= 60) {{
            var hours = Math.floor(mins / 60);
            var remMins = mins % 60;
            return hours + "h " + remMins + "m";
        }}
        return mins + "m " + String(secs).padStart(2, "0") + "s";
    }}

    function upcomingHtml(nowMs) {{
        var upcoming = schedule
            .map(function(evt) {{
                var eventMs = new Date(evt.event_iso).getTime();
                var remaining = Math.round((eventMs - nowMs) / 1000);
                if (remaining < -60) {{
                    return null;
                }}
                return {{
                    eventMs: eventMs,
                    remaining: remaining,
                    evt: evt
                }};
            }})
            .filter(function(item) {{ return item !== null; }})
            .sort(function(a, b) {{ return a.eventMs - b.eventMs; }})
            .slice(0, 12);

        if (!upcoming.length) {{
            return '<div class="empty-state">No future flagged events with precise times are currently armed.</div>';
        }}

        return upcoming.map(function(item) {{
            var evt = item.evt;
            var stamp = new Date(item.eventMs);
            var title = esc(evt.title);
            var subBits = [
                esc(stamp.getFullYear() + "-" + String(stamp.getMonth() + 1).padStart(2, "0") + "-" + String(stamp.getDate()).padStart(2, "0") + " " +
                esc(String(stamp.getHours()).padStart(2, "0") + ":" + String(stamp.getMinutes()).padStart(2, "0")))
            ];
            if (evt.country) {{
                subBits.push(esc(evt.country));
            }}
            if (evt.category) {{
                subBits.push(esc(evt.category));
            }}
            subBits.push("FLAG");
            if (evt.important) {{
                subBits.push("KEY");
            }}
            var noteHtml = evt.note ? "<br><span style='color:#c8ddea'>" + esc(evt.note) + "</span>" : "";
            return (
                '<div class="alarm-feed-item alarm-green">'
                + "<b>" + title + "</b><br>"
                + '<span style="color:#94a3b8">' + subBits.join(" \\u00b7 ") + " \\u00b7 " + esc(fmtRemaining(Math.max(item.remaining, 0))) + "</span>"
                + noteHtml
                + "</div>"
            );
        }}).join("");
    }}

    function pushAlert(message, cssClass) {{
        if (!alertsEl) {{
            return;
        }}
        if (alertsEl.querySelector(".empty-state")) {{
            alertsEl.innerHTML = "";
        }}
        var row = document.createElement("div");
        row.className = "alarm-feed-item " + (cssClass || "");
        row.innerHTML = "<b>" + esc(new Date().toLocaleTimeString()) + "</b> \\u00b7 " + esc(message);
        alertsEl.insertBefore(row, alertsEl.firstChild);
        while (alertsEl.children.length > 8) {{
            alertsEl.removeChild(alertsEl.lastChild);
        }}
    }}

    function refresh() {{
        var nowMs = Date.now();
        var activeCount = schedule.filter(function(evt) {{
            return new Date(evt.event_iso).getTime() >= nowMs - 60000;
        }}).length;

        if (statusEl) {{
            if (enabled) {{
                statusEl.innerHTML =
                    "Alarm monitor armed for <b>" + activeCount + "</b> flagged timed event"
                    + (activeCount == 1 ? "" : "s")
                    + ". Warnings fire at <b>T-5:00</b> and <b>T-1:30</b>.";
            }} else {{
                statusEl.innerHTML =
                    "Alarm monitor paused. Armed events remain listed below, but no sound or countdown alarms will fire until re-enabled.";
            }}
        }}

        if (upcomingEl) {{
            upcomingEl.innerHTML = upcomingHtml(nowMs);
        }}

        if (!enabled) {{
            return;
        }}

        schedule.forEach(function(evt) {{
            var eventMs = new Date(evt.event_iso).getTime();
            var remaining = Math.round((eventMs - nowMs) / 1000);

            [
                {{lead: 300, label: "5-minute", css: "alarm-yellow", soundCount: 1}},
                {{lead: 90, label: "90-second", css: "alarm-red", soundCount: 2}}
            ].forEach(function(rule) {{
                var firedKey = evt.event_id + "_" + rule.lead;
                if (store.fired[firedKey]) {{
                    return;
                }}
                if (remaining <= rule.lead && remaining > (rule.lead - triggerWindow)) {{
                    store.fired[firedKey] = true;
                    var bits = [rule.label + " warning", evt.title];
                    if (evt.time_text) {{
                        bits.push(evt.time_text);
                    }}
                    if (evt.country) {{
                        bits.push(evt.country);
                    }}
                    pushAlert(bits.join(" \\u00b7 "), rule.css);
                    beep(rule.soundCount);
                    try {{
                        console.log("MARKET EVENT ALARM:", bits.join(" \\u00b7 "));
                    }} catch (ignore) {{}}
                }}
            }});
        }});

        if (enabled) {{
            store.timer = window.setInterval(refresh, pollMs);
        }}
    }}();
}})();
</script>
"""

    html = html.replace("__KEY__", key)
    html = html.replace("__PREVIEW__", preview_html)
    html = html.replace("__ENABLED__", "true" if enabled else "false")
    html = html.replace("__SOUND_ENABLED__", "true" if sound_enabled else "false")
    html = html.replace("__POLL_MS__", str(int(poll_seconds * 1000)))
    html = html.replace("__TRIGGER_WINDOW__", str(trigger_window))
    html = html.replace("__SCHEDULE__", schedule_json)

    with self.alarm_out:
        clear_output(wait=True)
        display(HTML(html))


# ---------------------------------------------------------------------------
# Alarm toggle callbacks
# ---------------------------------------------------------------------------

def _alarm_on_toggle(self, change):
    if not hasattr(self, "alarm_enabled_toggle"):
        return
    enabled = bool(change.get("new"))
    self.alarm_enabled_toggle.icon = "bell" if enabled else "bell-slash"
    self.alarm_enabled_toggle.button_style = "warning" if enabled else ""
    self.render_alarm_monitor()
    self.log(
        "Alarm monitor armed for flagged timed events."
        if enabled else
        "Alarm monitor paused."
    )


def _alarm_on_sound_change(self, change):
    self.render_alarm_monitor()


# ---------------------------------------------------------------------------
# Alarm-aware build_widgets
# ---------------------------------------------------------------------------

_original_build_widgets_for_alarm = BloombergEventCalendarApp.build_widgets


def _patched_build_widgets_with_alarm(self, months_per_page_default):
    _original_build_widgets_for_alarm(self, months_per_page_default)

    default_alarm_enabled = bool(globals().get("ALARM_ENABLED_BY_DEFAULT", True))
    default_sound_enabled = bool(globals().get("ALARM_SOUND_ENABLED_BY_DEFAULT", True))

    self.alarm_poll_seconds = max(3, int(globals().get("ALARM_CHECK_INTERVAL_SECONDS", 5)))
    self._alarm_monitor_key = "market-event-calendar-flag-alerts"

    self.alarm_enabled_toggle = widgets.ToggleButton(
        description="Alarm Monitor",
        value=default_alarm_enabled,
        icon="bell" if default_alarm_enabled else "bell-slash",
        button_style="warning" if default_alarm_enabled else "",
        layout=widgets.Layout(width="170px"),
    )

    self.alarm_sound_checkbox = widgets.Checkbox(
        description="Sound",
        value=default_sound_enabled,
        layout=widgets.Layout(width="110px"),
    )

    self.alarm_status_html = widgets.HTML(
        '<div class="event-app page-caption">'
        "Alarm monitor watches flagged events that have a concrete date and time. "
        "Warnings fire at T-5:00 and T-1:30. Timeless or TBD events are skipped."
        "</div>"
    )

    self.alarm_out = widgets.Output(layout=widgets.Layout(width="100%"))

    alarm_title = widgets.HTML('<div class="event-app section-title">Flagged Event Alarm</div>')
    alarm_controls = widgets.HBox([self.alarm_enabled_toggle, self.alarm_sound_checkbox])
    alarm_note = widgets.HTML(
        '<div class="event-app page-caption">'
        "Uses the notebook/browser local clock against the event date and time shown in the app. "
        "Only flagged events are armed."
        "</div>"
    )

    self.alarm_section = widgets.VBox([alarm_title, alarm_controls, alarm_note, self.alarm_status_html, self.alarm_out])
    self.controls_panel.children = tuple(list(self.controls_panel.children) + [self.alarm_section])

    self.render_alarm_monitor()


# ---------------------------------------------------------------------------
# Alarm-aware wire_events
# ---------------------------------------------------------------------------

_original_wire_events_for_alarm = BloombergEventCalendarApp._wire_events


def _patched_wire_events_with_alarm(self):
    _original_wire_events_for_alarm(self)
    if hasattr(self, "alarm_enabled_toggle"):
        self.alarm_enabled_toggle.observe(self.on_alarm_toggle, names="value")
    if hasattr(self, "alarm_sound_checkbox"):
        self.alarm_sound_checkbox.observe(self.on_alarm_sound_change, names="value")


# ---------------------------------------------------------------------------
# Alarm-aware apply_filters
# ---------------------------------------------------------------------------

_original_apply_filters_for_alarm = BloombergEventCalendarApp.apply_filters


def _patched_apply_filters_with_alarm(self, _=None):
    _original_apply_filters_for_alarm(self, _)
    try:
        self.render_alarm_monitor()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Alarm-aware on_flag_toggled
# ---------------------------------------------------------------------------

_original_on_flag_toggled_for_alarm = BloombergEventCalendarApp.on_flag_toggled


def _patched_on_flag_toggled_with_alarm(self, change):
    _original_on_flag_toggled_for_alarm(self, change)
    try:
        self.render_alarm_monitor()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Alarm monkey-patch assignments (module level)
# ---------------------------------------------------------------------------

BloombergEventCalendarApp.fetch_economic_calendar = _patched_fetch_economic_calendar_with_central_banks
BloombergEventCalendarApp.prepare_economic_events = _patched_prepare_economic_events_with_calendar_alias
BloombergEventCalendarApp._alarm_parse_timestamp = _alarm_parse_timestamp
BloombergEventCalendarApp._alarm_candidates = _alarm_candidates
BloombergEventCalendarApp.render_alarm_monitor = _render_alarm_monitor
BloombergEventCalendarApp.on_alarm_toggle = _alarm_on_toggle
BloombergEventCalendarApp.on_alarm_sound_change = _alarm_on_sound_change
BloombergEventCalendarApp._build_widgets = _patched_build_widgets_with_alarm
BloombergEventCalendarApp._wire_events = _patched_wire_events_with_alarm
BloombergEventCalendarApp.apply_filters = _patched_apply_filters_with_alarm
BloombergEventCalendarApp.on_flag_toggled = _patched_on_flag_toggled_with_alarm

# ---------------------------------------------------------------------------
# Final patch: one-cell, fast, lazy-rendered tabular explorer
# ---------------------------------------------------------------------------

APP_CSS = APP_CSS + """
<style>
.fast-tabs-widget .p-TabBar-tab,
.fast-tabs-widget .lm-TabBar-tab {
    background: rgba(18, 24, 32, 0.98);
    border: 1px solid rgba(85, 103, 129, 0.34);
    color: #d7dde6;
}
.fast-tabs-widget .p-TabBar-tab.p-mod-current,
.fast-tabs-widget .lm-TabBar-tab.lm-mod-current {
    background: linear-gradient(180deg, rgba(35, 51, 74, 0.98), rgba(18, 28, 40, 0.98));
    border-color: rgba(112, 164, 255, 0.52);
    color: #fff;
}
.fast-explorer-caption {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.5;
    margin-bottom: 8px;
}
.fast-explorer-pager {
    display: flex;
    align-items: center;
    gap: 8px;
    margin-bottom: 10px;
}
.fast-explorer-pager .widget-button button {
    background: rgba(19, 26, 40, 0.98);
    color: #dce614;
    border: 1px solid rgba(93, 111, 136, 0.34);
    border-radius: 10px;
    padding: 0 12px;
}
.fast-explorer-pager .widget-button button:hover {
    background: rgba(28, 40, 58, 0.98);
    border-color: rgba(112, 164, 255, 0.43);
}
.fast-explorer-shell {
    border: 1px solid rgba(82, 99, 124, 0.34);
    border-radius: 14px;
    background: linear-gradient(180deg, rgba(14, 20, 28, 0.99), rgba(10, 15, 22, 0.99));
    overflow: hidden;
}
.fast-explorer-scroller {
    height: 620px;
    overflow: auto;
    background: rgba(10, 15, 22, 0.98);
}
.fast-header-row {
    display: flex;
    align-items: stretch;
    border-bottom: 1px solid rgba(88, 104, 130, 0.34);
    background: linear-gradient(180deg, rgba(24, 31, 40, 0.98), rgba(17, 23, 31, 0.98));
}
.fast-header-cell,
.fast-cell-block {
    box-sizing: border-box;
    border-right: 1px solid rgba(78, 94, 120, 0.18);
}
.fast-header-cell:last-child,
.fast-cell-block:last-child {
    border-right: none;
}
.fast-header-cell {
    padding: 9px 10px;
    color: #93a2b6;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}
.fast-data-row {
    display: flex;
    align-items: stretch;
    border-bottom: 1px solid rgba(78, 92, 114, 0.14);
    background: rgba(13, 18, 24, 0.98);
}
.fast-data-row.row-even {
    background: rgba(15, 21, 29, 0.98);
}
.fast-data-row.important-row {
    background: rgba(250, 204, 21, 0.17);
}
.fast-data-row.flagged-row {
    background: rgba(239, 68, 68, 0.19);
}
.fast-data-row.selected-row {
    box-shadow: inset 0 0 0 1px rgba(112, 164, 255, 0.65);
}
.fast-cell-block {
    padding: 9px 10px;
    min-height: 48px;
    color: #e7edf5;
    display: flex;
    align-items: center;
    overflow: hidden;
}
.fast-cell-right {
    justify-content: flex-end;
    font-variant-numeric: tabular-nums;
}
.fast-cell-center {
    justify-content: center;
}
.fast-date-stack {
    display: flex;
    flex-direction: column;
    line-height: 1.18;
}
.fast-date-main {
    color: #eef3fb;
    font-size: 12px;
    font-weight: 700;
    font-variant-numeric: tabular-nums;
}
.fast-date-sub {
    color: #97a6ba;
    font-size: 11px;
    margin-top: 3px;
    font-variant-numeric: tabular-nums;
}
.fast-country-text {
    color: #dbe4f2;
    font-size: 12px;
    font-weight: 600;
}
.fast-event-button button {
    width: 100%;
    background: transparent;
    color: #eef3fb;
    border: none;
    border-radius: 0;
    text-align: left;
    padding: 0;
    min-height: 28px;
    box-shadow: none;
    font-size: 12px;
    font-weight: 700;
}
.fast-event-button button:hover {
    color: #fff;
}
.fast-event-button button:focus {
    outline: none;
    box-shadow: none;
}
.fast-number-text {
    width: 100%;
    text-align: right;
    font-variant-numeric: tabular-nums;
}
.fast-day-divider {
    padding: 8px 12px;
    color: #dbe412;
    font-size: 12px;
    font-weight: 700;
    background: linear-gradient(180deg, rgba(23, 31, 42, 0.98), rgba(17, 23, 31, 0.98));
    border-top: 1px solid rgba(88, 104, 130, 0.24);
    border-bottom: 1px solid rgba(88, 104, 130, 0.18);
}
.fast-empty-state {
    padding: 18px 16px;
}
.fast-pane-label {
    color: var(--muted);
    font-size: 12px;
}
</style>
"""

# =============================================================================
# Fast explorer patch: replace Output-based tab rendering with direct widget construction
# =============================================================================

# =============================================================================
# Fast explorer patch: replace Output-based tab rendering with direct widget construction
# =============================================================================

_FAST_TAB_ORDER = ["ALL", "ECONOMIC", "EARNINGS", "COMMODITIES", "HOLIDAYS", "IMPORTANT"]

_FAST_OUTPUT_ATTRS = {
    "ALL": "all_list",
    "ECONOMIC": "econ_list",
    "EARNINGS": "earnings_list",
    "COMMODITIES": "comdty_list",
    "HOLIDAYS": "holiday_list",
    "IMPORTANT": "watch_list",
}

_FAST_PAGE_SIZE = 28

_FAST_COLUMNS = [
    ("Date / Time", "15%", "datetime", "left"),
    ("Country", "10%", "country", "left"),
    ("Event Name", "33%", "event", "left"),
    ("Period", "10%", "period", "left"),
    ("Prior", "8%", "prior", "right"),
    ("Survey", "8%", "survey", "right"),
    ("Actual", "8%", "actual", "right"),
    ("Beat / Miss", "8%", "outcome", "center"),
]

_FAST_TABLE_WIDTH = "100%"


def _fast_cell_widget(label, width, align="left", html_content=False, extra_class="fast-header-cell"):
    if label is None:
        label = ""
    if html_content:
        inner = label
    else:
        inner = safe_html(str(label))
    align_class = {"left": "fast-left", "center": "fast-center", "right": "fast-right"}.get(align, "fast-left")
    html = f'<div class="event-app {extra_class} {align_class}">{inner}</div>'
    col = widgets.HTML(html, layout=widgets.Layout(width=width, min_width="0", flex=f"0 0 {width}"))
    col.add_class("fast-col")
    return col


def _fast_header_row_widget():
    cells = []
    for label, width, key, align in _FAST_COLUMNS:
        align = "left"
        if key == "outcome":
            align = "center"
        elif key in {"prior", "survey", "actual"}:
            align = "right"
        col = _fast_cell_widget(label, width, align=align, html_content=False, extra_class="fast-header-cell")
        cells.append(col)
    row = widgets.HBox(cells, layout=widgets.Layout(width=_FAST_TABLE_WIDTH))
    row.add_class("fast-header-row")
    return row


def _fast_day_header_widget(label):
    return widgets.HTML(
        f'<div class="event-app fast-day-divider">{safe_html(label)}</div>',
        layout=widgets.Layout(width=_FAST_TABLE_WIDTH),
    )


def _fast_row_widget(self, row, stripe_even=False):
    cells = [
        _fast_cell_widget(row.get("display_datetime_html"), _FAST_COLUMNS[0][1], align="left", html_content=True),
        _fast_cell_widget(row.get("display_country"), _FAST_COLUMNS[1][1], align="left", html_content=False),
        _fast_event_button_widget(self, row, _FAST_COLUMNS[2][1]),
        _fast_cell_widget(row.get("display_period"), _FAST_COLUMNS[3][1], align="left", html_content=False),
        _fast_cell_widget(row.get("display_prior"), _FAST_COLUMNS[4][1], align="right", html_content=False),
        _fast_cell_widget(row.get("display_survey"), _FAST_COLUMNS[5][1], align="right", html_content=False),
        _fast_cell_widget(row.get("display_actual"), _FAST_COLUMNS[6][1], align="right", html_content=False),
        _fast_cell_widget(row.get("display_outcome_html"), _FAST_COLUMNS[7][1], align="center", html_content=True),
    ]
    row_box = widgets.HBox(cells, layout=widgets.Layout(width=_FAST_TABLE_WIDTH))
    row_box.add_class("fast-data-row")
    if stripe_even:
        row_box.add_class("row-even")
    if bool(row.get("_important_flag")):
        row_box.add_class("important-row")
    if bool(row.get("_flagged_flag")):
        row_box.add_class("flagged-row")
    if row.get("event_id") == self.current_event_id:
        row_box.add_class("selected-row")
    return row_box


def _fast_update_pager(self, tab_name, total_rows, start_idx, end_idx, max_page):
    current_page = self.explorer_page_by_tab.get(tab_name, 0)
    self.explorer_prev_button.disabled = current_page <= 0
    self.explorer_next_button.disabled = current_page >= max_page
    if total_rows <= 0:
        label = "Pane 1 of 1 - 0 rows"
    else:
        label = f"Pane {current_page + 1} of {max_page + 1} \u00b7 Rows {start_idx + 1}-{end_idx} of {total_rows}"
    self.explorer_page_label.value = f'<div class="event-app fast-pane-label">{safe_html(label)}</div>'


def _fast_render_tab(self, tab_name):
    output_widget = getattr(self, _FAST_OUTPUT_ATTRS[tab_name], None)
    df = getattr(self, "_explorer_frames", {}).get(tab_name)

    with output_widget:
        clear_output(wait=True)
        if df is None or df.empty:
            _fast_update_pager(self, tab_name, 0, 0, 0, 0)
            display(HTML('<div class="event-app fast-empty-state"><div class="empty-state">No events in this view.</div></div>'))
            return

        total_rows = len(df)
        max_page = max(0, math.ceil(total_rows / self.explorer_page_size) - 1)
        current_page = min(max(0, self.explorer_page_by_tab.get(tab_name, 0)), max_page)
        self.explorer_page_by_tab[tab_name] = current_page

        start_idx = current_page * self.explorer_page_size
        end_idx = min(total_rows, start_idx + self.explorer_page_size)
        page_df = df.iloc[start_idx:end_idx].copy()

        _fast_update_pager(self, tab_name, total_rows, start_idx, end_idx, max_page)

        table_children = [_fast_header_row_widget()]
        stripe_even = False

        for day_value, day_group in page_df.groupby("_day_key", sort=True):
            day_label = pd.Timestamp(day_value).strftime("%a %Y-%m-%d")
            total = len(day_group)
            key_count = int(day_group["_important_flag"].sum())
            flag_count = int(day_group["_flagged_flag"].sum())
            header_bits = [f"{day_label} - {total} event{'s' if total > 1 else ''}"]
            if key_count:
                header_bits.append(f"KEY({key_count})")
            if flag_count:
                header_bits.append(f"FLAG({flag_count})")
            holiday_codes = (
                day_group.loc[day_group["category"] == "Holiday", "ticker"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
            if holiday_codes:
                header_bits.append("Closed: " + ", ".join(sorted(holiday_codes)))
            table_children.append(_fast_day_header_widget(" \u00b7 ".join(header_bits)))

            for row in day_group.to_dict("records"):
                table_children.append(_fast_row_widget(self, row, stripe_even=stripe_even))
                stripe_even = not stripe_even

        table = widgets.VBox(table_children, layout=widgets.Layout(width=_FAST_TABLE_WIDTH))
        scroller = widgets.Box([table], layout=widgets.Layout(width="100%", height="620px", overflow="auto"))
        scroller.add_class("fast-explorer-scroller")
        shell = widgets.Box([scroller], layout=widgets.Layout(width="100%"))
        shell.add_class("fast-explorer-shell")
        display(shell)


def _fast_prev_explorer_page(self, _=None):
    tab_name = _fast_active_tab_name(self)
    current_page = self.explorer_page_by_tab.get(tab_name, 0)
    if current_page <= 0:
        return
    self.explorer_page_by_tab[tab_name] = current_page - 1
    _fast_render_tab(self, tab_name)


def _fast_next_explorer_page(self, _=None):
    tab_name = _fast_active_tab_name(self)
    df = getattr(self, "_explorer_frames", {}).get(tab_name)
    total_rows = 0 if df is None else len(df)
    max_page = max(0, math.ceil(max(total_rows, 1) / self.explorer_page_size) - 1)
    current_page = self.explorer_page_by_tab.get(tab_name, 0)
    if current_page >= max_page:
        return
    self.explorer_page_by_tab[tab_name] = current_page + 1
    _fast_render_tab(self, tab_name)


_previous_build_widgets_fast = BloombergEventCalendarApp._build_widgets


def _fast_build_widgets(self, months_per_page_default):
    _previous_build_widgets_fast(self, months_per_page_default)

    self.explorer_page_size = int(globals().get("EXPLORER_PAGE_SIZE", _FAST_PAGE_SIZE))
    self.explorer_page_by_tab = {tab_name: 0 for tab_name in _FAST_TAB_ORDER}
    self.explorer_prev_button = widgets.Button(description="Previous Pane", icon="arrow-left")
    self.explorer_next_button = widgets.Button(description="Next Pane", icon="arrow-right")
    self.explorer_page_label = widgets.HTML('<div class="event-app fast-pane-label">Pane 1 of 1 - 0 rows</div>')

    self.tabs.add_class("fast-tabs-widget")

    if len(self.root.children) >= 5:
        list_detail_box = self.root.children[4]
        if len(list_detail_box.children) >= 2:
            list_card, detail_card = list_detail_box.children
            list_card.layout.width = "58%"
            detail_card.layout.width = "42%"

    self.render_monthly_page()
    _fast_render_tab(self, _fast_active_tab_name(self))


_previous_render_monthly_fast = BloombergEventCalendarApp.render_monthly_page


def _fast_render_monthly_page(self):
    try:
        start_ts, end_ts = self._validate_dates()
        months = self._all_months_in_range(start_ts, end_ts)
        page_size = max(1, int(self.months_per_page.value or 1))
        max_start = max(0, len(months) - page_size)
        current_start = min(max(0, getattr(self, "month_page_start", 0)), max_start)
        visible_months = months[current_start:current_start + page_size]
        cal = calendar.Calendar(firstweekday=0)
        max_weeks = max(
            len(cal.monthdatescalendar(period.year, period.month)) for period in visible_months
        ) or 6

        target_height = max(980, 280 + max_weeks * 122)
        self.day_out.layout.height = str(target_height) + "px"
        self.day_out.layout.overflow = "auto"
        self.day_panel.layout.min_height = str(target_height + 64) + "px"
    except Exception:
        pass
    return _previous_render_monthly_fast(self)


_previous_wire_events_fast = BloombergEventCalendarApp._wire_events


def _fast_wire_events(self):
    _previous_wire_events_fast(self)
    self.explorer_prev_button.on_click(self.previous_explorer_pane)
    self.explorer_next_button.on_click(self.next_explorer_pane)


_previous_apply_filters_fast = BloombergEventCalendarApp.apply_filters


def _fast_apply_filters(self, _=None):
    self.explorer_page_by_tab = {tab_name: 0 for tab_name in _FAST_TAB_ORDER}
    self.month_page_start = 0
    return _previous_apply_filters_fast(self, _)


def _fast_refresh_lists(self):
    _fast_prepare_explorer_frames(self)
    self.render_watchlist_box()

    filtered_ids = set(self.filtered_events["event_id"]) if self.filtered_events is not None and not self.filtered_events.empty else set()
    if self.current_event_id and self.current_event_id not in filtered_ids:
        self.current_event_id = None

    if self.current_event_id is None and self.filtered_events is not None and not self.filtered_events.empty:
        all_df = getattr(self, "_explorer_frames", {}).get("ALL", pd.DataFrame())
        if all_df is not None and not all_df.empty:
            first_row = all_df.iloc[0]
            self.current_event_id = first_row["event_id"]
            event_date = first_row.get("event_date")
            if not is_missing(event_date):
                self.selected_day = pd.Timestamp(event_date).date()

    if self.current_event_id:
        self.render_detail(self.current_event_id)
    else:
        self.render_detail(None)

    _fast_render_tab(self, _fast_active_tab_name(self))
    self.render_selected_day_detail()


def _fast_get_active_list_widget(self):
    return getattr(self, _FAST_OUTPUT_ATTRS[_fast_active_tab_name(self)])


def _fast_on_tab_change(self, change):
    tab_name = _fast_active_tab_name(self)
    active_df = getattr(self, "_explorer_frames", {}).get(tab_name, pd.DataFrame())
    if active_df is not None and not active_df.empty:
        active_ids = set(active_df["event_id"])
        if self.current_event_id not in active_ids:
            first_row = active_df.iloc[0]
            self.current_event_id = first_row["event_id"]
            event_date = first_row.get("event_date")
            if not is_missing(event_date):
                self.selected_day = pd.Timestamp(event_date).date()
            self.render_detail(self.current_event_id)
            self.render_selected_day_detail()

    self.render_monthly_page()
    _fast_render_tab(self, _fast_active_tab_name(self))


def _fast_active_tab_name(self):
    idx = self.tabs.selected_index if getattr(self.tabs, "selected_index", None) is not None else 0
    idx = max(0, min(int(idx), len(_FAST_TAB_ORDER) - 1))
    return _FAST_TAB_ORDER[idx]


BloombergEventCalendarApp._build_widgets = _fast_build_widgets
BloombergEventCalendarApp._wire_events = _fast_wire_events
BloombergEventCalendarApp.apply_filters = _fast_apply_filters
BloombergEventCalendarApp.refresh_lists = _fast_refresh_lists
BloombergEventCalendarApp.get_active_list_widget = _fast_get_active_list_widget
BloombergEventCalendarApp.on_tab_change = _fast_on_tab_change
BloombergEventCalendarApp.previous_explorer_pane = _fast_prev_explorer_page
BloombergEventCalendarApp.next_explorer_pane = _fast_next_explorer_page
BloombergEventCalendarApp.render_monthly_page = _fast_render_monthly_page


# =============================================================================
# Stable explorer patch: replace Output-based tab rendering with direct widget containers
# to avoid blank rows, hidden tab content, and unnecessary horizontal scrolling
# =============================================================================

APP_CSS = APP_CSS + """
<style>
.stable-explorer-caption {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.5;
    margin-bottom: 8px;
}

.stable-tabs-widget .p-TabBar-tab,
.stable-tabs-widget .lm-TabBar-tab {
    background: rgba(18, 24, 32, 0.98);
    border: 1px solid rgba(85, 103, 129, 0.34);
    color: #d7dde6;
}

.stable-tabs-widget .p-TabBar-tab.p-mod-current,
.stable-tabs-widget .lm-TabBar-tab.lm-mod-current {
    background: linear-gradient(180deg, rgba(35, 51, 74, 0.98), rgba(18, 28, 40, 0.98));
    border-color: rgba(112, 164, 255, 0.52);
    color: #fff;
}

.stable-explorer-shell {
    width: 100%;
    height: 660px;
    overflow: auto;
    border: 1px solid rgba(82, 99, 124, 0.34);
    border-radius: 14px;
    background: linear-gradient(180deg, rgba(14, 20, 28, 0.99), rgba(10, 15, 22, 0.99));
}

.stable-header-row {
    display: flex;
    align-items: stretch;
    width: 100%;
    border-bottom: 1px solid rgba(88, 104, 130, 0.34);
    background: linear-gradient(180deg, rgba(24, 31, 40, 0.98), rgba(17, 23, 31, 0.98));
    position: sticky;
    top: 0;
    z-index: 4;
}

.stable-day-divider {
    width: 100%;
    padding: 8px 12px;
    color: #dbe4f2;
    font-size: 12px;
    font-weight: 700;
    letter-spacing: 0.01em;
    background: linear-gradient(180deg, rgba(23, 31, 42, 0.98), rgba(17, 23, 31, 0.98));
    border-top: 1px solid rgba(88, 104, 130, 0.24);
    border-bottom: 1px solid rgba(88, 104, 130, 0.18);
}

.stable-data-row {
    display: flex;
    align-items: stretch;
    width: 100%;
    border-bottom: 1px solid rgba(78, 92, 114, 0.14);
    background: rgba(13, 18, 24, 0.98);
}

.stable-data-row.row-even {
    background: rgba(15, 21, 29, 0.98);
}

.stable-data-row.important-row {
    background: rgba(250, 204, 21, 0.17);
}

.stable-data-row.flagged-row {
    background: rgba(239, 68, 68, 0.20);
}

.stable-data-row.selected-row {
    box-shadow: inset 0 0 0 1px rgba(112, 164, 255, 0.65);
}

.stable-col {
    box-sizing: border-box;
    min-width: 0;
    border-right: 1px solid rgba(78, 94, 120, 0.18);
}

.stable-col:last-child {
    border-right: none;
}

.stable-cell {
    min-height: 46px;
    padding: 8px 10px;
    color: #e7edf5;
    display: flex;
    align-items: center;
    overflow: hidden;
    white-space: normal;
    word-break: break-word;
    line-height: 1.25;
    font-size: 12px;
}

.stable-header-cell {
    min-height: 38px;
    padding: 9px 10px;
    color: #93a2b6;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}

.stable-right {
    justify-content: flex-end;
    text-align: right;
    font-variant-numeric: tabular-nums;
}

.stable-center {
    justify-content: center;
    text-align: center;
}

.stable-left {
    justify-content: flex-start;
    text-align: left;
}

.stable-date-stack {
    display: flex;
    flex-direction: column;
    line-height: 1.18;
}

.stable-date-main {
    color: #edf3fb;
    font-size: 12px;
    font-weight: 700;
    font-variant-numeric: tabular-nums;
}

.stable-date-sub {
    color: #97a6ba;
    font-size: 11px;
    margin-top: 3px;
    font-variant-numeric: tabular-nums;
}

.stable-country-text {
    color: #dbe4f2;
    font-size: 12px;
    font-weight: 600;
}

.stable-event-button button {
    width: 100%;
    min-width: 0;
    background: transparent;
    color: #eef3fb;
    border: none;
    border-radius: 0;
    text-align: left;
    padding: 0;
    box-shadow: none;
    font-size: 12px;
    font-weight: 700;
    white-space: normal;
    line-height: 1.25;
}

.stable-event-button button:hover {
    color: #ffffff;
}

.stable-event-button button:focus {
    outline: none;
    box-shadow: none;
}

.stable-empty-state {
    padding: 18px 16px;
}
</style>
"""

_STABLE_TAB_ORDER = ["ALL", "ECONOMIC", "EARNINGS", "COMMODITIES", "HOLIDAYS", "IMPORTANT"]

_STABLE_OUTPUT_ATTRS = {
    "ALL": "all_list",
    "ECONOMIC": "econ_list",
    "EARNINGS": "earnings_list",
    "COMMODITIES": "comdty_list",
    "HOLIDAYS": "holiday_list",
    "IMPORTANT": "watch_list",
}

_STABLE_PAGE_SIZE = 28

_STABLE_COLUMNS = [
    ("Date / Time", "15%", "datetime", "left"),
    ("Country", "10%", "country", "left"),
    ("Event Name", "33%", "event", "left"),
    ("Period", "10%", "period", "left"),
    ("Prior", "8%", "prior", "right"),
    ("Survey", "8%", "survey", "right"),
    ("Actual", "8%", "actual", "right"),
    ("Beat / Miss", "8%", "outcome", "center"),
]


def _stable_make_tab_container():
    container = widgets.VBox([], layout=widgets.Layout(width="100%", height="660px", overflow="auto"))
    container.add_class("stable-explorer-shell")
    return container


def _stable_active_tab_name(self):
    idx = self.tabs.selected_index if getattr(self.tabs, "selected_index", None) is not None else 0
    idx = max(0, min(int(idx), len(_STABLE_TAB_ORDER) - 1))
    return _STABLE_TAB_ORDER[idx]


def _stable_layout_kwargs(width):
    return widgets.Layout(width=width, min_width="0", flex=f"0 0 {width}")


def _stable_html_cell(content, width, align="left", html_content=False, header=False):
    if content is None:
        content = ""
    if html_content:
        inner = content
    else:
        text_value = text_or_blank(content, strip=True) or ""
        inner = safe_html(text_value)
    align_class = {"left": "stable-left", "center": "stable-center", "right": "stable-right"}.get(align, "stable-left")
    base_class = "stable-header-cell" if header else "stable-cell"
    html = f'<div class="event-app {base_class} {align_class}">{inner}</div>'
    outer = widgets.HTML(html, layout=_stable_layout_kwargs(width))
    outer.add_class("stable-col")
    return outer


def _stable_datetime_html(row):
    event_date = row.get("event_date")
    if is_missing(event_date):
        date_text = ""
    else:
        date_text = pd.Timestamp(event_date).strftime("%Y-%m-%d")
    time_text = fmt_time_text(row.get("event_time"))
    return (
        '<div class="event-app stable-date-stack">'
        f'<div class="stable-date-main">{safe_html(date_text)}</div>'
        f'<div class="stable-date-sub">{safe_html(time_text)}</div>'
        "</div>"
    )


def _stable_country_text(self, row):
    country = _strip_country_word(row.get("country"))
    if country:
        return country
    ticker = row.get("ticker")
    if ticker:
        return str(ticker)
    return "-"


def _stable_period_text(row):
    period = row.get("period")
    if is_missing(period):
        return "-"
    text_value = text_or_blank(period, strip=True)
    return text_value or "-"


def _stable_metric_text(row, field_name):
    if field_name == "prior":
        return fmt_metric(row.get("prior"), row.get("scaling_factor"))
    if field_name == "survey":
        return fmt_metric(row.get("survey"), row.get("scaling_factor"))
    if field_name == "actual":
        return _row_actual_text(row)
    return "-"


def _stable_event_button_cell(self, row, width):
    title = text_or_blank(row.get("title"), strip=True) or "Untitled event"
    tooltip_bits = [title]
    country = _stable_country_text(self, row)
    if country and country != "-":
        tooltip_bits.append(country)
    period = _stable_period_text(row)
    if period and period != "-":
        tooltip_bits.append(period)
    btn = widgets.Button(
        description=truncate_display_text(title, 80),
        tooltip=" | ".join(tooltip_bits),
        layout=widgets.Layout(width="100%", min_width="0"),
    )
    btn.add_class("stable-event-button")
    if row.get("event_id") == self.current_event_id:
        btn.add_class("list-event-button-selected")
    btn.on_click(lambda _, event_id=row["event_id"]: _activate_event_selection(self, event_id))
    wrapper = widgets.Box([btn], layout=_stable_layout_kwargs(width))
    wrapper.add_class("stable-col")
    return wrapper


def _stable_header_row_widget():
    cells = []
    for label, width, _, align in _STABLE_COLUMNS:
        cells.append(_stable_html_cell(label, width, align=align, html_content=False, header=True))
    row = widgets.HBox(cells, layout=widgets.Layout(width="100%"))
    row.add_class("stable-header-row")
    return row


def _stable_day_header_widget(label):
    return widgets.HTML(
        f'<div class="event-app stable-day-divider">{safe_html(label)}</div>',
        layout=widgets.Layout(width="100%"),
    )


def _stable_row_widget(self, row, stripe_even=False):
    cells = [
        _stable_html_cell(_stable_datetime_html(row), _STABLE_COLUMNS[0][1], align="left", html_content=True),
        _stable_html_cell(_stable_country_text(self, row), _STABLE_COLUMNS[1][1], align="left", html_content=False),
        _stable_event_button_cell(self, row, _STABLE_COLUMNS[2][1]),
        _stable_html_cell(_stable_period_text(row), _STABLE_COLUMNS[3][1], align="left", html_content=False),
        _stable_html_cell(_stable_metric_text(row, "prior"), _STABLE_COLUMNS[4][1], align="right", html_content=False),
        _stable_html_cell(_stable_metric_text(row, "survey"), _STABLE_COLUMNS[5][1], align="right", html_content=False),
        _stable_html_cell(_stable_metric_text(row, "actual"), _STABLE_COLUMNS[6][1], align="right", html_content=False),
        _stable_html_cell(outcome_badge_html(row), _STABLE_COLUMNS[7][1], align="center", html_content=True),
    ]
    row_box = widgets.HBox(cells, layout=widgets.Layout(width="100%"))
    row_box.add_class("stable-data-row")
    if stripe_even:
        row_box.add_class("row-even")
    if bool(row.get("_important_flag")):
        row_box.add_class("important-row")
    if bool(row.get("_flagged_flag")):
        row_box.add_class("flagged-row")
    if row.get("event_id") == self.current_event_id:
        row_box.add_class("selected-row")
    return row_box


def _stable_update_pager(self, tab_name, total_rows, start_idx, end_idx, max_page):
    current_page = self.explorer_page_by_tab.get(tab_name, 0)
    self.explorer_prev_button.disabled = current_page <= 0
    self.explorer_next_button.disabled = current_page >= max_page
    if total_rows <= 0:
        label = "Pane 1 of 1 - 0 rows"
    else:
        label = f"Pane {current_page + 1} of {max_page + 1} \u00b7 Rows {start_idx + 1}-{end_idx} of {total_rows}"
    self.explorer_page_label.value = f'<div class="event-app fast-pane-label">{safe_html(label)}</div>'


def _stable_render_tab(self, tab_name):
    container = getattr(self, _STABLE_OUTPUT_ATTRS[tab_name], None)
    if container is None:
        return
    df = getattr(self, "_explorer_frames", {}).get(tab_name)
    if df is None or df.empty:
        _stable_update_pager(self, tab_name, 0, 0, 0, 0)
        container.children = (widgets.HTML('<div class="event-app stable-empty-state"><div class="empty-state">No events in this view.</div></div>'),)
        self._stable_rendered_tabs[tab_name] = True
        return

    total_rows = len(df)
    max_page = max(0, math.ceil(total_rows / self.explorer_page_size) - 1)
    current_page = min(max(0, self.explorer_page_by_tab.get(tab_name, 0)), max_page)
    self.explorer_page_by_tab[tab_name] = current_page

    start_idx = current_page * self.explorer_page_size
    end_idx = min(total_rows, start_idx + self.explorer_page_size)
    page_df = df.iloc[start_idx:end_idx].copy()

    _stable_update_pager(self, tab_name, total_rows, start_idx, end_idx, max_page)

    children = [_stable_header_row_widget()]
    stripe_even = False
    for day_value, day_group in page_df.groupby("_day_key", sort=True):
        day_label = pd.Timestamp(day_value).strftime("%a %Y-%m-%d")
        total = len(day_group)
        key_count = int(day_group["_important_flag"].sum())
        flag_count = int(day_group["_flagged_flag"].sum())
        header_bits = [f"{day_label} - {total} event{'s' if total > 1 else ''}"]
        if key_count:
            header_bits.append(f"KEY({key_count})")
        if flag_count:
            header_bits.append(f"FLAG({flag_count})")
        holiday_codes = (
            day_group.loc[day_group["category"] == "Holiday", "ticker"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )
        if holiday_codes:
            header_bits.append("Closed: " + ", ".join(sorted(holiday_codes)))
        children.append(_stable_day_header_widget(" \u00b7 ".join(header_bits)))

        for row in day_group.to_dict("records"):
            children.append(_stable_row_widget(self, row, stripe_even=stripe_even))
            stripe_even = not stripe_even

    container.children = tuple(children)
    self._stable_rendered_tabs[tab_name] = True


_previous_build_widgets_stable = BloombergEventCalendarApp._build_widgets


def _stable_build_widgets(self, months_per_page_default):
    _previous_build_widgets_stable(self, months_per_page_default)

    self.explorer_page_size = int(globals().get("EXPLORER_PAGE_SIZE", _STABLE_PAGE_SIZE))
    self.explorer_page_by_tab = {tab_name: 0 for tab_name in _STABLE_TAB_ORDER}
    self._stable_rendered_tabs = {tab_name: False for tab_name in _STABLE_TAB_ORDER}

    self.all_list = _stable_make_tab_container()
    self.econ_list = _stable_make_tab_container()
    self.earnings_list = _stable_make_tab_container()
    self.comdty_list = _stable_make_tab_container()
    self.holiday_list = _stable_make_tab_container()
    self.watch_list = _stable_make_tab_container()

    self.tabs.children = [
        self.all_list,
        self.econ_list,
        self.earnings_list,
        self.comdty_list,
        self.holiday_list,
        self.watch_list,
    ]

    for i, title in enumerate(_STABLE_TAB_ORDER):
        self.tabs.set_title(i, title)
    self.tabs.add_class("stable-tabs-widget")

    try:
        list_detail_box = self.root.children[-1]
        list_card, detail_card = list_detail_box.children
        list_card.layout.width = "58%"
        detail_card.layout.width = "42%"

        caption = widgets.HTML(
            '<div class="event-app stable-explorer-caption">'
            "Fast tabular explorer. Only the active tab and current pane are rendered. "
            "Click an event name to sync the detail pane, notes, selected-day slate, and month highlights."
            "</div>"
        )

        pager_bar = widgets.HBox(
            [self.explorer_prev_button, self.explorer_next_button, self.explorer_page_label],
            layout=widgets.Layout(width="100%", align_items="center", justify_content="flex-start"),
        )
        pager_bar.add_class("fast-explorer-pager")
        list_card.children = (caption, pager_bar, self.tabs)
    except Exception:
        pass

    self.tabs.selected_index = 0
    self.day_out.layout.height = "980px"
    self.day_out.layout.overflow = "auto"
    self.day_panel.layout.min_height = "1040px"


def _stable_refresh_lists(self):
    _fast_prepare_explorer_frames(self)
    self._stable_rendered_tabs = {tab_name: False for tab_name in _STABLE_TAB_ORDER}
    self.render_watchlist_box()

    filtered_ids = set(self.filtered_events["event_id"]) if self.filtered_events is not None and not self.filtered_events.empty else set()
    if self.current_event_id and self.current_event_id not in filtered_ids:
        self.current_event_id = None

    if self.current_event_id is None and self.filtered_events is not None and not self.filtered_events.empty:
        all_df = getattr(self, "_explorer_frames", {}).get("ALL", pd.DataFrame())
        if all_df is not None and not all_df.empty:
            first_row = all_df.iloc[0]
            self.current_event_id = first_row["event_id"]
            event_date = first_row.get("event_date")
            if not is_missing(event_date):
                self.selected_day = pd.Timestamp(event_date).date()

    if self.current_event_id:
        self.render_detail(self.current_event_id)
    else:
        self.render_detail(None)

    _stable_render_tab(self, _stable_active_tab_name(self))
    self.render_selected_day_detail()


def _stable_get_active_list_widget(self):
    return getattr(self, _STABLE_OUTPUT_ATTRS[_stable_active_tab_name(self)])


def _stable_on_tab_change(self, change):
    tab_name = _stable_active_tab_name(self)
    active_df = getattr(self, "_explorer_frames", {}).get(tab_name, pd.DataFrame())
    if active_df is not None and not active_df.empty:
        active_ids = set(active_df["event_id"])
        if self.current_event_id not in active_ids:
            first_row = active_df.iloc[0]
            self.current_event_id = first_row["event_id"]
            event_date = first_row.get("event_date")
            if not is_missing(event_date):
                self.selected_day = pd.Timestamp(event_date).date()
            self.render_detail(self.current_event_id)
            self.render_selected_day_detail()

    self.render_monthly_page()
    _stable_render_tab(self, tab_name)


def _stable_prev_explorer_page(self, _=None):
    tab_name = _stable_active_tab_name(self)
    current_page = self.explorer_page_by_tab.get(tab_name, 0)
    if current_page <= 0:
        return
    self.explorer_page_by_tab[tab_name] = current_page - 1
    _stable_render_tab(self, tab_name)


def _stable_next_explorer_page(self, _=None):
    tab_name = _stable_active_tab_name(self)
    df = getattr(self, "_explorer_frames", {}).get(tab_name)
    total_rows = 0 if df is None else len(df)
    max_page = max(0, math.ceil(max(total_rows, 1) / self.explorer_page_size) - 1)
    current_page = self.explorer_page_by_tab.get(tab_name, 0)
    if current_page >= max_page:
        return
    self.explorer_page_by_tab[tab_name] = current_page + 1
    _stable_render_tab(self, tab_name)


BloombergEventCalendarApp._build_widgets = _stable_build_widgets
BloombergEventCalendarApp.refresh_lists = _stable_refresh_lists
BloombergEventCalendarApp.get_active_list_widget = _stable_get_active_list_widget
BloombergEventCalendarApp.on_tab_change = _stable_on_tab_change
BloombergEventCalendarApp.previous_explorer_pane = _stable_prev_explorer_page
BloombergEventCalendarApp.next_explorer_pane = _stable_next_explorer_page


# =============================================================================
# Text explorer patch: revert to a fast, stable, select-based explorer with
# clearer fixed-width rows and explicit dark-mode styling.
# =============================================================================

TEXT_EXPLORER_EXTRA_CSS = """
<style>
.event-app-shell .event-list-select {
    width: 100% !important;
}

.event-app-shell .event-list-select select {
    width: 100% !important;
    min-width: 100% !important;
    min-height: 720px !important;
    height: 720px !important;
    background: #0f1729 !important;
    color: #e5eefb !important;
    border: 1px solid #24324a !important;
    border-radius: 12px !important;
    box-shadow: inset 0 0 1px rgba(93, 168, 255, 0.04) !important;
    font-family: "SFMono-Regular", Menlo, Monaco, Consolas, "Liberation Mono", monospace !important;
    font-size: 10.4px !important;
    line-height: 1.35 !important;
    white-space: pre !important;
    letter-spacing: 0 !important;
    padding: 4px !important;
    overflow-x: hidden !important;
}

.event-app-shell .event-list-select select:focus {
    outline: none !important;
    border-color: #5da8ff !important;
    box-shadow: 0 0 0 1px rgba(93, 168, 255, 0.30) !important;
}

.event-app-shell .event-list-select option {
    background: #0f1729 !important;
    color: #e5eefb !important;
    font-family: inherit !important;
    font-size: inherit !important;
    line-height: 1.35 !important;
}

.event-app-shell .event-list-select option:checked,
.event-app-shell .event-list-select option[selected] {
    background: #16233a linear-gradient(0deg, #16233a, #16233a) !important;
    color: #fff !important;
}

.event-app-shell .text-explorer-caption {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.5;
    margin-bottom: 8px;
}

.event-app-shell .text-explorer-caption b {
    color: var(--text);
}
</style>
"""


def _text_compact_datetime(row):
    event_date = row.get("event_date")
    try:
        date_text = pd.Timestamp(event_date).strftime("%m/%d")
    except Exception:
        date_text = ""
    time_text = fmt_time_text(row.get("event_time"))
    return f"{date_text} {time_text}"


def _text_period_value(row):
    category = text_or_blank(row.get("category"), strip=True)
    if category == "Economic":
        return text_or_blank(row.get("period"), strip=True) or ""
    if category == "Earnings":
        return text_or_blank(row.get("ticker"), strip=True) or ""
    if category == "Commodity":
        delivery = row.get("delivery_date")
        if is_missing(delivery):
            delivery = row.get("EXPIRE_DT0_VALUE")
        if not is_missing(delivery):
            try:
                return pd.Timestamp(delivery).strftime("%m/%d")
            except Exception:
                return text_or_blank(delivery, strip=True)
        return text_or_blank(coalesce(row.get("underlying_ticker"), row.get("ticker")), strip=True) or ""
    if category == "Holiday":
        return "Closed"
    return text_or_blank(row.get("period"), strip=True) or ""


def _text_event_title(self, row, view_name="ALL"):
    title = text_or_blank(row.get("title"), strip=True) or "Untitled"
    category = text_or_blank(row.get("category"), strip=True) or "Event"
    view_name = str(view_name or "ALL").upper()
    if view_name in ("ALL", "IMPORTANT") and category != "Economic":
        prefix = self._list_category_code(category)
        return f"[{prefix}] {title}"
    return title


def _text_metric_or_blank(row, field_name):
    value = row.get(field_name)
    if is_missing(value):
        return ""
    return fmt_metric(value, row.get("scaling_factor"))


def _text_outcome_value(row):
    try:
        outcome = release_outcome(row)
        return text_or_blank(outcome.get("label"), strip=True) or "N/A"
    except Exception:
        return "N/A"


def _text_tag_value(self, row, view_name="ALL"):
    event_id = row["event_id"]
    important = self.is_event_important(event_id, row=row)
    flagged = self._row_flagged(event_id)
    if important and flagged:
        return "BOTH"
    if flagged:
        return "FLG"
    if important:
        return "KEY"
    view_name = str(view_name or "ALL").upper()
    if view_name in ("ALL", "IMPORTANT"):
        return self._list_category_code(row.get("category"))
    return ""


def _text_list_separator_label(self, event_date):
    ts = pd.Timestamp(event_date)
    return f"{'=' * 6} {ts.strftime('%Y-%m-%d (%a)')} {'=' * 26}"


def _text_list_header_label(self, view_name):
    cells = [
        display_cell("DT/TM", 11, default=""),
        display_cell("CTRY", 4, default=""),
        display_cell("EVENT", 21, default=""),
        display_cell("PER", 6, default=""),
        display_cell("PRIOR", 7, align="right", default=""),
        display_cell("SURV", 7, align="right", default=""),
        display_cell("ACT", 7, align="right", default=""),
        display_cell("B/M", 5, default=""),
        display_cell("TAG", 4, default=""),
    ]
    return join_display_cells(cells)


def _text_row_list_label(self, row, view_name="ALL"):
    category = text_or_blank(row.get("category"), strip=True)
    prior = _text_metric_or_blank(row, "prior")
    survey = _text_metric_or_blank(row, "survey")
    actual = _text_metric_or_blank(row, "actual")
    outcome = _text_outcome_value(row)
    if category not in ("Economic", "Earnings"):
        prior = ""
        survey = ""
        actual = ""
        outcome = ""
    country = self._row_country_code(row) or ""
    cells = [
        display_cell(_text_compact_datetime(row), 11, default=""),
        display_cell(country, 4, default=""),
        display_cell(_text_event_title(self, row, view_name=view_name), 21, default=""),
        display_cell(_text_period_value(row), 6, default=""),
        display_cell(prior, 7, align="right", default=""),
        display_cell(survey, 7, align="right", default=""),
        display_cell(actual, 7, align="right", default=""),
        display_cell(outcome, 5, default=""),
        display_cell(_text_tag_value(self, row, view_name=view_name), 4, default=""),
    ]
    return join_display_cells(cells)


def _text_build_widgets(self, months_per_page_default):
    _previous_build_widgets_fast(self, months_per_page_default)

    selectors = [
        self.all_list,
        self.econ_list,
        self.earnings_list,
        self.comdty_list,
        self.holiday_list,
        self.watch_list,
    ]
    for selector in selectors:
        selector.layout = widgets.Layout(width="100%", height="720px")
        try:
            selector.rows = 28
        except Exception:
            pass
        selector.add_class("event-list-select")

    self.tabs.layout = widgets.Layout(width="100%")
    self.tabs.selected_index = 0
    self.tabs.add_class("text-explorer-tabs")

    self.day_out.layout.height = "1000px"
    self.day_out.layout.overflow = "auto"
    self.day_panel.layout.min_height = "1060px"

    try:
        list_detail_box = self.root.children[-1]
        list_card, detail_card = list_detail_box.children
        list_card.layout.width = "56%"
        detail_card.layout.width = "44%"
        caption = widgets.HTML(
            '<div class="event-app text-explorer-caption">'
            "<b>Explorer</b> - stable text-based tabs with fixed-width rows for quick scanning. "
            "Columns read <b>DT/TM | CTRY | EVENT | PER | PRIOR | SURV | ACT | B/M | TAG</b>. "
            "Tags: <b>KEY</b>=important, <b>FLG</b>=flagged, <b>BOTH</b>=important and flagged. "
            '"Non-economic rows keep the same layout and leave release-metric columns blank where not applicable."'
            "</div>"
        )

        list_card.children = (caption, self.tabs)
    except Exception:
        pass

    self.root.children = (widgets.HTML(TEXT_EXPLORER_EXTRA_CSS),) + tuple(self.root.children)


def _text_wire_events(self):
    _previous_wire_events_fast(self)


def _text_apply_filters(self, _=None):
    self.month_page_start = 0
    return _previous_apply_filters_fast(self, _)


def _text_refresh_lists(self):
    preferred = self.current_event_id

    econ_df = self.filtered_events[self.filtered_events["category"] == "Economic"]
    earnings_df = self.filtered_events[self.filtered_events["category"] == "Earnings"]
    comdty_df = self.filtered_events[self.filtered_events["category"] == "Commodity"]
    holiday_df = self.filtered_events[self.filtered_events["category"] == "Holiday"]
    watch_df = self.filtered_events[self.filtered_events["event_id"].map(lambda x: self.is_event_important(x))]

    self._set_list_options(self.all_list, self.filtered_events, preferred, view_name="ALL")
    self._set_list_options(self.econ_list, econ_df, preferred, view_name="ECONOMIC")
    self._set_list_options(self.earnings_list, earnings_df, preferred, view_name="EARNINGS")
    self._set_list_options(self.comdty_list, comdty_df, preferred, view_name="COMMODITIES")
    self._set_list_options(self.holiday_list, holiday_df, preferred, view_name="HOLIDAYS")
    self._set_list_options(self.watch_list, watch_df, preferred, view_name="IMPORTANT")

    self.render_watchlist_box()

    active_widget = self.get_active_list_widget()
    active_values = [
        value for _, value in active_widget.options
        if value is not None and not self._is_separator_value(value)
    ]

    if preferred in active_values:
        active_widget.value = preferred
    elif active_values:
        active_widget.value = active_values[0]
    else:
        if self.current_event_id and not self.filtered_events.empty:
            self.render_detail(self.current_event_id)
        else:
            self.render_detail(None)


def _text_get_active_list_widget(self):
    idx = self.tabs.selected_index or 0
    idx = max(0, min(idx, 5))
    return [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list][idx]


def _text_on_tab_change(self, change):
    widget = self.get_active_list_widget()
    values = [
        value for _, value in widget.options
        if value is not None and not self._is_separator_value(value)
    ]
    if self.current_event_id in values:
        widget.value = self.current_event_id
    elif values:
        widget.value = values[0]
    else:
        if self.current_event_id:
            self.render_detail(self.current_event_id)
        else:
            self.render_detail(None)


BloombergEventCalendarApp._build_widgets = _text_build_widgets
BloombergEventCalendarApp._wire_events = _text_wire_events
BloombergEventCalendarApp.apply_filters = _text_apply_filters
BloombergEventCalendarApp.refresh_lists = _text_refresh_lists
BloombergEventCalendarApp.get_active_list_widget = _text_get_active_list_widget
BloombergEventCalendarApp.on_tab_change = _text_on_tab_change
BloombergEventCalendarApp._list_separator_label = _text_list_separator_label
BloombergEventCalendarApp._list_header_label = _text_list_header_label
BloombergEventCalendarApp._row_list_label = _text_row_list_label


# =============================================================================
# Hardcoded options and futures expiry schedule supplied by the user.
# The calendar now uses this embedded schedule instead of the old live commodity-expiry pull.
# =============================================================================

CUSTOM_HARDCODED_COMMODITY_EXPIRES_TSV = """\
Option Contract\tUnderlying Contract\tOption Expiration\tUnderlying Maturity
May-26\tCLK6\t4/16/2025\t4/21/2026
Jun-26\tCLM6\t5/14/2026\t5/19/2026
Jul-26\tCLN6\t6/16/2026\t6/22/2026
Aug-26\tCLQ6\t7/16/2026\t7/21/2026
Sep-26\tCLU6\t8/17/2026\t8/20/2026
Oct-26\tCLV6\t9/17/2026\t9/22/2026
Nov-26\tCLX6\t10/15/2026\t10/20/2026
Dec-26\tCLZ6\t11/17/2026\t11/20/2026
Jan-27\tCLF7\t12/16/2026\t12/21/2026
Feb-27\tCLG7\t1/14/2027\t1/20/2027
# ... (full schedule truncated)
"""


def _parse_hardcoded_commodity_schedule(tsv_text=None):
    tsv_text = CUSTOM_HARDCODED_COMMODITY_EXPIRES_TSV if tsv_text is None else tsv_text
    if tsv_text is None or not str(tsv_text).strip():
        return pd.DataFrame(columns=["Option Contract", "Underlying Contract", "Option Expiration", "Underlying Maturity"])
    try:
        df = pd.read_csv(StringIO(str(tsv_text).strip()), sep="\t")
    except Exception:
        return pd.DataFrame(columns=["Option Contract", "Underlying Contract", "Option Expiration", "Underlying Maturity"])
    expected = ["Option Contract", "Underlying Contract", "Option Expiration", "Underlying Maturity"]
    for col in expected:
        if col not in df.columns:
            df[col] = None
    df = df[expected].copy()
    for col in ["Option Contract", "Underlying Contract"]:
        df[col] = df[col].fillna("").astype(str).str.strip()
    for col in ["Option Expiration", "Underlying Maturity"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    df = df[
        df["Option Contract"].ne("")
        & df["Underlying Contract"].ne("")
        & df["Option Expiration"].notna()
        & df["Underlying Maturity"].notna()
    ].drop_duplicates().reset_index(drop=True)
    return df


def _build_hardcoded_commodity_events(start_ts=None, end_ts=None):
    raw_df = _parse_hardcoded_commodity_schedule()
    columns = BloombergEventCalendarApp.master_columns
    if raw_df.empty:
        return pd.DataFrame(columns=columns)

    rows = []
    for rec in raw_df.to_dict("records"):
        option_label = text_or_blank(rec.get("Option Contract"), strip=True)
        underlying_contract = text_or_blank(rec.get("Underlying Contract"), strip=True)
        option_exp = pd.to_datetime(rec.get("Option Expiration"), errors="coerce")
        underlying_mat = pd.to_datetime(rec.get("Underlying Maturity"), errors="coerce")
        if not option_label or not underlying_contract:
            continue

        if pd.notna(option_exp):
            rows.append({
                "event_id": make_event_id("CMDTYOPT", underlying_contract, option_label, option_exp.date()),
                "event_date": option_exp,
                "event_time": None,
                "category": "Commodity",
                "title": f"{underlying_contract} option expiry",
                "subtitle": f"Option ({option_label}) | Underlying maturity {underlying_mat.date() if pd.notna(underlying_mat) else '-'}",
                "country": None,
                "ticker": underlying_contract,
                "company": None,
                "period": option_label,
                "survey": None,
                "actual": None,
                "prior": None,
                "revision": None,
                "scaling_factor": None,
                "market_cap_usd": None,
                "future_price": None,
                "currency": None,
                "underlying_ticker": underlying_contract,
                "delivery_date": underlying_mat,
                "expire_group_id": option_label,
                "orig_ids": "Option expiry",
            })

        if pd.notna(underlying_mat):
            rows.append({
                "event_id": make_event_id("CMDTFUT", underlying_contract, option_label, underlying_mat.date()),
                "event_date": underlying_mat,
                "event_time": None,
                "category": "Commodity",
                "title": f"{underlying_contract} future maturity",
                "subtitle": f"Related option ({option_label}) | Option expiry {option_exp.date() if pd.notna(option_exp) else '-'}",
                "country": None,
                "ticker": underlying_contract,
                "company": None,
                "period": option_label,
                "survey": None,
                "actual": None,
                "prior": None,
                "revision": None,
                "scaling_factor": None,
                "market_cap_usd": None,
                "future_price": None,
                "currency": None,
                "underlying_ticker": underlying_contract,
                "delivery_date": underlying_mat,
                "expire_group_id": option_label,
                "orig_ids": "Future maturity",
            })

    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(columns=columns)

    for col in columns:
        if col not in out.columns:
            out[col] = None

    out["event_date"] = pd.to_datetime(out["event_date"], errors="coerce")
    out["delivery_date"] = pd.to_datetime(out["delivery_date"], errors="coerce")
    out = out.drop_duplicates(subset=["event_id"]).copy()

    if start_ts is not None and end_ts is not None:
        start_ts = pd.Timestamp(start_ts).normalize()
        end_ts = pd.Timestamp(end_ts).normalize()
        out = out[out["event_date"].between(start_ts, end_ts)]

    out["sort_time"] = out["event_time"].fillna("99:99")
    out = out.sort_values("event_date").drop(columns=["sort_time"])
    return out[columns].reset_index(drop=True)


def _hardcoded_prepare_commodity_events(self, df):
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame(columns=self.master_columns)
    if isinstance(df, pd.DataFrame) and "event_id" in df.columns and "event_date" in df.columns:
        out = df.copy()
        for col in self.master_columns:
            if col not in out.columns:
                out[col] = None
        out["event_date"] = pd.to_datetime(out["event_date"], errors="coerce")
        out["delivery_date"] = pd.to_datetime(out["delivery_date"], errors="coerce")
        out = out[out["event_date"].notna()].copy()
        return out[self.master_columns].reset_index(drop=True)
    return _original_prepare_commodity_events(self, df)


def _hardcoded_fetch_commodity_calendar(self, start_ts, end_ts):
    return _build_hardcoded_commodity_events(start_ts, end_ts)


def _hardcoded_load_from_dataframes(self, eco_df=None, earnings_df=None, comdty_df=None):
    eco_df = eco_df.copy() if eco_df is not None else pd.DataFrame()
    earnings_df = earnings_df.copy() if earnings_df is not None else pd.DataFrame()

    eco_events = self.prepare_economic_events(eco_df)
    earnings_events = self.prepare_earnings_events(earnings_df)

    date_starts = []
    date_ends = []
    for frame in [eco_events, earnings_events]:
        if frame is not None and not frame.empty:
            date_starts.append(pd.to_datetime(frame["event_date"], errors="coerce").min())
            date_ends.append(pd.to_datetime(frame["event_date"], errors="coerce").max())

    if date_starts and date_ends:
        inferred_start = min(pd.Timestamp(x).normalize() for x in date_starts if pd.notna(x))
        inferred_end = max(pd.Timestamp(x).normalize() for x in date_ends if pd.notna(x))
    else:
        inferred_start, inferred_end = self._validate_dates()

    commodity_events = _build_hardcoded_commodity_events(inferred_start, inferred_end)

    frames = [frame for frame in [eco_events, earnings_events, commodity_events] if frame is not None and not frame.empty]
    if frames:
        self.core_events = pd.concat(frames, ignore_index=True)
        self.core_events["event_date"] = pd.to_datetime(self.core_events["event_date"], errors="coerce")
        self.core_events = self.core_events[self.core_events["event_date"].notna()].copy()
        self.core_events["sort_time"] = self.core_events["event_time"].fillna("99:99")
        self.core_events["category_order"] = self.core_events["category"].map(CATEGORY_ORDER).fillna(99)
        self.core_events = self.core_events.sort_values(
            ["event_date", "category_order", "sort_time", "title"]
        ).drop(columns=["sort_time", "category_order"]).reset_index(drop=True)
        self.start_picker.value = inferred_start.date()
        self.end_picker.value = inferred_end.date()
        self.selected_day = inferred_start.date()
    else:
        self.core_events = pd.DataFrame(columns=self.master_columns)
        self.selected_day = self.start_picker.value

    self.month_page_start = 0
    self.apply_filters()
    holiday_count = len(self.prepare_holiday_events(self._validate_dates()))
    hardcoded_rows = len(_parse_hardcoded_commodity_schedule())
    self.log(
        f"Loaded {len(self.core_events)} events + {holiday_count} market holiday markers. "
        f"Commodity events are using the embedded hardcoded options/futures expiry schedule ({hardcoded_rows} source rows)."
    )


def _hardcoded_event_fields(self, row):
    if (text_or_blank(row.get("category"), strip=True) == "Commodity"
            and text_or_blank(row.get("orig_ids"), strip=True) in ("Option expiry", "Future maturity")):
        return [
            ("Event Date", fmt_value(row.get("event_date"))),
            ("Contract", fmt_value(coalesce(row.get("underlying_ticker"), row.get("ticker")))),
            ("Type", fmt_value(row.get("orig_ids"), 0)),
            ("Option Label", fmt_value(row.get("expire_group_id"), 0)),
            ("Underlying Maturity", fmt_value(row.get("delivery_date"))),
            ("Detail", fmt_value(row.get("subtitle"), 0)),
        ]
    return _original_event_fields(self, row)


def _hardcoded_list_detail_text(self, row):
    if (text_or_blank(row.get("category"), strip=True) == "Commodity"
            and text_or_blank(row.get("orig_ids"), strip=True) in ("Option expiry", "Future maturity")):
        ticker = text_or_blank(coalesce(row.get("underlying_ticker"), row.get("ticker")), strip=True) or ""
        kind = text_or_blank(row.get("orig_ids"), strip=True) or ""
        option_label = text_or_blank(row.get("expire_group_id"), strip=True) or ""
        maturity = fmt_value(row.get("delivery_date"))
        return f"{ticker} | {kind} | {option_label} | Fut mat {maturity}"
    return _original_list_detail_text(self, row)


def _hardcoded_list_header_label(self, view_name):
    view_name = str(view_name or "ALL").upper()
    if view_name == "COMMODITIES":
        cells = [
            display_cell("MARK", 4, default=""),
            display_cell("TIME", 5, default=""),
            display_cell("EVENT", 24, default=""),
            display_cell("TICKER", 10, default=""),
            display_cell("MATURITY", 10, default=""),
            display_cell("TYPE", 12, default=""),
        ]
        return join_display_cells(cells)
    return _original_list_header_label(self, view_name)


def _hardcoded_row_list_label(self, row, view_name="ALL"):
    view_name = str(view_name or "ALL").upper()
    if (view_name == "COMMODITIES" and text_or_blank(row.get("category"), strip=True) == "Commodity"
            and text_or_blank(row.get("orig_ids"), strip=True) in ("Option expiry", "Future maturity")):
        marker = self._row_marker_text(row["event_id"], row=row)
        event_time = fmt_time_text(row.get("event_time"))
        title = text_or_blank(row.get("title"), strip=True) or "Untitled"
        ticker = text_or_blank(coalesce(row.get("underlying_ticker"), row.get("ticker")), strip=True) or ""
        delivery = fmt_value(row.get("delivery_date"))
        kind = text_or_blank(row.get("orig_ids"), strip=True)
        option_label = text_or_blank(row.get("expire_group_id"), strip=True)
        type_text = f"OPT" if kind == "Option expiry" else f"FUT ({option_label})".strip()
        cells = [
            display_cell(marker, 4, default=""),
            display_cell(event_time, 5),
            display_cell(title, 24),
            display_cell(ticker, 10),
            display_cell(delivery, 10),
            display_cell(type_text, 12),
        ]
        return join_display_cells(cells)
    return _original_row_list_label(self, row, view_name)


_original_prepare_commodity_events = BloombergEventCalendarApp.prepare_commodity_events
_original_load_from_dataframes = BloombergEventCalendarApp.load_from_dataframes
_original_event_fields = BloombergEventCalendarApp._event_fields
_original_list_detail_text = BloombergEventCalendarApp._list_detail_text
_original_list_header_label = BloombergEventCalendarApp._list_header_label
_original_row_list_label = BloombergEventCalendarApp._row_list_label

BloombergEventCalendarApp.prepare_commodity_events = _hardcoded_prepare_commodity_events
BloombergEventCalendarApp.fetch_commodity_calendar = _hardcoded_fetch_commodity_calendar
BloombergEventCalendarApp.load_from_dataframes = _hardcoded_load_from_dataframes
BloombergEventCalendarApp._event_fields = _hardcoded_event_fields
BloombergEventCalendarApp._list_detail_text = _hardcoded_list_detail_text
BloombergEventCalendarApp._list_header_label = _hardcoded_list_header_label
BloombergEventCalendarApp._row_list_label = _hardcoded_row_list_label


# =============================================================================
# Explorer polish + custom events patch
# - Slightly larger explorer font
# - TAG column moved to the far left
# - Stable color-coded left tag markers for KEY / FLG / BOTH
# - Custom event add / delete with persistent CSV storage
# =============================================================================

if "CUSTOM_EVENTS_FILE" not in globals():
    CUSTOM_EVENTS_FILE = "custom_events.csv"

CATEGORY_ORDER = dict(CATEGORY_ORDER)
CATEGORY_ORDER["Custom"] = 4
LIST_CATEGORY_CODES = dict(LIST_CATEGORY_CODES)
LIST_CATEGORY_CODES["Custom"] = "CUST"

APP_CSS = APP_CSS + """
<style>
.badge-custom {
    color: #7dd3fc;
    background: rgba(125, 211, 252, 0.12);
    border-color: rgba(125, 211, 252, 0.28);
}

.custom-event-form .widget-label {
    min-width: 88px !important;
}

.custom-event-note {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.45;
}
</style>
"""

EXPLORER_UI_PATCH_CSS = """
<style>
.event-app-shell .event-list-select select {
    min-height: 790px !important;
    height: 790px !important;
    font-size: 11.8px !important;
    line-height: 1.42 !important;
    padding: 6px 5px !important;
}

.event-app-shell .event-list-select option {
    font-size: 11.8px !important;
    line-height: 1.42 !important;
    padding-top: 1px !important;
    padding-bottom: 1px !important;
}

.event-app-shell .explorer-enhanced-caption {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.45;
    margin-bottom: 8px;
}

.event-app-shell .explorer-enhanced-caption b {
    color: var(--text);
}
</style>
"""


def _custom_events_empty_frame(self):
    return pd.DataFrame(columns=self.master_columns)


def _normalize_custom_event_time(value):
    raw = text_or_blank(value, strip=True)
    if not raw:
        return None
    raw = raw.upper().replace(" ", "")
    if re.fullmatch(r"\d{1,2}:\d{2}", raw):
        hour_text, minute_text = raw.split(":")
        hour = int(hour_text)
        minute = int(minute_text)
    elif re.fullmatch(r"\d{3,4}", raw):
        if len(raw) == 3:
            hour = int(raw[0])
            minute = int(raw[1:])
        else:
            hour = int(raw[:2])
            minute = int(raw[2:])
    else:
        parsed = pd.to_datetime(raw, errors="coerce")
        if pd.isna(parsed):
            raise ValueError("Custom event time must be blank or in HH:MM format.")
        hour = int(parsed.hour)
        minute = int(parsed.minute)
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError("Custom event time must be blank or in HH:MM format.")
    return f"{hour:02d}:{minute:02d}"


def _custom_event_id_from_values(event_date, event_time, title, country="", ticker="", period=""):
    if is_missing(event_date):
        date_text = ""
    else:
        date_text = pd.Timestamp(event_date).strftime("%Y-%m-%d")
    return make_event_id(
        "CUST",
        date_text,
        text_or_blank(event_time, strip=True),
        normalize_key_text(country),
        normalize_key_text(ticker),
        normalize_key_text(period),
        normalize_key_text(title),
    )


def _custom_event_row_from_form(self):
    if getattr(self, "custom_event_date", None) is None or self.custom_event_date.value is None:
        raise ValueError("Choose a custom-event date first.")
    title = text_or_blank(getattr(self, "custom_event_title", None).value if getattr(self, "custom_event_title", None) is not None else "", strip=True)
    if not title:
        raise ValueError("Custom events need a title.")
    time_value = _normalize_custom_event_time(
        getattr(self, "custom_event_time", None).value if getattr(self, "custom_event_time", None) is not None else ""
    )
    event_date = pd.Timestamp(self.custom_event_date.value).normalize()
    country = _strip_country_word(text_or_blank(
        getattr(self, "custom_event_country", None).value if getattr(self, "custom_event_country", None) is not None else "",
        strip=True,
    ))
    ticker = text_or_blank(
        getattr(self, "custom_event_ticker", None).value if getattr(self, "custom_event_ticker", None) is not None else "",
        strip=True,
    )
    period = text_or_blank(
        getattr(self, "custom_event_period", None).value if getattr(self, "custom_event_period", None) is not None else "",
        strip=True,
    )
    subtitle = text_or_blank(
        getattr(self, "custom_event_subtitle", None).value if getattr(self, "custom_event_subtitle", None) is not None else "",
        strip=True,
    )

    event_id = _custom_event_id_from_values(event_date, time_value, title, country=country, ticker=ticker, period=period)

    row = {col: None for col in self.master_columns}
    row.update({
        "event_id": event_id,
        "event_date": event_date,
        "event_time": time_value,
        "category": "Custom",
        "title": title,
        "subtitle": subtitle or "User-defined custom event",
        "country": country or None,
        "ticker": ticker or None,
        "company": None,
        "period": period or None,
        "survey": None,
        "actual": None,
        "prior": None,
        "revision": None,
        "scaling_factor": None,
        "market_cap_usd": None,
        "future_price": None,
        "currency": None,
        "underlying_ticker": None,
        "delivery_date": None,
        "expire_group_id": None,
        "orig_ids": "Custom event",
    })
    return row


def _load_custom_events_from_disk(self, silent=True):
    path = getattr(self, "custom_events_path", Path(CUSTOM_EVENTS_FILE))
    self.custom_events_path = path
    self.custom_events = self._custom_events_empty_frame()
    if not path.exists():
        return False

    def _log(msg):
        if not silent:
            self.log(msg)

    try:
        df = pd.read_csv(path)
    except Exception as exc:
        _log(f"Could not load custom events from {path.resolve()}: {exc}")
        return False

    if df is None or df.empty:
        self.custom_events = self._custom_events_empty_frame()
        _log(f"Loaded 0 custom events from {path.resolve()}")
        return True

    for col in self.master_columns:
        if col not in df.columns:
            df[col] = None

    df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce")
    if "delivery_date" in df.columns:
        df["delivery_date"] = pd.to_datetime(df["delivery_date"], errors="coerce")
    else:
        df["delivery_date"] = None

    df["category"] = "Custom"

    if "event_time" in df.columns:
        normalized_times = []
        for raw in df["event_time"].fillna("").tolist():
            try:
                normalized_times.append(_normalize_custom_event_time(raw))
            except Exception:
                normalized_times.append(text_or_blank(raw, strip=True) or None)
        df["event_time"] = normalized_times
    else:
        df["event_time"] = None

    if "event_id" not in df.columns:
        df["event_id"] = ""

    recomputed_ids = []
    for row in df.fillna("").to_dict("records"):
        existing = text_or_blank(row.get("event_id"), strip=True)
        if existing:
            recomputed_ids.append(existing)
        else:
            recomputed_ids.append(
                _custom_event_id_from_values(
                    row.get("event_date"),
                    row.get("event_time"),
                    row.get("title"),
                    country=row.get("country"),
                    ticker=row.get("ticker"),
                    period=row.get("period"),
                )
            )

    df["event_id"] = recomputed_ids

    df = df[df["event_date"].notna()].copy()
    if df.empty:
        self.custom_events = self._custom_events_empty_frame()
        _log(f"Loaded 0 custom events from {path.resolve()}")
        return True

    for col in self.master_columns:
        if col not in df.columns:
            df[col] = None

    df = (
        df[self.master_columns]
        .drop_duplicates(subset=["event_id"], keep="last")
        .sort_values(["event_date", "event_time", "title"], na_position="last")
        .reset_index(drop=True)
    )

    self.custom_events = df
    _log(f"Loaded {len(df)} custom events from {path.resolve()}")
    return True


def _save_custom_events_to_disk(self):
    path = getattr(self, "custom_events_path", Path(CUSTOM_EVENTS_FILE))
    self.custom_events_path = path
    path.parent.mkdir(parents=True, exist_ok=True)

    df = getattr(self, "custom_events", self._custom_events_empty_frame()).copy()
    for col in self.master_columns:
        if col not in df.columns:
            df[col] = None
    if not df.empty:
        df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce").dt.strftime("%Y-%m-%d")
        if "delivery_date" in df.columns:
            df["delivery_date"] = pd.to_datetime(df["delivery_date"], errors="coerce").dt.strftime("%Y-%m-%d")
        df = df[self.master_columns].sort_values(["event_date", "event_time", "title"], na_position="last")
    else:
        df = self._custom_events_empty_frame()
    df.to_csv(path, index=False)


def _custom_events_in_range(self, start_ts, end_ts):
    df = getattr(self, "custom_events", self._custom_events_empty_frame()).copy()
    if df is None or df.empty:
        return self._custom_events_empty_frame()
    df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce")
    df = df[df["event_date"].notna()].copy()
    df = df[df["event_date"].between(pd.Timestamp(start_ts).normalize(), pd.Timestamp(end_ts).normalize())]
    if df.empty:
        return self._custom_events_empty_frame()
    return df[self.master_columns].reset_index(drop=True)


def _clear_custom_event_form(self, _=None):
    if getattr(self, "custom_event_date", None) is not None:
        self.custom_event_date.value = None
    if getattr(self, "custom_event_time", None) is not None:
        self.custom_event_time.value = ""
    if getattr(self, "custom_event_country", None) is not None:
        self.custom_event_country.value = ""
    if getattr(self, "custom_event_ticker", None) is not None:
        self.custom_event_ticker.value = ""
    if getattr(self, "custom_event_period", None) is not None:
        self.custom_event_period.value = ""
    if getattr(self, "custom_event_title", None) is not None:
        self.custom_event_title.value = ""
    if getattr(self, "custom_event_subtitle", None) is not None:
        self.custom_event_subtitle.value = ""


def _sync_custom_form_from_selected(self, event_id):
    if not event_id:
        return
    row = self._find_event_row(event_id)
    if row is None:
        return
    category = text_or_blank(row.get("category") if isinstance(row, dict) else row.get("category"), strip=True)
    if category != "Custom":
        return
    try:
        event_date = row.get("event_date") if hasattr(row, "get") else row["event_date"]
    except Exception:
        event_date = None
    if getattr(self, "custom_event_date", None) is not None:
        self.custom_event_date.value = pd.Timestamp(event_date).date() if not is_missing(event_date) else None
    if getattr(self, "custom_event_time", None) is not None:
        self.custom_event_time.value = text_or_blank(row.get("event_time") if hasattr(row, "get") else None, strip=True)
    if getattr(self, "custom_event_country", None) is not None:
        self.custom_event_country.value = text_or_blank(row.get("country") if hasattr(row, "get") else None, strip=True)
    if getattr(self, "custom_event_ticker", None) is not None:
        self.custom_event_ticker.value = text_or_blank(row.get("ticker") if hasattr(row, "get") else None, strip=True)
    if getattr(self, "custom_event_period", None) is not None:
        self.custom_event_period.value = text_or_blank(row.get("period") if hasattr(row, "get") else None, strip=True)
    if getattr(self, "custom_event_title", None) is not None:
        self.custom_event_title.value = text_or_blank(row.get("title") if hasattr(row, "get") else None, strip=True)
    if getattr(self, "custom_event_subtitle", None) is not None:
        self.custom_event_subtitle.value = text_or_blank(row.get("subtitle") if hasattr(row, "get") else None, strip=True)


def _add_custom_event(self, _=None):
    try:
        row = self._custom_event_row_from_form()
    except Exception as exc:
        self.log(str(exc))
        return

    if not hasattr(self, "custom_events"):
        self.custom_events = self._custom_events_empty_frame()

    event_date = pd.Timestamp(row["event_date"]).normalize()
    event_id = row["event_id"]

    if self.start_picker.value is None or event_date.date() < self.start_picker.value:
        self.start_picker.value = event_date.date()
    if self.end_picker.value is None or event_date.date() > self.end_picker.value:
        self.end_picker.value = event_date.date()

    custom_df = self.custom_events.copy()
    if not custom_df.empty:
        custom_df = custom_df[custom_df["event_id"] != event_id].copy()

    new_row_df = pd.DataFrame({col: [row.get(col) for col in self.master_columns]}, columns=self.master_columns)
    custom_df = pd.concat([custom_df, new_row_df], ignore_index=True)
    custom_df["event_date"] = pd.to_datetime(custom_df["event_date"], errors="coerce")
    custom_df = (
        custom_df[self.master_columns]
        .drop_duplicates(subset=["event_id"], keep="last")
        .sort_values(["event_date", "event_time", "title"], na_position="last")
        .reset_index(drop=True)
    )

    self.custom_events = custom_df
    self._save_custom_events_to_disk()

    if event_id in self.state:
        self._sync_state_metadata(event_id, row=pd.Series(row), touch_timestamp=False)
        self._save_state_to_disk()

    selected_categories = list(self.category_filter.value)
    if "Custom" not in selected_categories:
        selected_categories.append("Custom")
        try:
            self.category_filter.value = tuple(selected_categories)
        except Exception:
            pass

    self.selected_day = event_date.date()
    self.current_event_id = event_id
    self.month_page_start = 0
    self.apply_filters()

    try:
        self.tabs.selected_index = 0
        self.all_list.value = event_id
    except Exception:
        pass
    self.render_detail(event_id)

    self.log(f"Saved custom event to {self.custom_events_path.resolve()}")


def _delete_selected_custom_event(self, _=None):
    target_id = None
    target_title = ""

    row = self._find_event_row(self.current_event_id) if self.current_event_id else None
    if row is not None:
        category = text_or_blank(row.get("category") if hasattr(row, "get") else None, strip=True)
        if category == "Custom":
            target_id = row["event_id"] if not isinstance(row, dict) else row.get("event_id")
            target_title = text_or_blank(row.get("title") if hasattr(row, "get") else None, strip=True)

    if not target_id:
        try:
            form_row = self._custom_event_row_from_form()
            target_id = form_row["event_id"]
            target_title = text_or_blank(form_row.get("title"), strip=True)
        except Exception:
            self.log("Select a custom event, or fill the custom-event form with the event you want to delete.")
            return

    custom_df = getattr(self, "custom_events", self._custom_events_empty_frame()).copy()
    if custom_df.empty or target_id not in set(custom_df["event_id"].astype(str)):
        self.log("That custom event was not found in the custom-events file.")
        return

    self.custom_events = custom_df[custom_df["event_id"].astype(str) != str(target_id)].copy().reset_index(drop=True)
    self._save_custom_events_to_disk()

    if target_id in self.state:
        self.state.pop(target_id, None)
        self._save_state_to_disk()

    if self.current_event_id == target_id:
        self.current_event_id = None

    self.apply_filters()
    self.render_detail(None)
    self.render_monthly_page()
    self.render_selected_day_detail()
    self.clear_custom_event_form()
    title_text = target_title or target_id
    self.log(f"Deleted custom event '{title_text}' from {self.custom_events_path.resolve()}")


def _enhanced_tag_value(self, row):
    event_id = row["event_id"]
    important = self.is_event_important(event_id, row=row)
    flagged = self._row_flagged(event_id)
    if flagged and important:
        return "\u25a0 BTH"
    if flagged:
        return "\u25a0 FLG"
    if important:
        return "\u25a0 KEY"
    category = text_or_blank(row.get("category"), strip=True)
    return {
        "Economic": "ECON",
        "Earnings": "EARN",
        "Commodity": "CMDY",
        "Holiday": "HOL",
        "Custom": "CUST",
    }.get(category, "")


def _enhanced_period_value(row):
    category = text_or_blank(row.get("category"), strip=True)
    if category == "Economic":
        return text_or_blank(row.get("period"), strip=True) or ""
    if category == "Earnings":
        return text_or_blank(row.get("ticker"), strip=True) or ""
    if category == "Commodity":
        if text_or_blank(row.get("orig_ids"), strip=True) in ("Option expiry", "Future maturity"):
            return text_or_blank(row.get("expire_group_id"), strip=True) or fmt_value(row.get("delivery_date"))
        return fmt_value(row.get("delivery_date")) if not is_missing(row.get("delivery_date")) else ""
    if category == "Holiday":
        return "Closed"
    if category == "Custom":
        return text_or_blank(row.get("period"), strip=True) or "Custom"
    return text_or_blank(row.get("period"), strip=True) or ""


def _enhanced_row_title(self, row):
    return text_or_blank(row.get("title"), strip=True) or "Untitled"


def _enhanced_list_header_label(self, view_name):
    cells = [
        display_cell("TAG", 6, default=""),
        display_cell("DT/TM", 11, default=""),
        display_cell("CTRY", 4, default=""),
        display_cell("EVENT", 20, default=""),
        display_cell("PER", 6, default=""),
        display_cell("PRIOR", 7, align="right", default=""),
        display_cell("SURV", 7, align="right", default=""),
        display_cell("ACT", 7, align="right", default=""),
        display_cell("B/M", 5, default=""),
    ]
    return join_display_cells(cells)


def _enhanced_row_list_label(self, row, view_name="ALL"):
    category = text_or_blank(row.get("category"), strip=True)
    prior = _text_metric_or_blank(row, "prior")
    survey = _text_metric_or_blank(row, "survey")
    actual = _text_metric_or_blank(row, "actual")
    outcome = _text_outcome_value(row)
    if category not in ("Economic", "Earnings"):
        prior = ""
        survey = ""
        actual = ""
        outcome = ""
    country = self._row_country_code(row) or ""
    cells = [
        display_cell(_enhanced_tag_value(self, row), 6, default=""),
        display_cell(_text_compact_datetime(row), 11, default=""),
        display_cell(country, 4, default=""),
        display_cell(_enhanced_row_title(self, row), 20, default=""),
        display_cell(_enhanced_period_value(row), 6, default=""),
        display_cell(prior, 7, align="right", default=""),
        display_cell(survey, 7, align="right", default=""),
        display_cell(actual, 7, align="right", default=""),
        display_cell(outcome, 5, default=""),
    ]
    return join_display_cells(cells)


_previous_build_widgets_custom = BloombergEventCalendarApp._build_widgets
_previous_wire_events_custom = BloombergEventCalendarApp._wire_events
_previous_compose_custom = BloombergEventCalendarApp.compose_events_for_range
_previous_reset_filters_custom = BloombergEventCalendarApp.reset_filters
_previous_find_event_row_custom = BloombergEventCalendarApp._find_event_row
_previous_watch_source_events_custom = BloombergEventCalendarApp._watch_source_events
_previous_event_fields_custom = BloombergEventCalendarApp._event_fields
_previous_render_detail_custom = BloombergEventCalendarApp.render_detail
_previous_export_snapshot_custom = BloombergEventCalendarApp._export_snapshot_dataframe


def _enhanced_build_widgets(self, months_per_page_default):
    _previous_build_widgets_custom(self, months_per_page_default)

    self.custom_events_path = Path(globals().get("CUSTOM_EVENTS_FILE", CUSTOM_EVENTS_FILE))
    self.custom_events = self._custom_events_empty_frame()
    self._load_custom_events_from_disk(silent=True)

    try:
        category_options = ["Economic", "Earnings", "Commodity", "Holiday", "Custom"]
        self.category_filter.options = category_options
        self.category_filter.value = tuple(category_options)
    except Exception:
        pass

    common_desc = {"description_width": "initial"}
    self.custom_event_date = widgets.DatePicker(description="Date", style=common_desc, layout=widgets.Layout(width="180px"))
    self.custom_event_time = widgets.Text(description="Time", placeholder="HH:MM", style=common_desc, layout=widgets.Layout(width="145px"))
    self.custom_event_country = widgets.Text(description="Country", placeholder="US / UK / Japan", style=common_desc,
                                             layout=widgets.Layout(width="190px"))
    self.custom_event_ticker = widgets.Text(description="Ticker", placeholder="Optional", style=common_desc, layout=widgets.Layout(width="170px"))
    self.custom_event_period = widgets.Text(description="Period", placeholder="Optional", style=common_desc, layout=widgets.Layout(width="180px"))
    self.custom_event_title = widgets.Text(description="Title", placeholder="Required: event name", style=common_desc,
                                           layout=widgets.Layout(width="100%"))
    self.custom_event_subtitle = widgets.Textarea(description="Details", placeholder="Optional subtitle / description", style=common_desc,
                                                  layout=widgets.Layout(width="100%", height="30px"))

    self.add_custom_event_button = widgets.Button(description="Add / Update Custom Event", button_style="info", icon="plus")
    self.delete_custom_event_button = widgets.Button(description="Delete Selected Custom Event", button_style="danger", icon="trash")
    self.clear_custom_form_button = widgets.Button(description="Clear Custom Form", icon="eraser")

    self.custom_events_note = widgets.HTML(
        f'<div class="event-app custom-event-note">'
        f'Custom events persist to <b>{safe_html(str(self.custom_events_path))}</b>. '
        f'Notes, flags, and Important markers still live in <b>{safe_html(str(self.state_path))}</b>, so an identical custom event automatically reuses the same '
        f'annotation state.'
        f'</div>'
    )

    custom_title = widgets.HTML('<div class="event-app section-title">Custom Events</div>')
    custom_form = widgets.VBox([
        widgets.HBox([self.custom_event_date, self.custom_event_time, self.custom_event_country]),
        widgets.HBox([self.custom_event_ticker, self.custom_event_period]),
        self.custom_event_title,
        self.custom_event_subtitle,
        widgets.HBox([self.add_custom_event_button, self.delete_custom_event_button, self.clear_custom_form_button]),
        self.custom_events_note,
    ])
    custom_form.add_class("custom-event-form")

    try:
        list_detail_box = self.root.children[-1]
        list_card, detail_card = list_detail_box.children
        detail_children = list(detail_card.children)
        insert_at = min(6, len(detail_children))
        detail_children[insert_at:insert_at] = [custom_title, custom_form]
        detail_card.children = tuple(detail_children)

        list_card.children = (
            widgets.HTML(
                '<div class="event-app explorer-enhanced-caption">'
                "<b>Explorer</b> - larger fixed-width text tabs for faster scanning. "
                'Columns read <b>TAG | DT/TM | CTRY | EVENT | PER | PRIOR | SURV | ACT | B/M</b>. '
                'The left TAG cell is color-coded: <b>\u25a0 KEY</b>=important, <b>\u25a0 FLG</b>=flagged, <b>\u25a0 BTH</b>=both. '
                '"Non-highlighted rows show their category code in the TAG cell."'
                "</div>"
            ),
            self.tabs,
        )
    except Exception:
        pass

    for selector in [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list]:
        selector.layout = widgets.Layout(width="100%", height="720px")
        try:
            selector.rows = 32
        except Exception:
            pass

    self.day_out.layout.height = "1120px"
    self.day_out.layout.overflow = "auto"
    self.day_panel.layout.min_height = "1180px"

    self.root.children = (widgets.HTML(EXPLORER_UI_PATCH_CSS),) + tuple(self.root.children)


def _enhanced_wire_events(self):
    _previous_wire_events_custom(self)
    self.add_custom_event_button.on_click(self.add_custom_event)
    self.delete_custom_event_button.on_click(self.delete_selected_custom_event)
    self.clear_custom_form_button.on_click(self.clear_custom_event_form)


def _compose_events_for_range_with_custom(self, start_ts, end_ts):
    frames = []
    base_df = _previous_compose_custom(self, start_ts, end_ts)
    if base_df is not None and not base_df.empty:
        frames.append(base_df.copy())
    custom_df = self._custom_events_in_range(start_ts, end_ts)
    if custom_df is not None and not custom_df.empty:
        frames.append(custom_df.copy())

    if not frames:
        return pd.DataFrame(columns=self.master_columns)

    events = pd.concat(frames, ignore_index=True)
    for col in self.master_columns:
        if col not in events.columns:
            events[col] = None
    events["event_date"] = pd.to_datetime(events["event_date"], errors="coerce")
    events = events[events["event_date"].notna()].copy()
    events["sort_time"] = events["event_time"].fillna("99:99")
    events["category_order"] = events["category"].map(CATEGORY_ORDER).fillna(99)
    events = (
        events[self.master_columns + ["sort_time", "category_order"]]
        .drop_duplicates(subset=["event_id"], keep="first")
        .sort_values(["event_date", "category_order", "sort_time", "title"])
        .drop(columns=["sort_time", "category_order"])
        .reset_index(drop=True)
    )
    return events


def _reset_filters_with_custom(self, _=None):
    self.search_box.value = ""
    category_options = ["Economic", "Earnings", "Commodity", "Holiday", "Custom"]
    self.category_filter.options = category_options
    self.category_filter.value = tuple(category_options)
    self.only_flagged.value = False
    self.only_watch.value = False
    self.month_page_start = 0
    self.apply_filters()


def _find_event_row_with_custom(self, event_id):
    row = _previous_find_event_row_custom(self, event_id)
    if row is not None:
        return row
    custom_df = getattr(self, "custom_events", None)
    if custom_df is not None and not custom_df.empty and "event_id" in custom_df.columns:
        match = custom_df[custom_df["event_id"] == event_id]
        if not match.empty:
            return match.iloc[0]
    return None


def _watch_source_events_with_custom(self):
    frames = []
    base_df = _previous_watch_source_events_custom(self)
    if base_df is not None and not base_df.empty:
        frames.append(base_df.copy())
    custom_df = getattr(self, "custom_events", None)
    if custom_df is not None and not custom_df.empty:
        frames.append(custom_df.copy())
    if not frames:
        return pd.DataFrame(columns=self.master_columns)
    df = pd.concat(frames, ignore_index=True)
    df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce")
    df = df.dropna(subset=["event_date"]).drop_duplicates(subset=["event_id"], keep="first")
    return df.sort_values(["event_date", "title"]).reset_index(drop=True)


def _event_fields_with_custom(self, row):
    if text_or_blank(row.get("category"), strip=True) == "Custom":
        return [
            ("Date", fmt_value(row.get("event_date"))),
            ("Time", fmt_time_text(row.get("event_time"))),
            ("Country", fmt_value(row.get("country"), 0)),
            ("Ticker", fmt_value(row.get("ticker"), 0)),
            ("Period", fmt_value(row.get("period"), 0)),
            ("Details", fmt_value(row.get("subtitle"), 0)),
        ]
    return _previous_event_fields_custom(self, row)


def _render_detail_with_custom_form(self, event_id):
    _previous_render_detail_custom(self, event_id)
    self._sync_custom_form_from_selected(event_id)


def _event_badges_html_with_custom(self, row, state):
    badge_map = {
        "Economic": "badge-econ",
        "Earnings": "badge-earn",
        "Commodity": "badge-cmdty",
        "Holiday": "badge-holiday",
        "Custom": "badge-custom",
    }
    html = f'<span class="badge {badge_map.get(row["category"], "badge-econ")}">{safe_html(row["category"])}</span>'
    if self.is_event_important(row["event_id"], row=row):
        source = self.important_source(row["event_id"], row=row)
        label = "important" if not source else f"important-{source}"
        html += f'<span class="badge badge-watch">{safe_html(label)}</span>'
    if self._row_flagged(row["event_id"]):
        html += '<span class="badge badge-flag">Flagged</span>'
    return html


def _export_snapshot_with_custom(self, df):
    base_mask = df["category"].astype(str).ne("Custom") if not df.empty else pd.Series([], dtype=bool)
    frames = []
    if not df.empty and base_mask.any():
        frames.append(_previous_export_snapshot_custom(self, df[base_mask]).copy())
    if not df.empty and (~base_mask).any():
        rows = []
        for row in df[~base_mask].to_dict("records"):
            state = self.state.get(row["event_id"], {})
            important = self.is_event_important(row["event_id"], row=row)
            flagged = self._row_flagged(row["event_id"])
            important_source = self.important_source(row["event_id"], row=row) or "-"
            note_text = text_or_blank(state.get("note"), strip=True).replace("\n", " | ")
            if len(note_text) > 280:
                note_text = note_text[:277] + "..."
            status_tags = []
            if important:
                status_tags.append("IMPORTANT")
            if flagged:
                status_tags.append("FLAGGED")
            status_text = " / ".join(status_tags) if status_tags else "-"
            notes_and_tags = " | ".join([status_text] + ([note_text] if note_text else [])) if (status_tags or note_text) else "-"
            rows.append({
                "Event ID": row["event_id"],
                "Date": pd.Timestamp(row["event_date"]).strftime("%Y-%m-%d"),
                "Day": pd.Timestamp(row["event_date"]).strftime("%a"),
                "Time": fmt_time_text(row.get("event_time")),
                "Category": "Custom",
                "Event": text_or_blank(row.get("title"), strip=True) or "Untitled",
                "Country": text_or_blank(row.get("country"), strip=True) or "-",
                "Ticker": text_or_blank(row.get("ticker"), strip=True) or text_or_blank(row.get("country"), strip=True) or "-",
                "Market / Ticker": text_or_blank(row.get("ticker"), strip=True) or "-",
                "Period": text_or_blank(row.get("period"), strip=True) or "-",
                "Period / Detail": text_or_blank(row.get("subtitle"), strip=True) or "User-defined custom event",
                "Survey / Expected": "-",
                "Survey": "-",
                "Prior": "-",
                "Actual": "-",
                "Revision": "-",
                "Scaling": "-",
                "Status": status_text,
                "Important": bool(important),
                "Flagged": bool(flagged),
                "Important Source": important_source,
                "Subtitle": text_or_blank(row.get("subtitle"), strip=True) or "-",
                "Notes": note_text or "-",
                "Notes / Tags": notes_and_tags,
            })
        frames.append(pd.DataFrame(rows))
    if not frames:
        return pd.DataFrame(columns=[
            "Event ID", "Date", "Day", "Time", "Category", "Event", "Country", "Ticker",
            "Market / Ticker", "Period", "Period / Detail", "Survey / Expected", "Survey",
            "Prior", "Actual", "Revision", "Scaling", "Status", "Important", "Flagged",
            "Important Source", "Subtitle", "Notes", "Notes / Tags",
        ])
    return pd.concat(frames, ignore_index=True)


BloombergEventCalendarApp._custom_events_empty_frame = _custom_events_empty_frame
BloombergEventCalendarApp._custom_event_row_from_form = _custom_event_row_from_form
BloombergEventCalendarApp._load_custom_events_from_disk = _load_custom_events_from_disk
BloombergEventCalendarApp._save_custom_events_to_disk = _save_custom_events_to_disk
BloombergEventCalendarApp._custom_events_in_range = _custom_events_in_range
BloombergEventCalendarApp._sync_custom_form_from_selected = _sync_custom_form_from_selected
BloombergEventCalendarApp.clear_custom_event_form = _clear_custom_event_form
BloombergEventCalendarApp.add_custom_event = _add_custom_event
BloombergEventCalendarApp.delete_selected_custom_event = _delete_selected_custom_event
BloombergEventCalendarApp._build_widgets = _enhanced_build_widgets
BloombergEventCalendarApp._wire_events = _enhanced_wire_events
BloombergEventCalendarApp.compose_events_for_range = _compose_events_for_range_with_custom
BloombergEventCalendarApp.reset_filters = _reset_filters_with_custom
BloombergEventCalendarApp._find_event_row = _find_event_row_with_custom
BloombergEventCalendarApp._watch_source_events = _watch_source_events_with_custom
BloombergEventCalendarApp._event_fields = _event_fields_with_custom
BloombergEventCalendarApp.render_detail = _render_detail_with_custom_form
BloombergEventCalendarApp.event_badges_html = _event_badges_html_with_custom
BloombergEventCalendarApp._export_snapshot_dataframe = _export_snapshot_with_custom
BloombergEventCalendarApp._list_header_label = _enhanced_list_header_label
BloombergEventCalendarApp._row_list_label = _enhanced_row_list_label


# =============================================================================
# Final explorer recovery patch: restore a stable Select-based tab explorer.
# This avoids the custom Output/table renderer regressions while preserving
# the shared notes / flags / important-state workflow.
# =============================================================================

EXPLORER_RECOVERY_CSS = """
<style>
.event-app-shell .event-list-select {
    width: 100% !important;
}

.event-app-shell .event-list-select select {
    width: 100% !important;
    min-width: 100% !important;
    min-height: 760px !important;
    height: 760px !important;
    background: #0f1729 !important;
    color: #e5eefb !important;
    border: 1px solid #24324a !important;
    border-radius: 12px !important;
    box-shadow: inset 0 0 1px rgba(93, 168, 255, 0.04) !important;
    font-family: "SFMono-Regular", Menlo, Monaco, Consolas, "Liberation Mono", monospace !important;
    font-size: 10.4px !important;
    line-height: 1.35 !important;
    white-space: pre !important;
    letter-spacing: 0 !important;
    padding: 4px !important;
    overflow-x: hidden !important;
}

.event-app-shell .event-list-select select:focus {
    outline: none !important;
    border-color: #5da8ff !important;
    box-shadow: 0 0 0 1px rgba(93, 168, 255, 0.30) !important;
}

.event-app-shell .event-list-select option {
    background: #0f1729 !important;
    color: #e5eefb !important;
    font-family: inherit !important;
    font-size: inherit !important;
    line-height: 1.35 !important;
}

.event-app-shell .event-list-select option:checked,
.event-app-shell .event-list-select option[selected] {
    background: #16233a linear-gradient(0deg, #16233a, #16233a) !important;
    color: #fff !important;
}

.explorer-recovery-caption {
    color: var(--muted);
    font-size: 12px;
    line-height: 1.45;
    margin-bottom: 8px;
}
</style>
"""


def _recovery_build_widgets(self, months_per_page_default):
    default_start = pd.Timestamp.today().date()
    default_end = (pd.Timestamp.today() + pd.Timedelta(days=45)).date()

    common_desc = {"description_width": "initial"}

    self.start_picker = widgets.DatePicker(description="Start", value=default_start, style=common_desc)
    self.end_picker = widgets.DatePicker(description="End", value=default_end, style=common_desc)
    self.refresh_button = widgets.Button(description="Refresh Data", button_style="primary", icon="refresh")
    self.apply_filters_button = widgets.Button(description="Apply Explorer Filters", icon="filter")

    self.reset_filters_button = widgets.Button(description="Reset Filters", icon="undo")

    self.search_box = widgets.Text(
        description="Search",
        placeholder="Event, ticker, company, country...",
        style=common_desc,
        layout=widgets.Layout(width="340px"),
    )

    self.category_filter = widgets.SelectMultiple(
        description="Explorer Categories",
        options=["Economic", "Earnings", "Commodity", "Holiday"],
        value=("Economic", "Earnings", "Commodity", "Holiday"),
        rows=5,
        style=common_desc,
        layout=widgets.Layout(width="260px"),
    )

    self.only_flagged = widgets.Checkbox(description="Flagged only", value=False, style=common_desc)
    self.only_watch = widgets.Checkbox(description="Important only", value=False, style=common_desc)

    self.months_per_page = widgets.BoundedIntText(
        description="Months / pane",
        value=max(1, int(months_per_page_default)),
        min=1,
        max=12,
        style=common_desc,
        layout=widgets.Layout(width="170px"),
    )

    self.prev_months_button = widgets.Button(description="Previous Pane", icon="arrow-left")
    self.next_months_button = widgets.Button(description="Next Pane", icon="arrow-right")
    self.month_page_label = widgets.HTML()
    self.month_page_note = widgets.HTML(
        '<div class="event-app page-caption">'
        "Monthly view marks holidays, important days, and flagged days. "
        "Click any day to open the full day slate with shared notes, flags, and release detail."
        "</div>"
    )

    self.status_out = widgets.Output()
    self.summary_out = widgets.Output()
    self.monthly_out = widgets.Output()
    self.day_out = widgets.Output(layout=widgets.Layout(height="1040px", overflow="auto"))
    self.detail_html = widgets.HTML()
    self.watchlist_out = widgets.Output()

    list_layout = widgets.Layout(width="100%", height="760px")
    self.all_list = widgets.Select(description="", options=[], rows=30, layout=list_layout)
    self.econ_list = widgets.Select(description="", options=[], rows=30, layout=list_layout)
    self.earnings_list = widgets.Select(description="", options=[], rows=30, layout=list_layout)
    self.comdty_list = widgets.Select(description="", options=[], rows=30, layout=list_layout)
    self.holiday_list = widgets.Select(description="", options=[], rows=30, layout=list_layout)
    self.watch_list = widgets.Select(description="", options=[], rows=30, layout=list_layout)

    for selector in [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list]:
        selector.add_class("event-list-select")

    self.tabs = widgets.Tab(children=[
        self.all_list,
        self.econ_list,
        self.earnings_list,
        self.comdty_list,
        self.holiday_list,
        self.watch_list,
    ])
    for i, title in enumerate(["ALL", "ECONOMIC", "EARNINGS", "COMMODITIES", "HOLIDAYS", "IMPORTANT"]):
        self.tabs.set_title(i, title)
    self.tabs.selected_index = 0
    self.tabs.layout = widgets.Layout(width="100%")

    self.flag_toggle = widgets.ToggleButton(
        description="Flag", icon="flag-o", value=False, layout=widgets.Layout(width="120px"),
    )

    self.watch_toggle = widgets.ToggleButton(
        description="Important", icon="star-o", value=False, layout=widgets.Layout(width="150px"),
    )

    self.note_area = widgets.Textarea(
        description="Notes",
        placeholder="Add custom notes, trading angles, reminders, scenario comments...",
        layout=widgets.Layout(width="100%", height="150px"),
        style=common_desc,
    )

    self.save_note_button = widgets.Button(description="Save Note", button_style="success", icon="save")
    self.clear_note_button = widgets.Button(description="Clear Note", icon="trash")
    self.export_state_button = widgets.Button(description="Sync CSV Now", icon="save")
    self.load_state_button = widgets.Button(description="Reload CSV", icon="refresh")
    self.clear_watchlist_button = widgets.Button(description="Clear Important List", icon="times")

    self.export_scope_dropdown = widgets.Dropdown(
        description="Export Scope",
        options=[
            ("Current visible pane", "pane"),
            ("Full selected range", "range"),
            ("Important / flagged only", "highlights"),
        ],
        value="pane",
        style=common_desc,
        layout=widgets.Layout(width="330px"),
    )

    self.export_document_button = widgets.Button(
        description="Export View CSV", icon="download", button_style="primary",
    )

    self.export_note = widgets.HTML(
        '<div class="event-app page-caption">'
        "Creates a rich CSV snapshot for the active view. Copy the CSV text from the box below into a local file if notebook downloads are blocked, then run the "
        "companion local builder notebook to format the report into a Wall Street-style Excel workbook."
        "</div>"
    )

    self.export_csv_path = widgets.HTML(
        '<div class="event-app page-caption">CSV export preview will appear here after each export.</div>'
    )

    self.export_csv_preview = widgets.Textarea(
        description="CSV Preview",
        placeholder="After export, the CSV text will appear here for copy / paste.",
        layout=widgets.Layout(width="100%", height="220px"),
        style=common_desc,
    )

    self.persistence_note = widgets.HTML(
        f'<div class="event-app page-caption">'
        f'Annotations auto-save to <b>{safe_html(str(self.state_path))}</b>. '
        f'Rule-based key releases from the embedded watch list remain Important even after clearing user-selected Important items.'
        f'</div>'
    )

    controls_title = widgets.HTML('<div class="event-app section-title">Controls</div>')
    controls_grid = widgets.VBox([
        widgets.HBox([self.start_picker, self.end_picker, self.refresh_button]),
        widgets.HBox([self.search_box, self.apply_filters_button, self.reset_filters_button]),
        widgets.HBox([self.category_filter, widgets.VBox([self.only_flagged, self.only_watch])]),
    ])
    self.controls_panel = widgets.VBox([controls_title, controls_grid])
    self.controls_panel.add_class("app-panel")

    # Alarm section retained from the earlier feature request.
    default_alarm_enabled = bool(globals().get("ALARM_ENABLED_BY_DEFAULT", True))
    default_sound_enabled = bool(globals().get("ALARM_SOUND_ENABLED_BY_DEFAULT", True))
    self.alarm_poll_seconds = max(3, int(globals().get("ALARM_CHECK_INTERVAL_SECONDS", 5)))
    self._alarm_monitor_key = "market-event-calendar-flag-alerts"

    self.alarm_enabled_toggle = widgets.ToggleButton(
        description="Alarm Monitor",
        value=default_alarm_enabled,
        icon="bell" if default_alarm_enabled else "bell-slash",
        button_style="warning" if default_alarm_enabled else "",
        layout=widgets.Layout(width="170px"),
    )

    self.alarm_sound_checkbox = widgets.Checkbox(
        description="Sound",
        value=default_sound_enabled,
        layout=widgets.Layout(width="110px"),
    )

    self.alarm_status_html = widgets.HTML(
        '<div class="event-app page-caption">'
        "Alarm monitor watches flagged events that have a concrete date and time. "
        'Warnings fire at T-5:00 and T-1:30. Timeless or TBD events are skipped.'
        "</div>"
    )

    self.alarm_out = widgets.Output(layout=widgets.Layout(width="100%"))

    alarm_title = widgets.HTML('<div class="event-app section-title">Flagged Event Alarm</div>')
    alarm_controls = widgets.HBox([self.alarm_enabled_toggle, self.alarm_sound_checkbox])
    alarm_note = widgets.HTML(
        '<div class="event-app page-caption">'
        "Uses the notebook/browser local clock against the event date and time shown in the app. "
        '"Only flagged events are armed."'
        "</div>"
    )

    self.alarm_section = widgets.VBox([alarm_title, alarm_controls, alarm_note, self.alarm_status_html, self.alarm_out])
    self.controls_panel.children = tuple(list(self.controls_panel.children) + [self.alarm_section])

    summary_title = widgets.HTML('<div class="event-app section-title">Explorer Summary</div>')
    self.summary_panel = widgets.VBox([summary_title, self.summary_out])
    self.summary_panel.add_class("app-panel")

    monthly_title = widgets.HTML('<div class="event-app section-title">Monthly Page</div>')
    month_nav = widgets.HBox([
        self.months_per_page,
        self.prev_months_button,
        self.next_months_button,
        self.month_page_label,
    ])
    self.monthly_panel = widgets.VBox([monthly_title, month_nav, self.month_page_note, self.monthly_out])
    self.monthly_panel.add_class("app-panel")
    self.monthly_panel.layout = widgets.Layout(width="66%")

    day_title = widgets.HTML('<div class="event-app section-title">Selected Day</div>')
    self.day_panel = widgets.VBox([day_title, self.day_out])
    self.day_panel.add_class("app-panel")
    self.day_panel.layout = widgets.Layout(width="34%", min_height="1100px")

    list_caption = widgets.HTML(
        '<div class="event-app explorer-recovery-caption">'
        "<b>Explorer</b> - stable fixed-width text tabs. "
        'Columns read <b>DT/TM | CTRY | EVENT | PER | PRIOR | SURV | ACT | B/M | TAG</b>. '
        'Tags: <b>KEY</b>=important, <b>FLG</b>=flagged, <b>BOTH</b>=both. '
        '"Select any row to sync the shared detail pane and notes."'
        "</div>"
    )

    list_card = widgets.VBox([list_caption, self.tabs], layout=widgets.Layout(width="56%"))
    list_card.add_class("app-panel")

    detail_title = widgets.HTML('<div class="event-app section-title">Event Detail & Notes</div>')
    detail_controls = widgets.HBox([
        self.flag_toggle,
        self.watch_toggle,
        self.save_note_button,
        self.clear_note_button,
    ])
    detail_actions = widgets.HBox([self.export_state_button, self.load_state_button, self.clear_watchlist_button])
    export_controls = widgets.HBox([self.export_scope_dropdown, self.export_document_button])
    detail_card = widgets.VBox([
        detail_title,
        self.detail_html,
        detail_controls,
        self.note_area,
        detail_actions,
        self.persistence_note,
        export_controls,
        self.export_note,
        self.export_csv_path,
        self.export_csv_preview,
        self.watchlist_out,
        self.status_out,
    ], layout=widgets.Layout(width="44%"))
    detail_card.add_class("app-panel")

    header_html = widgets.HTML(
        APP_CSS + TEXT_EXPLORER_EXTRA_CSS + EXPLORER_RECOVERY_CSS +
        """
        <div class="event-app">
            <div class="app-title">Market Event Calendar</div>
            <div class="app-subtitle">
                Dark-themed notebook calendar for economic releases, large-cap earnings, hardcoded commodity expiries, market holidays, and embedded key
                macro events.
            </div>
            <div class="app-footnote">
                This build restores the robust Select-based explorer so the ALL, ECONOMIC, EARNINGS, COMMODITIES, HOLIDAYS, and IMPORTANT tabs remain
                readable and responsive.
            </div>
        </div>
        """
    )

    self.root = widgets.VBox([
        header_html,
        self.controls_panel,
        self.summary_panel,
        widgets.HBox([self.monthly_panel, self.day_panel]),
        widgets.HBox([list_card, detail_card]),
    ], layout=widgets.Layout(width="100%"))
    self.root.add_class("event-app-shell")

    if hasattr(self, "render_alarm_monitor"):
        self.render_alarm_monitor()


def _recovery_wire_events(self):
    self.refresh_button.on_click(self.refresh_data)
    self.apply_filters_button.on_click(self.apply_filters)
    self.reset_filters_button.on_click(self.reset_filters)
    self.save_note_button.on_click(self.save_note)
    self.clear_note_button.on_click(self.clear_note)
    self.export_state_button.on_click(self.export_state)
    self.load_state_button.on_click(self.load_state)
    self.clear_watchlist_button.on_click(self.clear_watchlist)
    self.export_document_button.on_click(self.export_current_view_csv)
    self.prev_months_button.on_click(self.previous_month_pane)
    self.next_months_button.on_click(self.next_month_pane)
    self.months_per_page.observe(self.on_months_per_page_changed, names="value")

    for selector in [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list]:
        selector.observe(self.on_event_selected, names="value")

    self.tabs.observe(self.on_tab_change, names="selected_index")
    self.flag_toggle.observe(self.on_flag_toggled, names="value")
    self.watch_toggle.observe(self.on_watch_toggled, names="value")
    if hasattr(self, "alarm_enabled_toggle"):
        self.alarm_enabled_toggle.observe(self.on_alarm_toggle, names="value")
    if hasattr(self, "alarm_sound_checkbox"):
        self.alarm_sound_checkbox.observe(self.on_alarm_sound_change, names="value")


def _recovery_get_active_list_widget(self):
    idx = self.tabs.selected_index or 0
    return [self.all_list, self.econ_list, self.earnings_list, self.comdty_list, self.holiday_list, self.watch_list][idx]


def _recovery_on_tab_change(self, change):
    widget = self.get_active_list_widget()
    values = [
        value for _, value in widget.options
        if value is not None and not self._is_separator_value(value)
    ]

    if self.current_event_id in values:
        widget.value = self.current_event_id
    elif values:
        widget.value = values[0]
    else:
        if self.current_event_id:
            self.render_detail(self.current_event_id)
        else:
            self.render_detail(None)


def _recovery_on_event_selected(self, change):
    new_value = change["new"]
    if new_value is None or self._is_separator_value(new_value):
        if self.current_event_id:
            self.render_detail(self.current_event_id)
        else:
            self.render_detail(None)
        return
    self.render_detail(new_value)


BloombergEventCalendarApp._build_widgets = _recovery_build_widgets
BloombergEventCalendarApp._wire_events = _recovery_wire_events
BloombergEventCalendarApp.get_active_list_widget = _recovery_get_active_list_widget
BloombergEventCalendarApp.on_tab_change = _recovery_on_tab_change
BloombergEventCalendarApp.on_event_selected = _recovery_on_event_selected


# =============================================================================
# Hedge Fund UX Enhancement Patch
# Bloomberg-terminal-inspired UI for rapid event identification and tracking
# =============================================================================

HF_UX_CSS = """
<style>
/* ── Hedge Fund UX: Enhanced Summary Dashboard ── */
.hf-summary-bar {
    display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 12px;
}
.hf-summary-card {
    flex: 1 1 100px; min-width: 90px; padding: 10px 12px;
    background: linear-gradient(180deg, rgba(20,30,50,0.95), rgba(10,18,32,0.98));
    border: 1px solid rgba(93,168,255,0.15); border-radius: 10px;
    text-align: center; position: relative; overflow: hidden;
}
.hf-summary-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px;
}
.hf-summary-card.hf-econ::before { background: #2ec4b6; }
.hf-summary-card.hf-earn::before { background: #a78bfa; }
.hf-summary-card.hf-cmdty::before { background: #fb7185; }
.hf-summary-card.hf-hol::before { background: #59e0b5; }
.hf-summary-card.hf-flag::before { background: #fb7185; }
.hf-summary-card.hf-key::before { background: #facc15; }
.hf-summary-card.hf-total::before { background: #5da8ff; }
.hf-summary-card.hf-days::before { background: #94a3b8; }
.hf-card-value { font-size: 22px; font-weight: 800; color: #e5eefb; letter-spacing: -0.02em; }
.hf-card-label { font-size: 9px; font-weight: 700; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.08em; margin-top: 2px; }

/* ── Next Key Event Countdown ── */
.hf-countdown-bar {
    display: flex; align-items: center; gap: 14px; padding: 10px 16px;
    background: linear-gradient(90deg, rgba(250,204,21,0.08), rgba(93,168,255,0.06));
    border: 1px solid rgba(250,204,21,0.2); border-radius: 10px; margin-bottom: 10px;
}
.hf-countdown-label { font-size: 10px; font-weight: 700; color: #facc15; text-transform: uppercase; letter-spacing: 0.1em; white-space: nowrap; }
.hf-countdown-event { font-size: 13px; font-weight: 600; color: #e5eefb; flex: 1; }
.hf-countdown-time { font-size: 15px; font-weight: 800; color: #facc15; font-variant-numeric: tabular-nums; white-space: nowrap; }
.hf-countdown-country { font-size: 11px; font-weight: 700; color: #94a3b8; padding: 2px 8px; background: rgba(148,163,184,0.1); border-radius: 6px; }

/* ── Quick Filter Buttons ── */
.hf-quick-filters { display: flex; gap: 6px; margin-bottom: 8px; flex-wrap: wrap; }
.hf-qf-btn button {
    font-size: 10px !important; font-weight: 700 !important; padding: 4px 12px !important;
    border-radius: 20px !important; text-transform: uppercase !important; letter-spacing: 0.05em !important;
    min-height: 28px !important; border: 1px solid rgba(93,168,255,0.3) !important;
    background: rgba(93,168,255,0.08) !important; color: #5da8ff !important;
    transition: all 0.15s ease !important;
}
.hf-qf-btn button:hover { background: rgba(93,168,255,0.2) !important; }
.hf-qf-active button { background: rgba(93,168,255,0.25) !important; border-color: #5da8ff !important; }

/* ── Beat/Miss Indicators ── */
.hf-beat { color: #34d399; font-weight: 700; }
.hf-miss { color: #fb7185; font-weight: 700; }
.hf-inline { color: #94a3b8; }
.hf-beat-dot { display: inline-block; width: 7px; height: 7px; border-radius: 50%; margin-right: 4px; }
.hf-beat-dot.beat { background: #34d399; }
.hf-beat-dot.miss { background: #fb7185; }
.hf-beat-dot.inline { background: #94a3b8; }

/* ── Calendar Heatmap Enhancement ── */
.calendar-day-box.hf-heat-1 { border-left: 3px solid rgba(93,168,255,0.3); }
.calendar-day-box.hf-heat-2 { border-left: 3px solid rgba(93,168,255,0.5); }
.calendar-day-box.hf-heat-3 { border-left: 3px solid rgba(93,168,255,0.7); }
.calendar-day-box.hf-heat-4 { border-left: 3px solid #5da8ff; }
.calendar-day-box.hf-today { outline: 2px solid #facc15; outline-offset: -1px; }
.calendar-day-box.hf-today .day-meta strong { color: #facc15; }

/* ── Enhanced Day Card with Category Border ── */
.hf-day-card-econ { border-left: 3px solid #2ec4b6 !important; }
.hf-day-card-earn { border-left: 3px solid #a78bfa !important; }
.hf-day-card-cmdty { border-left: 3px solid #fb7185 !important; }
.hf-day-card-hol { border-left: 3px solid #59e0b5 !important; }

/* ── Upcoming Events (next 24h highlight) ── */
.hf-upcoming { position: relative; }
.hf-upcoming::after {
    content: 'UPCOMING'; position: absolute; top: 4px; right: 6px;
    font-size: 8px; font-weight: 800; color: #facc15; letter-spacing: 0.1em;
    padding: 1px 5px; background: rgba(250,204,21,0.12); border-radius: 4px;
    border: 1px solid rgba(250,204,21,0.25);
}

/* ── Event Detail: Actual vs Survey comparison ── */
.hf-metric-compare {
    display: flex; align-items: baseline; gap: 6px;
}
.hf-metric-actual { font-size: 16px; font-weight: 800; }
.hf-metric-vs { font-size: 11px; color: #94a3b8; }
.hf-metric-survey { font-size: 13px; color: #94a3b8; }
.hf-delta-chip {
    display: inline-block; padding: 2px 8px; border-radius: 6px;
    font-size: 10px; font-weight: 700; margin-left: 6px;
}
.hf-delta-beat { background: rgba(52,211,153,0.15); color: #34d399; border: 1px solid rgba(52,211,153,0.3); }
.hf-delta-miss { background: rgba(251,113,133,0.15); color: #fb7185; border: 1px solid rgba(251,113,133,0.3); }
.hf-delta-inline { background: rgba(148,163,184,0.1); color: #94a3b8; border: 1px solid rgba(148,163,184,0.2); }

/* ── Priority Score ── */
.hf-priority { font-size: 10px; letter-spacing: -1px; }
.hf-priority-high { color: #facc15; }
.hf-priority-med { color: #5da8ff; }
.hf-priority-low { color: #475569; }

/* ── Compact List Alternating Rows ── */
.event-list-select select option:nth-child(even) { background: rgba(20,30,50,0.3); }

/* ── Today marker in monthly view ── */
.hf-today-marker { font-size: 9px; color: #facc15; font-weight: 800; letter-spacing: 0.05em; }
</style>
"""

APP_CSS = APP_CSS + HF_UX_CSS


def _hf_beat_miss_html(row):
    """Return beat/miss indicator HTML for an economic release."""
    actual = row.get("actual")
    survey = row.get("survey")
    if is_missing(actual) or is_missing(survey):
        return ""
    try:
        a, s = float(actual), float(survey)
    except (ValueError, TypeError):
        return ""
    if math.isclose(a, s, rel_tol=1e-6, abs_tol=1e-9):
        return '<span class="hf-delta-chip hf-delta-inline">IN LINE</span>'
    if a > s:
        delta = a - s
        return f'<span class="hf-delta-chip hf-delta-beat">▲ BEAT +{delta:.2f}</span>'
    delta = s - a
    return f'<span class="hf-delta-chip hf-delta-miss">▼ MISS -{delta:.2f}</span>'


def _hf_priority_score(app, row):
    """Compute a 1-5 priority score for event importance ranking."""
    score = 1
    event_id = row.get("event_id", "")
    category = text_or_blank(row.get("category"), strip=True)
    if app.is_event_important(event_id, row=row):
        score += 2
    if app._row_flagged(event_id):
        score += 1
    if category == "Economic" and has_value(row.get("actual")):
        score += 1
    if category in ("Economic", "Earnings"):
        score = min(score + 1, 5)
    return min(score, 5)


def _hf_priority_stars(score):
    """Return star HTML for priority score."""
    if score >= 4:
        return f'<span class="hf-priority hf-priority-high">{"★" * score}</span>'
    if score >= 2:
        return f'<span class="hf-priority hf-priority-med">{"★" * score}</span>'
    return f'<span class="hf-priority hf-priority-low">{"★" * score}</span>'


def _hf_next_key_event_html(app):
    """Build countdown bar showing time until next key event."""
    now = pd.Timestamp.now()
    today = now.normalize()
    df = app.filtered_events if not app.filtered_events.empty else app.events
    if df.empty:
        return ""
    future = df[pd.to_datetime(df["event_date"], errors="coerce") >= today].copy()
    if future.empty:
        return ""
    # Prefer important events
    important_mask = future["event_id"].map(lambda x: app.is_event_important(x))
    key_future = future[important_mask]
    if key_future.empty:
        key_future = future
    key_future = key_future.sort_values("event_date")
    row = key_future.iloc[0]
    event_date = pd.to_datetime(row["event_date"])
    title = text_or_blank(row.get("title"), strip=True) or "Untitled"
    country = app._display_country_code(row.to_dict()) if hasattr(row, "to_dict") else ""
    time_text = text_or_blank(row.get("event_time"), strip=True) or "TBA"
    category = text_or_blank(row.get("category"), strip=True)
    badge_class = {"Economic": "hf-econ", "Earnings": "hf-earn", "Commodity": "hf-cmdty", "Holiday": "hf-hol"}.get(category, "")

    delta = event_date - now
    days = delta.days
    hours = delta.seconds // 3600
    if days > 0:
        countdown = f"{days}d {hours}h"
    elif hours > 0:
        mins = (delta.seconds % 3600) // 60
        countdown = f"{hours}h {mins}m"
    else:
        mins = max(0, delta.seconds // 60)
        countdown = f"{mins}m" if mins > 0 else "NOW"

    return f'''<div class="event-app hf-countdown-bar">
        <div class="hf-countdown-label">Next Key Event</div>
        <div class="hf-countdown-country">{safe_html(country)}</div>
        <div class="hf-countdown-event">{safe_html(title)} <span style="color:#94a3b8;font-size:11px">({safe_html(time_text)})</span></div>
        <div class="hf-countdown-time">T-{countdown}</div>
    </div>'''


def _hf_render_dashboard(self):
    """Enhanced dashboard with countdown bar, summary cards, and quick filters."""
    df = self.filtered_events
    total = len(df)
    event_days = int(df["event_date"].dt.normalize().nunique()) if not df.empty else 0
    econ = len(df[df["category"] == "Economic"]) if not df.empty else 0
    earnings = len(df[df["category"] == "Earnings"]) if not df.empty else 0
    comdty = len(df[df["category"] == "Commodity"]) if not df.empty else 0
    holidays = len(df[df["category"] == "Holiday"]) if not df.empty else 0
    flagged = int(sum(self._row_flagged(eid) for eid in df.get("event_id", [])))
    important_ids = {eid for eid in df.get("event_id", []) if self.is_event_important(eid)}
    watch = len(important_ids)

    countdown_html = _hf_next_key_event_html(self)

    summary_html = f'''
    <div class="event-app">
    {countdown_html}
    <div class="hf-summary-bar">
        <div class="hf-summary-card hf-total"><div class="hf-card-value">{total}</div><div class="hf-card-label">Events</div></div>
        <div class="hf-summary-card hf-days"><div class="hf-card-value">{event_days}</div><div class="hf-card-label">Days</div></div>
        <div class="hf-summary-card hf-econ"><div class="hf-card-value">{econ}</div><div class="hf-card-label">Economic</div></div>
        <div class="hf-summary-card hf-earn"><div class="hf-card-value">{earnings}</div><div class="hf-card-label">Earnings</div></div>
        <div class="hf-summary-card hf-cmdty"><div class="hf-card-value">{comdty}</div><div class="hf-card-label">Commodity</div></div>
        <div class="hf-summary-card hf-hol"><div class="hf-card-value">{holidays}</div><div class="hf-card-label">Holidays</div></div>
        <div class="hf-summary-card hf-flag"><div class="hf-card-value">{flagged}</div><div class="hf-card-label">Flagged</div></div>
        <div class="hf-summary-card hf-key"><div class="hf-card-value">{watch}</div><div class="hf-card-label">Key Events</div></div>
    </div>
    </div>
    '''
    with self.summary_out:
        clear_output(wait=True)
        display(HTML(summary_html))

    self.render_monthly_page()
    self.render_selected_day_detail()


def _hf_build_day_box(self, day, current_month, start_day, end_day, summary):
    """Enhanced day box with heatmap coloring, today marker, and priority dots."""
    in_range = start_day <= day <= end_day
    in_month = day.month == current_month
    total = summary["counts"].get(day, 0)
    econ = summary["econ"].get(day, 0)
    ern = summary["earnings"].get(day, 0)
    cmd = summary["commodity"].get(day, 0)
    hol = summary["holiday"].get(day, 0)
    key_count = summary["watch"].get(day, 0)
    flag_count = summary["flagged"].get(day, 0)
    hol_codes = summary["holiday_codes"].get(day, [])
    is_today = day == pd.Timestamp.now().date()

    day_label = str(day.day)
    if is_today:
        day_label = f"{day.day}"

    day_button = widgets.Button(
        description=day_label,
        layout=widgets.Layout(width="100%"),
        disabled=not in_range,
        tooltip=f"{day.isoformat()}"
    )
    day_button.add_class("calendar-day-btn")
    if in_range:
        day_button.on_click(lambda _, selected_day=day: self.on_day_clicked(selected_day))

    meta_lines = []
    if is_today:
        meta_lines.append('<span class="hf-today-marker">TODAY</span>')
    if total:
        parts = []
        if key_count:
            parts.append(f'<span style="color:#facc15">★{key_count}</span>')
        if flag_count:
            parts.append(f'<span style="color:#fb7185">⚑{flag_count}</span>')
        parts.append(f"<strong>{total}</strong>")
        meta_lines.append(" ".join(parts))

    bucket_parts = []
    if econ:
        bucket_parts.append(f'<span style="color:#2ec4b6">{econ}E</span>')
    if ern:
        bucket_parts.append(f'<span style="color:#a78bfa">{ern}R</span>')
    if cmd:
        bucket_parts.append(f'<span style="color:#fb7185">{cmd}C</span>')
    if bucket_parts:
        meta_lines.append(" ".join(bucket_parts))

    chip_text = ""
    if hol_codes:
        shown = hol_codes[:3]
        chips = " ".join(f'<span class="holiday-chip">{safe_html(code)}</span>' for code in shown)
        if len(hol_codes) > 3:
            chips += f'<span class="holiday-chip">+{len(hol_codes) - 3}</span>'
        chip_text = chips

    meta_html = widgets.HTML(
        f'<div class="event-app day-meta">{"<br>".join(meta_lines)}{"<div>" + chip_text + "</div>" if chip_text else ""}</div>'
    )

    box = widgets.VBox([day_button, meta_html], layout=widgets.Layout(width="100%"))
    box.add_class("calendar-day-box")
    if not in_month:
        box.add_class("out-month")
    if is_today:
        box.add_class("hf-today")
    if total:
        box.add_class("has-events")
        heat = min(4, max(1, (total + 1) // 3))
        box.add_class(f"hf-heat-{heat}")
    if hol:
        box.add_class("holiday-day")
    if key_count:
        box.add_class("watch-day")
    if flag_count:
        box.add_class("flagged-day")
    if self.selected_day == day:
        box.add_class("selected-day")
    if not in_range:
        box.add_class("no-click")
    return box


def _hf_render_detail(self, event_id):
    """Enhanced detail panel with beat/miss indicator and priority score."""
    self.current_event_id = event_id
    if not event_id:
        self.detail_html.value = (
            '<div class="event-app"><div class="empty-state">'
            'Select an event to view details, notes, flags, and important status.'
            '</div></div>'
        )
        self._suspend_state_events = True
        self.note_area.value = ""
        self.flag_toggle.value = False
        self.watch_toggle.value = False
        self.flag_toggle.icon = "flag-o"
        self.watch_toggle.icon = "star-o"
        self.watch_toggle.disabled = False
        self.watch_toggle.tooltip = ""
        self._suspend_state_events = False
        return

    base_df = self.events if not self.events.empty else self.filtered_events
    row_df = base_df[base_df["event_id"] == event_id]
    if row_df.empty:
        self.detail_html.value = '<div class="event-app"><div class="empty-state">Selected event is no longer in the active date range.</div></div>'
        return

    row = row_df.iloc[0]
    row_dict = row.to_dict()
    if event_id in self.state:
        self._sync_state_metadata(event_id, row=row, touch_timestamp=False)
    state = self.state.get(event_id, {})
    auto_important = self.is_auto_important_event(event_id, row=row_dict)
    important_value = self.is_event_important(event_id, row=row_dict)
    category = text_or_blank(row.get("category"), strip=True)

    self._suspend_state_events = True
    self.flag_toggle.value = self._row_flagged(event_id)
    self.watch_toggle.value = important_value
    self.note_area.value = state.get("note", "")
    self.flag_toggle.icon = "flag" if self.flag_toggle.value else "flag-o"
    self.watch_toggle.icon = "star" if self.watch_toggle.value else "star-o"
    self.watch_toggle.disabled = auto_important
    self.watch_toggle.tooltip = "Always important: matched embedded key-event rule." if auto_important else ""
    self._suspend_state_events = False

    badge_html = self._event_badges_html(row_dict, state)
    priority = _hf_priority_score(self, row_dict)
    priority_html = _hf_priority_stars(priority)

    subtitle = text_or_blank(row.get("subtitle"), strip=True)
    event_time = text_or_blank(row.get("event_time"), strip=True)
    if event_time and event_time not in subtitle:
        subtitle = f"{subtitle} | {event_time}" if subtitle else event_time

    # Beat/miss for economic releases
    beat_miss_html = ""
    metric_compare_html = ""
    if category == "Economic":
        beat_miss_html = _hf_beat_miss_html(row_dict)
        actual = row.get("actual")
        survey = row.get("survey")
        prior = row.get("prior")
        scale = text_or_blank(row.get("scaling_factor"), strip=True)
        if has_value(actual):
            actual_text = fmt_metric(actual, scale)
            survey_text = fmt_metric(survey, scale) if has_value(survey) else "—"
            prior_text = fmt_metric(prior, scale) if has_value(prior) else "—"
            metric_compare_html = f'''
            <div class="hf-metric-compare" style="margin:8px 0">
                <span class="hf-metric-actual">{safe_html(actual_text)}</span>
                <span class="hf-metric-vs">act vs</span>
                <span class="hf-metric-survey">{safe_html(survey_text)} exp</span>
                <span class="hf-metric-vs">|</span>
                <span class="hf-metric-survey">{safe_html(prior_text)} prior</span>
                {beat_miss_html}
            </div>'''

    fields = self._event_fields(row_dict)
    field_html = "".join(
        f'<div class="field"><div class="field-label">{safe_html(label)}</div><div class="field-value">{safe_html(value)}</div></div>'
        for label, value in fields
    )

    # Category border class
    border_class = {"Economic": "hf-day-card-econ", "Earnings": "hf-day-card-earn", "Commodity": "hf-day-card-cmdty", "Holiday": "hf-day-card-hol"}.get(category, "")

    note_preview = state.get("note", "").strip()
    note_html = ""
    if note_preview:
        note_html = f'<div class="note-preview"><b>Saved note</b><br>{safe_html(note_preview)}</div>'
    important_note = ""
    if auto_important:
        important_note = '<div class="note-preview"><b>Rule-based important event</b><br>This release matches the embedded key-event list and is always highlighted as important.</div>'

    self.detail_html.value = f'''
    <div class="event-app {border_class}" style="padding-left:12px">
    <div style="display:flex;justify-content:space-between;align-items:center">
        <div>{badge_html}</div>
        <div>{priority_html}</div>
    </div>
    <div class="detail-title">{safe_html(row.get("title", ""))}</div>
    <div class="detail-subtitle">{safe_html(subtitle)}</div>
    {metric_compare_html}
    <div class="field-grid">{field_html}</div>
    {important_note}
    {note_html}
    </div>
    '''


def _hf_render_selected_day_detail(self):
    """Enhanced day detail with category-colored cards and beat/miss indicators."""
    if self.selected_day is None:
        with self.day_out:
            clear_output(wait=True)
            display(HTML('<div class="event-app day-scroll-box"><div class="empty-state">Click a day to inspect the full state.</div></div>'))
        return

    day_df = self.events[self.events["event_date"].dt.date == self.selected_day].copy()
    if day_df.empty:
        html = (
            '<div class="event-app day-scroll-box">'
            f'<div class="day-panel-header"><div class="day-panel-title">{safe_html(str(self.selected_day))}</div>'
            '<div class="day-panel-subtitle">No events or holiday markers for this date.</div></div>'
            '<div class="empty-state">No events in the selected range for this date.</div></div>'
        )
        with self.day_out:
            clear_output(wait=True)
            display(HTML(html))
        return

    day_df["category_order"] = day_df["category"].map(CATEGORY_ORDER).fillna(99)
    day_df["sort_time"] = day_df["event_time"].fillna("99:99")
    day_df = day_df.sort_values(["category_order", "sort_time", "title"]).drop(columns=["category_order", "sort_time"])

    holiday_codes = day_df.loc[day_df["category"] == "Holiday", "ticker"].dropna().astype(str).unique().tolist()
    header_subtitle = f"{len(day_df)} event{'s' if len(day_df) != 1 else ''}"
    if holiday_codes:
        header_subtitle += f" · Markets closed: {', '.join(sorted(holiday_codes))}"

    cards = []
    for row in day_df.to_dict("records"):
        row_dict = row if isinstance(row, dict) else row
        state = self.state.get(row_dict.get("event_id", ""), {})
        category = text_or_blank(row_dict.get("category"), strip=True)
        border_class = {"Economic": "hf-day-card-econ", "Earnings": "hf-day-card-earn", "Commodity": "hf-day-card-cmdty", "Holiday": "hf-day-card-hol"}.get(category, "")

        fields = self._event_fields(row_dict)
        shown_fields = []
        for label, value in fields:
            if label in ("Event Date",):
                continue
            if value in ("—", None, ""):
                continue
            shown_fields.append(
                f'<div class="day-card-field"><div class="day-card-field-label">{safe_html(label)}</div>'
                f'<div class="day-card-field-value">{safe_html(value)}</div></div>'
            )

        beat_miss = _hf_beat_miss_html(row_dict) if category == "Economic" else ""

        note_text = text_or_blank(state.get("note", ""), strip=True)
        note_html = f'<div class="note-preview"><b>Saved note</b><br>{safe_html(note_text)}</div>' if note_text else ""
        subtitle_bits = []
        event_time = text_or_blank(row_dict.get("event_time"), strip=True)
        subtitle_text = text_or_blank(row_dict.get("subtitle"), strip=True)
        if event_time and event_time != "—":
            subtitle_bits.append(event_time)
        if subtitle_text:
            subtitle_bits.append(subtitle_text)

        title_text = text_or_blank(row_dict.get("title"), strip=True) or "Untitled event"
        important = self.is_event_important(row_dict.get("event_id", ""), row=row_dict)
        flagged = self._row_flagged(row_dict.get("event_id", ""))
        markers = ""
        if important:
            markers += '<span style="color:#facc15;margin-right:4px">★</span>'
        if flagged:
            markers += '<span style="color:#fb7185;margin-right:4px">⚑</span>'

        cards.append(f'''
        <div class="day-card {border_class}">
            <div style="display:flex;justify-content:space-between;align-items:center">
                <div class="day-card-title">{markers}{safe_html(title_text)}</div>
                <div>{beat_miss}</div>
            </div>
            <div class="day-card-subtitle">{safe_html(" | ".join(subtitle_bits))}</div>
            <div class="day-card-grid">{"".join(shown_fields)}</div>
            {note_html}
        </div>''')

    html = f'''
    <div class="event-app day-scroll-box">
        <div class="day-panel-header">
            <div class="day-panel-title">{safe_html(str(self.selected_day))}</div>
            <div class="day-panel-subtitle">{header_subtitle}</div>
        </div>
        {"".join(cards)}
    </div>'''

    with self.day_out:
        clear_output(wait=True)
        display(HTML(html))


def _hf_build_widgets(self, months_per_page_default):
    """Add quick-filter buttons to the widget tree."""
    _previous_hf_build_widgets(self, months_per_page_default)

    # Quick filter buttons
    common_layout = widgets.Layout(width="auto")
    self.hf_btn_today = widgets.Button(description="Today", layout=common_layout)
    self.hf_btn_today.add_class("hf-qf-btn")
    self.hf_btn_week = widgets.Button(description="This Week", layout=common_layout)
    self.hf_btn_week.add_class("hf-qf-btn")
    self.hf_btn_7d = widgets.Button(description="Next 7 Days", layout=common_layout)
    self.hf_btn_7d.add_class("hf-qf-btn")
    self.hf_btn_key_only = widgets.Button(description="Key Events Only", layout=common_layout)
    self.hf_btn_key_only.add_class("hf-qf-btn")
    self.hf_btn_econ_only = widgets.Button(description="Econ Only", layout=common_layout)
    self.hf_btn_econ_only.add_class("hf-qf-btn")
    self.hf_btn_reset = widgets.Button(description="Reset", layout=common_layout, button_style="warning")
    self.hf_btn_reset.add_class("hf-qf-btn")

    self.hf_quick_bar = widgets.HBox(
        [self.hf_btn_today, self.hf_btn_week, self.hf_btn_7d, self.hf_btn_key_only, self.hf_btn_econ_only, self.hf_btn_reset],
    )
    self.hf_quick_bar.add_class("hf-quick-filters")

    # Insert quick filter bar after controls panel
    if hasattr(self, "controls_panel") and self.controls_panel.children:
        self.controls_panel.children = tuple(self.controls_panel.children) + (self.hf_quick_bar,)


def _hf_wire_events(self):
    """Wire quick-filter button events."""
    _previous_hf_wire_events(self)

    def _set_range(start, end):
        self.start_picker.value = start
        self.end_picker.value = end
        self.month_page_start = 0
        self.apply_filters()

    def on_today(_):
        today = pd.Timestamp.today().date()
        _set_range(today, today)

    def on_week(_):
        today = pd.Timestamp.today()
        start = (today - pd.Timedelta(days=today.weekday())).date()
        end = (start + pd.Timedelta(days=6) if isinstance(start, date) else start)
        if isinstance(start, date):
            end = (pd.Timestamp(start) + pd.Timedelta(days=6)).date()
        _set_range(start, end)

    def on_7d(_):
        today = pd.Timestamp.today().date()
        _set_range(today, (pd.Timestamp(today) + pd.Timedelta(days=7)).date())

    def on_key_only(_):
        self.only_watch.value = True
        self.only_flagged.value = False
        self.apply_filters()

    def on_econ_only(_):
        self.category_filter.value = ("Economic",)
        self.apply_filters()

    def on_reset(_):
        self.reset_filters()

    if hasattr(self, "hf_btn_today"):
        self.hf_btn_today.on_click(on_today)
        self.hf_btn_week.on_click(on_week)
        self.hf_btn_7d.on_click(on_7d)
        self.hf_btn_key_only.on_click(on_key_only)
        self.hf_btn_econ_only.on_click(on_econ_only)
        self.hf_btn_reset.on_click(on_reset)


# Apply hedge fund UX monkey-patches
_previous_hf_build_widgets = BloombergEventCalendarApp._build_widgets
_previous_hf_wire_events = BloombergEventCalendarApp._wire_events

BloombergEventCalendarApp._build_widgets = _hf_build_widgets
BloombergEventCalendarApp._wire_events = _hf_wire_events
BloombergEventCalendarApp.render_dashboard = _hf_render_dashboard
BloombergEventCalendarApp._build_day_box = _hf_build_day_box
BloombergEventCalendarApp.render_detail = _hf_render_detail
BloombergEventCalendarApp.render_selected_day_detail = _hf_render_selected_day_detail


# =============================================================================
# Launch the app
# =============================================================================

app = BloombergEventCalendarApp(
    eco_universe=CUSTOM_ECO_UNIVERSE,
    comdty_tickers=CUSTOM_COMDTY_TICKERS,
    top_n_market_cap=TOP_N_MARKET_CAP,
    state_path=ANNOTATIONS_FILE,
    legacy_state_path=LEGACY_STATE_FILE,
    export_dir=EXPORT_DIR,
    market_holidays_tsv=CUSTOM_MARKET_HOLIDAYS_TSV,
    market_code_map=CUSTOM_MARKET_CODE_MAP,
    always_important_tsv=CUSTOM_ALWAYS_IMPORTANT_TSV,
    important_country_aliases=CUSTOM_IMPORTANT_COUNTRY_ALIASES,
    months_per_page_default=MONTHS_PER_PANE_DEFAULT,
)

app.eco_calendar_types = CUSTOM_ECO_CALENDAR_TYPES
app.alarm_poll_seconds = max(3, int(ALARM_CHECK_INTERVAL_SECONDS))

if hasattr(app, "render_alarm_monitor"):
    app.render_alarm_monitor()

app.display()

required_frames = ["df_ECO", "df_EQ_Earnings"]
missing_frames = [name for name in required_frames if name not in globals()]

if not missing_frames:
    app.load_from_dataframes(df_ECO, df_EQ_Earnings, None)
else:
    app.log(
        "Existing dataframes were not found in this session: "
        + ", ".join(missing_frames)
        + ". Set the date range and click 'Refresh Data' to pull the data live. "
        + "Commodity events will still use the embedded hardcoded options/futures expiry schedule."
    )
